"""
desktop_app.py — настольное Windows-приложение поверх торгового движка (main.py
и всё, что он использует). Это ТОЧКА ВХОДА для .exe/установщика — см. BUILD.md.

Всё — счёт, брокер, символы, сделки, лог, график equity, настройки, новости,
чат с Claude — открывается ВНУТРИ этого окна (вкладки), браузер для управления
с компьютера не нужен. Веб-дашборд (Flask) продолжает работать в фоне только
ради доступа С ТЕЛЕФОНА по Wi-Fi — на самом компьютере им пользоваться не
обязательно.

Два режима интерфейса (переключаются на вкладке "Обзор", см. UI_MODE в config.py):
  - Простой     — Обзор / Символы / Сделки / Лог / Настройка / Как пользоваться.
  - Продвинутый — плюс Брокер / Equity / Новости / Chat AI.
    Вкладка "Настройка" ВСЕГДА видна (в обоих режимах — чтобы точно не потерялась):
    быстрые переключатели (профиль/режим/пауза/звук) + ПОЛНЫЙ набор
    input-параметров торговой логики (как у MQL5-советника): индикаторы,
    риск/лот, TP/SL, BE/трейлинг/Profit Lock/частичное закрытие, защитные
    фильтры, режим рынка, контекст, AI, новости, автообучение — редактируется
    прямо в интерфейсе, без открытия config.py руками.
    Вкладка "Как пользоваться" — подробная инструкция по всем вкладкам, тоже
    всегда видна.

ВАЖНО про архитектуру:
  1) Собранный PyInstaller-ом .exe уже содержит весь Python-рантайм внутри —
     отдельного "python.exe" на машине пользователя может не быть. Поэтому
     торговый цикл (main.main()) запускается в ФОНОВОМ ПОТОКЕ этого же
     процесса, а не отдельным подпроцессом. Кнопка "Стоп" останавливает его
     через threading.Event (main.py проверяет его каждую итерацию).
  2) config.py при сборке НЕ встраивается внутрь .exe (см. build_exe.bat:
     --exclude-module config) — он остаётся отдельным редактируемым файлом
     рядом с программой. Поэтому первым делом (до импорта любых внутренних
     модулей проекта) мы добавляем папку с .exe в sys.path.
  3) Веб-дашборд (Flask) поднимается ОДИН РАЗ при старте приложения — иначе
     повторный запуск бота попытался бы занять уже занятый порт 5000.
  4) Вкладки читают то же самое состояние (dashboard_state.get_snapshot(),
     control.py), что и веб-дашборд — никакого HTTP не нужно, всё в одном процессе.
  5) MT5 трогает ТОЛЬКО поток бота (main.main()) и, отдельно, кнопка "Проверить
     подключение" на вкладке "Брокер" (но она заблокирована, пока бот работает,
     чтобы не дёргать API из двух потоков одновременно).
"""

import sys
import os

if getattr(sys, "frozen", False):
    _app_dir = os.path.dirname(sys.executable)
else:
    _app_dir = os.path.dirname(os.path.abspath(__file__))
if _app_dir not in sys.path:
    sys.path.insert(0, _app_dir)
os.chdir(_app_dir)

import base64
import csv
import importlib
import logging
import atexit
import multiprocessing
import re
import threading
import time
import webbrowser
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog, ttk

def _ensure_config(app_dir: str) -> str:
    """Настройки рядом с программой обязаны существовать. Возвращает, что сделано.

    ПОЧЕМУ ЭТО ЗДЕСЬ. config.py намеренно НЕ встраивается внутрь .exe — он
    остаётся отдельным редактируемым файлом рядом с программой. И его может
    не оказаться: первая установка, ручное удаление, неудачное обновление,
    распаковка не в ту папку.

    Раньше в этом случае программа падала прямо на строке `import config` —
    ещё до того, как появлялось хоть одно окно. А собранная БЕЗ КОНСОЛИ
    программа при падении показывает системное окно с ошибкой, которого на
    рабочем столе можно и не заметить, и висит, пока в нём не нажмут «ОК».
    Снаружи это выглядит ровно как «нет отклика от программы, виснет».

    Теперь недостающие настройки создаются из эталона — как и должно быть
    при первом запуске. Молча заменять СУЩЕСТВУЮЩИЙ config.py нельзя: там
    личные ключи и пароли владельца."""
    настройки = os.path.join(app_dir, "config.py")
    if os.path.exists(настройки):
        return ""
    эталон = os.path.join(app_dir, "config.py.example")
    if not os.path.exists(эталон):
        return "нет ни config.py, ни config.py.example"
    try:
        with open(эталон, "r", encoding="utf-8") as f:
            текст = f.read()
        with open(настройки, "w", encoding="utf-8") as f:
            f.write(текст)
        return "создан config.py из config.py.example"
    except OSError as e:
        return f"не удалось создать config.py: {e}"


_config_note = _ensure_config(_app_dir)

try:
    import config as cfg
except Exception as e:  # noqa: BLE001
    # Падать здесь — значит зависнуть невидимым окном ошибки (см. выше).
    # Лучше сказать это словами и уйти с ненулевым кодом: так причину видно
    # и человеку, и проверке при сборке.
    print(f"НАСТРОЙКИ НЕ ЗАГРУЖЕНЫ: {type(e).__name__}: {e}")
    print(f"Папка программы: {_app_dir}")
    if _config_note:
        print(f"Попытка починить: {_config_note}")
    print("Положите файл config.py рядом с программой и запустите ещё раз.")
    sys.exit(2)

LOG_FILE = os.path.join(_app_dir, "scalper.log")
logging.basicConfig(
    level=getattr(logging, cfg.LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler()],
)
log = logging.getLogger("desktop_app")

import main as bot_engine
import dashboard_state as ds
import news_calendar
import news_providers
import trading_schedule as tsched
import mt5_connector as mt5c
import mt5_install
import param_help
import config_migrate
import news_autostart
import single_instance
import runtime_events
import settings_backup
import ui_theme
import ui_layout
import version as app_version
import cloud_journal
import bridge_host
import diagnostics
import updater
import telegram_signals as tgs
import telegram_reader as tgr
import secure_store
import strategies as strategies_mod
import safe_files
from control import control

try:
    import pystray
    from PIL import Image, ImageDraw
    TRAY_AVAILABLE = True
except ImportError:
    TRAY_AVAILABLE = False
    log.info("pystray/Pillow не установлены — работаю без значка в трее.")

APP_TITLE = "AI Scalper Pro"

PROFILE_OPTIONS = [
    ("Консервативный", "conservative"),
    ("Сбалансированный", "balanced"),
    ("Агрессивный", "aggressive"),
    ("Истеричка (YOLO)", "hysteric"),
]
MODE_OPTIONS = [
    ("Скальпинг", "scalping"),
    ("Новости", "news_trading"),
    ("Оба", "both"),
]
ADVANCED_TAB_NAMES = ["Брокер", "Equity", "Новости", "Сигналы", "Чат"]
# "Источники" видна всегда: это единственное место, где источники данных
# включаются и выключаются — прятать его в продвинутый режим нельзя.
# График календаря на вкладке "Новости": на сколько часов вперёд смотрим и
# какой высоты полоса. Больше 24 ч смысла нет — засечки сливаются.
NEWS_CHART_HOURS = 12
NEWS_CHART_HEIGHT = 96
NEWS_LABEL_MIN_GAP_PX = 26   # ближе этого подписи валют наезжают друг на друга
# "Настройка" (все настройки + input-параметры) и "Как пользоваться" видны ВСЕГДА,
# в обоих режимах интерфейса — чтобы их точно не потерять в простом режиме.

# ---- Вкладка "Параметры" (продвинутый режим) -------------------------------
# Полный список "входных параметров" торговой логики — как input-параметры
# MQL5-советника: любой из них можно выставить вручную, без правки config.py
# руками. (key, тип, группа, подпись, варианты_для_choice).
# тип: "int" | "float" | "bool" | "choice" | "secret" (шифруется при сохранении)

# Что показывается в поле секрета, когда ключ уже сохранён: сам ключ в
# интерфейс не выводится, чтобы его нельзя было подсмотреть через плечо.
SECRET_PLACEHOLDER = "••••••••  (ключ сохранён, впишите новый чтобы заменить)"
ADVANCED_PARAMS = [
    # Реальная торговля выключена в образце настроек по умолчанию, поэтому
    # включатель обязан быть ВИДИМЫМ. Иначе свежепоставленная программа молча
    # не торговала бы, а единственным способом это исправить было бы
    # редактирование config.py в блокноте — то есть ровно то, чего владелец
    # делать не должен.
    ("LIVE_TRADING", "bool", "Общее",
     "Реальная торговля (выкл. = только запись в журнал, ордера не отправляются)", None),
    ("POLL_SECONDS", "int", "Общее", "Частота опроса рынка, сек", None),
    ("TIMEFRAME", "choice", "Общее", "Рабочий таймфрейм", ["M1", "M5", "M15", "M30", "H1", "H4", "D1"]),
    ("TREND_TIMEFRAME", "choice", "Общее", "Старший ТФ тренда (должен быть старше рабочего)",
     ["M1", "M5", "M15", "M30", "H1", "H4", "D1"]),
    ("MAGIC_NUMBER", "int", "Общее", "Magic number (НЕ менять во время работы бота!)", None),
    ("LOG_LEVEL", "choice", "Общее", "Уровень логирования", ["DEBUG", "INFO", "WARNING", "ERROR"]),

    ("EMA_FAST_PERIOD", "int", "Индикаторы", "EMA быстрая — период", None),
    ("EMA_SLOW_PERIOD", "int", "Индикаторы", "EMA медленная — период", None),
    ("EMA_TREND_PERIOD", "int", "Индикаторы", "EMA тренда (старший ТФ) — период", None),
    ("ADX_PERIOD", "int", "Индикаторы", "ADX — период", None),
    ("RSI_PERIOD", "int", "Индикаторы", "RSI — период", None),
    ("ATR_PERIOD", "int", "Индикаторы", "ATR — период", None),
    ("ATR_AVG_PERIOD", "int", "Индикаторы", "ATR — период среднего", None),
    ("ADX_MIN_LEVEL", "int", "Индикаторы", "ADX — минимальный уровень тренда", None),
    ("RSI_OVERBOUGHT", "int", "Индикаторы", "RSI — уровень перекупленности", None),
    ("RSI_OVERSOLD", "int", "Индикаторы", "RSI — уровень перепроданности", None),

    ("PULLBACK_TOLERANCE_POINTS", "int", "Price Action / откат", "Допуск отката, пункты", None),
    ("BODY_PERCENT_MIN", "int", "Price Action / откат", "Мин. % тела свечи", None),
    ("MAX_WICK_PERCENT", "int", "Price Action / откат", "Макс. % тени свечи", None),

    ("USE_VOLUME_FILTER", "bool", "Объём", "Учитывать тиковый объём", None),
    ("VOLUME_AVG_PERIOD", "int", "Объём", "Период среднего объёма", None),

    ("AUTO_ADAPT_TO_SYMBOL", "bool", "Автонастройка под инструмент",
     "Авто-масштабирование порогов под ATR инструмента", None),

    ("MIN_BARS_BETWEEN_REVERSAL", "int", "Защитные проверки",
     "Анти-дребезг: баров ожидания перед разворотом (0 = без ожидания)", None),
    ("USE_SPREAD_FILTER", "bool", "Защитные проверки", "Фильтр по спреду", None),
    ("MAX_SPREAD_POINTS", "int", "Защитные проверки", "Макс. спред, пункты", None),
    ("USE_VOLATILITY_SPIKE_GUARD", "bool", "Защитные проверки", "Защита от скачков волатильности", None),
    ("VOLATILITY_SPIKE_MULTIPLIER", "float", "Защитные проверки", "Множитель скачка волатильности", None),
    ("USE_ROLLOVER_GUARD", "bool", "Защитные проверки",
     "Пауза на роллoвер, полночь брокера (выключена)", None),
    ("ROLLOVER_HOUR_SERVER", "int", "Защитные проверки", "Час роллoвера (серверное время)", None),
    ("ROLLOVER_GUARD_MINUTES", "int", "Защитные проверки", "Длительность паузы роллoвера, мин", None),

    ("USE_AI_SIGNAL", "bool", "AI-сигнал", "Использовать внешний AI-сигнал", None),
    ("AI_PROVIDER", "choice", "AI-сигнал", "Провайдер AI", ["claude", "openai"]),
    ("AI_SIGNAL_WEIGHT", "int", "AI-сигнал", "Вес AI-сигнала в score", None),
    ("AI_SIGNAL_CACHE_SECONDS", "int", "AI-сигнал", "Кэш ответа AI, сек", None),
    ("AI_SIGNAL_REQUIRE_DIRECTION", "bool", "AI-сигнал", "Требовать совпадения направления с AI", None),
    # Ключи вводятся прямо здесь и сохраняются ЗАШИФРОВАННЫМИ (тип "secret"):
    # раньше их можно было вписать только руками в config.py
    ("ANTHROPIC_API_KEY", "secret", "AI-сигнал", "Ключ Claude (console.anthropic.com)", None),
    ("OPENAI_API_KEY", "secret", "AI-сигнал", "Ключ OpenAI (нужен только для провайдера openai)", None),

    ("USE_CUSTOM_STRATEGY", "bool", "Собственная стратегия",
     "Использовать собственную стратегию программы (custom_strategy.py)", None),
    ("CUSTOM_STRATEGY_WEIGHT", "int", "Собственная стратегия", "Вес собственной стратегии в score", None),

    ("USE_MULTI_INDICATOR", "bool", "Доп. индикаторы",
     "Подмешивать MACD/Bollinger/Stochastic в score (multi_indicator.py)", None),
    ("MULTI_INDICATOR_WEIGHT", "int", "Доп. индикаторы", "Вес группы индикаторов в score", None),
    ("MACD_FAST_PERIOD", "int", "Доп. индикаторы", "MACD: период быстрой EMA", None),
    ("MACD_SLOW_PERIOD", "int", "Доп. индикаторы", "MACD: период медленной EMA", None),
    ("MACD_SIGNAL_PERIOD", "int", "Доп. индикаторы", "MACD: период сигнальной линии", None),
    ("BB_PERIOD", "int", "Доп. индикаторы", "Bollinger Bands: период SMA", None),
    ("BB_STD_MULT", "float", "Доп. индикаторы", "Bollinger Bands: множитель стандартного отклонения", None),
    ("STOCH_K_PERIOD", "int", "Доп. индикаторы", "Stochastic: период %K", None),
    ("STOCH_D_PERIOD", "int", "Доп. индикаторы", "Stochastic: период %D", None),

    ("USE_MARKET_REGIME_FILTER", "bool", "Режим рынка", "Учитывать режим рынка (тренд/флэт)", None),
    ("USE_ADAPTIVE_SCORE_WEIGHTS", "bool", "Режим рынка",
     "Умнее веса score: усиливать Тренд/Откат в тренде, RSI во флэте", None),
    ("REGIME_ER_PERIOD", "int", "Режим рынка", "Efficiency Ratio — период", None),
    ("REGIME_ADX_TREND_LEVEL", "int", "Режим рынка", "ADX — порог тренда", None),
    ("REGIME_ADX_RANGE_LEVEL", "int", "Режим рынка", "ADX — порог флэта", None),
    ("REGIME_ER_TREND_LEVEL", "float", "Режим рынка", "ER — порог тренда", None),
    ("REGIME_ER_RANGE_LEVEL", "float", "Режим рынка", "ER — порог флэта", None),
    ("REGIME_CONFIRM_BARS", "int", "Режим рынка", "Баров подтверждения смены режима", None),
    ("REGIME_RANGE_PENALTY", "int", "Режим рынка", "Штраф score во флэте", None),
    ("REGIME_TREND_BONUS", "int", "Режим рынка", "Бонус score в тренде", None),

    ("USE_PA_HARD_GATE", "bool", "Анти-'зеркало' фильтры",
     "Обязательное подтверждение свечи (иначе score=0) — против входа на развороте", None),
    ("BLOCK_ENTRY_IN_RANGE", "bool", "Анти-'зеркало' фильтры",
     "Полный блок входа во флэте (не просто штраф)", None),
    ("USE_EXHAUSTION_FILTER", "bool", "Анти-'зеркало' фильтры",
     "Блокировать вход на уже растянутой (перегретой) свече", None),
    ("EXHAUSTION_RANGE_ATR_RATIO", "float", "Анти-'зеркало' фильтры",
     "Порог: диапазон свечи > этого × средний ATR -> блок входа", None),

    ("USE_MARKET_CONTEXT", "bool", "Контекст рынка", "Учитывать коррелирующие инструменты", None),
    ("CONTEXT_EMA_PERIOD", "int", "Контекст рынка", "EMA контекста — период", None),
    ("CONTEXT_SCORE_WEIGHT", "int", "Контекст рынка", "Вес контекста в score", None),

    ("USE_SCORE_FILTER", "bool", "Score-фильтр", "Входить только по score-порогу профиля", None),

    ("USE_MAX_PROFIT_RIDE", "bool", "Стопы / TP",
     "Без фикс. TP — тянуть максимум прибыли трейлингом/Profit Lock (пока не выбьет стоп)", None),
    ("RISK_REWARD_RATIO", "float", "Стопы / TP", "Risk/Reward (если TP не в деньгах; не используется при макс. профите)", None),
    ("MIN_RISK_REWARD_RATIO", "float", "Стопы / TP",
     "КРИТИЧНО: мин. Risk/Reward — TP никогда не меньше SL x это число, даже если денежная цель профиля скромнее", None),
    ("TP_MIN_POINTS", "int", "Стопы / TP", "Мин. TP, пункты (не используется при макс. профите)", None),
    ("TP_MAX_POINTS", "int", "Стопы / TP", "Макс. TP, пункты (не используется при макс. профите)", None),

    ("USE_BREAK_EVEN", "bool", "Break Even", "Переносить в безубыток", None),
    ("BREAK_EVEN_ATR_MULTIPLIER", "float", "Break Even", "Триггер BE, множитель ATR", None),
    ("BREAK_EVEN_OFFSET_POINTS", "int", "Break Even", "Отступ BE, пункты", None),

    ("USE_TRAILING_STOP", "bool", "Трейлинг-стоп", "Трейлинг-стоп по ATR", None),
    ("TRAILING_ATR_MULTIPLIER", "float", "Трейлинг-стоп", "Множитель ATR трейлинга", None),
    ("TRAILING_MIN_POINTS", "int", "Трейлинг-стоп", "Мин. дистанция трейлинга, пункты", None),
    ("TRAILING_STEP_MIN_POINTS", "int", "Трейлинг-стоп", "Мин. шаг трейлинга, пункты", None),

    ("USE_PROFIT_LOCK_TRAILING", "bool", "Profit Lock", "Фиксация части пиковой прибыли", None),
    ("PROFIT_LOCK_START_POINTS", "int", "Profit Lock", "Старт Profit Lock, пункты (ATR-порог)", None),
    ("PROFIT_LOCK_START_R_FRACTION", "float", "Profit Lock",
     "КРИТИЧНО: лок не стартует, пока прибыль не достигнет этой доли риска сделки (1.0 = не раньше 1R)", None),
    ("PROFIT_LOCK_PERCENT", "int", "Profit Lock",
     "% пиковой прибыли фиксировать (запасной вариант, если ступенчатая фиксация ниже выключена)", None),
    ("USE_TIERED_PROFIT_LOCK", "bool", "Profit Lock",
     "Ступенчатая фиксация: чем выше пик прибыли, тем больший % запирается (уровни — в config.py)", None),
    ("POSITION_MONITOR_SECONDS", "float", "Profit Lock",
     "Как часто (сек) проверять УЖЕ открытые позиции между полными проходами по всем парам", None),

    ("USE_DAILY_LOSS_LIMIT", "bool", "Просадка / серии убытков",
     "Дневной порог убытка (выключен: бот работает всё торговое время)", None),
    ("USE_MAX_DRAWDOWN_LIMIT", "bool", "Просадка / серии убытков",
     "Лимит общей просадки (выключен: торговля не останавливается)", None),
    ("MAX_CONSECUTIVE_LOSSES", "int", "Просадка / серии убытков",
     "Серия убытков подряд, после которой объём снижается", None),
    ("PAUSE_MINUTES_AFTER_LOSS_STREAK", "int", "Просадка / серии убытков",
     "Пауза после серии убытков, минут (0 = без паузы, торговля не прерывается)", None),
    ("USE_LOSS_STREAK_RISK_SCALING", "bool", "Просадка / серии убытков", "Снижать риск при серии убытков", None),
    ("MIN_LOSS_STREAK_RISK_MULTIPLIER", "float", "Просадка / серии убытков",
     "Мин. множитель риска при серии убытков", None),

    ("MAX_SPREAD_COST_PERCENT_OF_TP", "float", "Издержки", "Макс. % TP, съедаемый спредом", None),
    ("ORDER_RETRY_ATTEMPTS", "int", "Издержки", "Повторов отправки ордера", None),

    ("NEWS_BREAKOUT_WINDOW_MIN", "int", "Новости (пороги)", "Окно реакции на новость, мин", None),
    ("NEWS_BREAKOUT_MIN_BODY_PCT", "int", "Новости (пороги)", "Мин. % тела свечи для пробоя", None),
    ("NEWS_VOLATILITY_SL_BOOST", "float", "Новости (пороги)", "Множитель SL на новостях", None),

    ("USE_AUTO_LEARNING", "bool", "Автообучение", "Адаптировать вес AI / порог по винрейту", None),
    ("AUTO_LEARNING_WINDOW", "int", "Автообучение", "Окно последних сделок по символу", None),
    ("UI_THEME", "str", "Оформление", "Тема окна: light (светлая) или dark (тёмная)", None),
    ("USE_SYMBOL_AUTO_OFF", "bool", "Автообучение", "Отключать инструмент, который стабильно в минусе", None),
    ("SYMBOL_AUTO_OFF_MIN_TRADES", "int", "Автообучение", "Сколько сделок нужно, прежде чем судить об инструменте", None),
    ("SYMBOL_AUTO_OFF_LOSS_PERCENT", "float", "Автообучение", "Потерял больше % счёта за окно — отключить инструмент", None),
    ("AUTO_LEARNING_MIN_TRADES", "int", "Автообучение", "Мин. сделок для начала адаптации", None),
    ("AI_WEIGHT_MULT_MIN", "float", "Автообучение", "Мин. множитель веса AI", None),
    ("AI_WEIGHT_MULT_MAX", "float", "Автообучение", "Макс. множитель веса AI", None),
    ("SCORE_THRESHOLD_ADJUST_MIN", "int", "Автообучение", "Мин. коррекция порога score", None),
    ("SCORE_THRESHOLD_ADJUST_MAX", "int", "Автообучение", "Макс. коррекция порога score", None),

    ("USE_CONFIG_HOT_RELOAD", "bool", "Автообновление", "Подхватывать правки config.py на лету", None),
    ("CONFIG_RELOAD_CHECK_SECONDS", "int", "Автообновление", "Период проверки config.py, сек", None),
    ("USE_AUTO_RECONNECT", "bool", "Автообновление", "Автопереподключение к MT5", None),
    ("RECONNECT_AFTER_FAILURES", "int", "Автообновление", "Неудач подряд до переподключения", None),
    ("HISTORY_SYNC_DAYS", "int", "Автообновление", "Синхр. с MetaTrader: глубина истории, дней", None),
    ("HISTORY_SYNC_SECONDS", "int", "Автообновление", "Синхр. с MetaTrader: период обновления, сек", None),

    ("USE_TRADING_HOURS", "bool", "Часы торговли", "Ограничить торговлю часами (иначе круглосуточно)", None),
    ("TRADING_START_HOUR", "int", "Часы торговли", "Час начала торговли, 0-23 (время сервера)", None),
    ("TRADING_END_HOUR", "int", "Часы торговли", "Час окончания торговли, 0-23 (время сервера)", None),

    ("USE_RISK_BASED_LOT", "bool", "Лот", "Считать лот по риску в % от эквити (иначе фиксированный)", None),
    ("LOT_FALLBACK", "float", "Лот", "Фиксированный / запасной лот", None),

    ("NEWS_HARD_BLOCK_WINDOW_MIN", "int", "Новости (пороги)",
     "Пауза рядом с важной новостью, мин (0 = без паузы)", None),
    ("NEWS_SOFT_PENALTY_WINDOW_MIN", "int", "Новости (пороги)",
     "Окно штрафа score рядом с новостью, мин (это НЕ пауза)", None),
    ("NEWS_SOFT_PENALTY_POINTS", "float", "Новости (пороги)", "Штраф score рядом с MODERATE-новостью, баллы", None),

    ("AI_SIGNAL_TIMEOUT_MS", "int", "AI-сигнал", "Таймаут запроса к AI, мс", None),

    ("USE_PARTIAL_CLOSE", "bool", "Частичное закрытие", "Частично закрывать позицию при достижении профита", None),
    ("PARTIAL_CLOSE_TRIGGER_POINTS", "int", "Частичное закрытие", "Профит для частичного закрытия, пункты", None),
    ("PARTIAL_CLOSE_PERCENT", "int", "Частичное закрытие", "% объёма закрывать частично", None),

    ("USE_TP_TIGHTEN", "bool", "Фиксация прибыли",
     "Подтягивать тейк-профит ближе к цене (никогда не отодвигать дальше)", None),
    ("TP_TIGHTEN_START_ATR", "float", "Фиксация прибыли", "Стартовая цель прибыли = ATR × это", None),
    ("TP_TIGHTEN_SHRINK_PER_MINUTE", "float", "Фиксация прибыли",
     "На сколько ужимать цель за минуту жизни сделки (0.10 = на 10%)", None),
    ("TP_TIGHTEN_MIN_FRACTION", "float", "Фиксация прибыли", "Ниже какой доли стартовой цели не опускаться", None),
    ("TP_TIGHTEN_MIN_PROFIT_POINTS", "int", "Фиксация прибыли", "Мин. прибыль от входа, пункты (TP всегда в плюсе)", None),
    ("TP_TIGHTEN_STEP_POINTS", "int", "Фиксация прибыли", "Мин. шаг переноса TP, пункты", None),
    ("TP_TIGHTEN_MIN_R", "float", "Фиксация прибыли",
     "КРИТИЧНО: цель не может стать меньше риска сделки x это число (1.0 = не меньше стопа)", None),

    ("USE_BREAK_EVEN_RESCUE", "bool", "Фиксация прибыли",
     "Спасать в безубыток сделку, которая просела и вернулась к нулю (стоп-лосс НЕ трогается)", None),
    ("BE_RESCUE_MIN_DRAWDOWN_POINTS", "int", "Фиксация прибыли", "Насколько глубоко надо было просесть, пункты", None),
    ("BE_RESCUE_AFTER_MINUTES", "int", "Фиксация прибыли", "И сколько провисеть до включения спасения, мин", None),
    ("BE_RESCUE_EXIT_POINTS", "int", "Фиксация прибыли", "Что считать 'нулём', пункты (запас на спред)", None),

    ("USE_TP_LEARNING", "bool", "Автообучение",
     "Учить цель прибыли по пикам прошлых сделок (медиана)", None),
    ("TP_LEARN_FRACTION", "float", "Автообучение", "Доля от медианы пика для TP (0.7 = выходить до разворота)", None),
    ("TP_LEARN_MIN_POINTS", "int", "Автообучение", "Мин. выученная цель, пункты", None),
    ("TP_LEARN_MAX_POINTS", "int", "Автообучение", "Макс. выученная цель, пункты", None),
    ("USE_LEARNING_PERSISTENCE", "bool", "Автообучение",
     "Сохранять выученное между запусками (без этого обучение обнуляется при каждом старте)", None),
    ("LEARNING_STATE_PATH", "str", "Автообучение", "Файл со статистикой обучения", None),

    ("LOG_CSV_PATH", "str", "Общее", "Имя CSV-файла журнала сделок", None),
]

RISK_PROFILE_FIELD_DEFS = [
    ("name", "str", "Название профиля"),
    ("risk_percent", "float", "Риск на сделку, % от эквити"),
    ("atr_sl_multiplier", "float", "Множитель ATR для SL"),
    ("use_money_tp", "bool", "TP в деньгах (иначе Risk/Reward)"),
    ("target_profit_money", "float", "Целевой TP, деньги"),
    ("min_score_to_trade", "int", "Порог score для входа"),
    ("max_open_positions", "int", "Макс. одновременных сделок"),
    ("max_trades_per_day", "int", "Макс. сделок в день"),
    ("daily_loss_limit_pct", "float", "Дневной порог убытка, % (0 = без порога)"),
    ("max_drawdown_pct", "float", "Лимит просадки, %"),
    ("max_total_risk_pct", "float", "Лимит совокупного риска, %"),
    ("ignore_soft_filters", "bool", "Игнорировать мягкие фильтры"),
    ("hedge_both_directions", "bool", "Хедж: при сигнале открывать сразу BUY и SELL (обычный SL на каждой ноге)"),
]


_MISSING = object()


def param_current_value(key: str):
    """Значение параметра для показа в поле вкладки «Настройка».

    Если параметра в вашем config.py ещё нет (он появился в новой версии, а
    файл настроек остался старым), берём значение по умолчанию из
    config.py.example. Раньше в этом случае подставлялась пустая строка: поле
    выглядело пустым, а «Сохранить» отвечал «Некорректные значения в полях:
    PROFIT_LOCK_START_R_FRACTION, ...» — пустоту нельзя превратить в число.
    config_migrate.sync() дописывает такие настройки в файл при запуске, а это
    — вторая линия обороны на случай, если файл не удалось изменить (нет прав,
    файл только для чтения)."""
    value = getattr(cfg, key, _MISSING)
    if value is not _MISSING:
        return value
    default = param_help.default_of(key)
    return "" if default is None else default


def _write_config_value(key: str, value_literal: str):
    """Переписывает ОДНУ строку присваивания в config.py, не трогая остальной
    файл. Запись атомарная (временный файл + os.replace) с проверкой
    синтаксиса ПЕРЕД заменой оригинала и резервной копией — сбой посреди
    записи (антивирус/отключение питания) не испортит config.py (см.
    safe_files.py). После записи доступ к файлу ограничивается текущим
    пользователем Windows."""
    config_path = os.path.join(_app_dir, "config.py")
    with open(config_path, "r", encoding="utf-8") as f:
        text = f.read()
    pattern = re.compile(rf"^{re.escape(key)}\s*=.*$", re.MULTILINE)
    new_line = f"{key} = {value_literal}"
    if pattern.search(text):
        text = pattern.sub(new_line, text, count=1)
    else:
        text = text.rstrip("\n") + f"\n{new_line}\n"
    safe_files.atomic_write_text(config_path, text, validate=safe_files.validate_python_syntax)
    safe_files.restrict_to_current_user(config_path)


def _write_config_block(key: str, new_value_literal: str):
    """Как _write_config_value, но для МНОГОСТРОЧНЫХ литералов (RISK_PROFILES,
    MARKET_CONTEXT — словари, которые в config.py форматированы на много строк
    ради читаемости). _write_config_value не годится для них — её regex
    заменяет только ПЕРВУЮ строку присваивания, оставляя "хвост" старого
    многострочного литерала висеть ниже как мусор. Эта функция вместо этого
    находит `key = ...` и определяет КОНЕЦ значения по балансу скобок (с
    учётом строк в кавычках, чтобы не сбиться на скобке внутри текста)."""
    config_path = os.path.join(_app_dir, "config.py")
    with open(config_path, "r", encoding="utf-8") as f:
        text = f.read()

    pattern = re.compile(rf"^{re.escape(key)}\s*=\s*", re.MULTILINE)
    m = pattern.search(text)
    if not m:
        new_text = text.rstrip("\n") + f"\n{key} = {new_value_literal}\n"
        safe_files.atomic_write_text(config_path, new_text, validate=safe_files.validate_python_syntax)
        safe_files.restrict_to_current_user(config_path)
        return

    start = m.end()
    depth = 0
    in_str = None
    started = False
    i = start
    while i < len(text):
        ch = text[i]
        if in_str:
            if ch == "\\":
                i += 2
                continue
            if ch == in_str:
                in_str = None
        elif ch in "\"'":
            in_str = ch
        elif ch in "([{":
            depth += 1
            started = True
        elif ch in ")]}":
            depth -= 1
            if started and depth == 0:
                i += 1
                break
        elif not started and ch == "\n":
            # значение без скобок на одной строке — на всякий случай, для этой
            # функции это не основной сценарий, но не должно зависать.
            break
        i += 1
    end = i

    new_text = text[:start] + new_value_literal + text[end:]
    safe_files.atomic_write_text(config_path, new_text, validate=safe_files.validate_python_syntax)
    safe_files.restrict_to_current_user(config_path)


def _format_risk_profiles(profiles: dict) -> str:
    """RISK_PROFILES использует ключи RiskProfile.XXX (enum) — repr() такого
    словаря НЕ является валидным Python-исходником (repr enum-члена выглядит
    как "<RiskProfile.CONSERVATIVE: 'conservative'>"). Поэтому собираем текст
    вручную в том же формате, что и оригинальный config.py."""
    lines = ["{"]
    for enum_member, params in profiles.items():
        parts = ", ".join(f"{k}={v!r}" for k, v in params.items())
        lines.append(f"    RiskProfile.{enum_member.name}: dict({parts}),")
    lines.append("}")
    return "\n".join(lines)


def _reload_cfg():
    """importlib.reload(cfg) + повторная расшифровка секретов (reload читает
    config.py заново с диска, где секреты снова в виде "enc:..." — нужно
    расшифровать их в памяти опять тем же паролем входа, что ввели при
    старте). Используй ВМЕСТО голого importlib.reload(cfg) везде в этом файле."""
    importlib.reload(cfg)
    pw = control.get_session_password()
    if pw:
        try:
            secure_store.unlock_config(cfg, pw)
        except ValueError as e:
            log.error("Не удалось расшифровать секреты после перечитывания config.py: %s", e)
    # Настройки только что менялись — обновляем постоянную копию. Она лежит
    # в папке пользователя и переживает и обновление, и перенос программы:
    # запуск свежескачанного .exe из другой папки больше не показывает
    # заводские настройки вместо ваших.
    try:
        settings_backup.save()
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось обновить копию настроек: %s", e)


def _migrate_legacy_secrets():
    """Одноразовая миграция при первом запуске новой версии: раньше
    config.py хранил пароль дашборда и остальные секреты (MT5/AI/новости)
    открытым текстом. Теперь пароль дашборда хранится только как хэш
    (DASHBOARD_PASSWORD_HASH), а остальные секреты — зашифрованы этим же
    паролем (см. secure_store.py). Если конфиг ещё старого формата —
    мигрируем автоматически ТЕМ ЖЕ паролем, что сейчас в файле, без каких-либо
    действий пользователя. Best-effort: если что-то пойдёт не так — оставляем
    всё как было (старый открытый режим), программа не ломается."""
    try:
        if secure_store.private_mode():
            # В приватном режиме секреты намеренно лежат открытым текстом.
            # Зашифровать их «на всякий случай» значило бы вернуть тот самый
            # тупик, ради выхода из которого режим и сделан: вход отключён,
            # пароля нет, расшифровать нечем.
            return
        if getattr(cfg, "DASHBOARD_PASSWORD_HASH", ""):
            return  # уже мигрировано
        legacy_password = getattr(cfg, "DASHBOARD_PASSWORD", "")
        if not legacy_password:
            return  # нечем шифровать и не с чем сравнивать при входе — оставляем как есть

        salt = getattr(cfg, "SECURITY_SALT", "") or secure_store.new_salt()
        password_hash = secure_store.hash_password(legacy_password, salt)

        for field in ("MT5_PASSWORD", "ANTHROPIC_API_KEY", "OPENAI_API_KEY"):
            value = getattr(cfg, field, "")
            if isinstance(value, str) and value and not value.startswith(secure_store.ENC_PREFIX):
                encrypted = secure_store.encrypt_value(value, legacy_password, salt)
                _write_config_value(field, repr(encrypted))

        news_keys = getattr(cfg, "NEWS_API_KEYS", {}) or {}
        if isinstance(news_keys, dict) and news_keys:
            new_news_keys = {}
            changed = False
            for k, v in news_keys.items():
                if isinstance(v, str) and v and not v.startswith(secure_store.ENC_PREFIX):
                    new_news_keys[k] = secure_store.encrypt_value(v, legacy_password, salt)
                    changed = True
                else:
                    new_news_keys[k] = v
            if changed:
                _write_config_value("NEWS_API_KEYS", repr(new_news_keys))

        _write_config_value("SECURITY_SALT", repr(salt))
        _write_config_value("DASHBOARD_PASSWORD_HASH", repr(password_hash))
        # Сам пароль убираем из файла в открытом виде — теперь проверяется
        # через хэш выше, хранить его вторым разом открытым текстом незачем.
        _write_config_value("DASHBOARD_PASSWORD", '""')

        importlib.reload(cfg)
        log.info("config.py мигрирован на шифрование секретов (первый запуск новой версии программы).")
    except Exception as e:
        log.exception("Миграция шифрования секретов не удалась — работаю в старом (открытом) режиме: %s", e)


def _harden_files():
    """Best-effort защита файлов при старте программы: ограничение доступа
    (Windows ACL, только текущий пользователь) для config.py/лога/журнала
    сделок, плюс проверка целостности журнала сделок (сверка с sha256 от
    прошлого запуска — предупреждает, если файл менялся снаружи программы)."""
    try:
        config_path = os.path.join(_app_dir, "config.py")
        log_path = LOG_FILE
        trades_path = os.path.join(_app_dir, getattr(cfg, "LOG_CSV_PATH", "trades_log.csv"))

        for p in (config_path, log_path, trades_path):
            safe_files.restrict_to_current_user(p)

        if os.path.exists(trades_path) and not safe_files.check_integrity(trades_path):
            msg = ("trades_log.csv изменился СНАРУЖИ программы с прошлого запуска. "
                   "Если это не ты правил файл руками — проверь компьютер на посторонний доступ.")
            log.warning(msg)
            control.push_notification("Проверка целостности", msg)
        if os.path.exists(trades_path):
            safe_files.mark_integrity_current(trades_path)
    except Exception as e:
        log.warning("Не удалось выполнить проверку целостности/ограничение доступа к файлам: %s", e)


# ---- "Запомнить пароль" на экране входа (Windows DPAPI, см. secure_store.py) ----
_REMEMBER_PATH = os.path.join(_app_dir, ".login_remember")


def _save_remembered_password(password: str):
    try:
        blob = secure_store.dpapi_protect(password.encode("utf-8"))
        with open(_REMEMBER_PATH, "wb") as f:
            f.write(base64.b64encode(blob))
        safe_files.restrict_to_current_user(_REMEMBER_PATH)
    except Exception as e:
        log.warning("Не удалось сохранить пароль для автозаполнения (не критично): %s", e)


def _load_remembered_password() -> str:
    try:
        if not os.path.exists(_REMEMBER_PATH):
            return ""
        with open(_REMEMBER_PATH, "rb") as f:
            blob = base64.b64decode(f.read())
        return secure_store.dpapi_unprotect(blob).decode("utf-8")
    except Exception:
        return ""


def _clear_remembered_password():
    try:
        if os.path.exists(_REMEMBER_PATH):
            os.remove(_REMEMBER_PATH)
    except Exception:
        pass


# Разделы вкладки "Настройка": 118 параметров в 25 группах — в одном списке
# их не найти. Раскладываем по смыслу на несколько подвкладок.
CONFIG_SECTIONS = [
    ("Торговля", ["Общее", "Часы торговли", "Лот", "Score-фильтр"]),
    ("Сигнал", ["Индикаторы", "Price Action / откат", "Объём", "Доп. индикаторы",
                "Собственная стратегия", "Режим рынка", "Контекст рынка",
                "Анти-'зеркало' фильтры", "Автонастройка под инструмент"]),
    ("Риск", ["Стопы / TP", "Break Even", "Трейлинг-стоп", "Profit Lock",
                      "Просадка / серии убытков", "Издержки", "Частичное закрытие",
                      "Фиксация прибыли"]),
    ("Защита", ["Защитные проверки", "Новости (пороги)"]),
    ("AI-сигнал", ["AI-сигнал"]),
    ("Обучение", ["Автообучение", "Автообновление"]),
    ("Вид", ["Оформление"]),
]


class App:
    def __init__(self):
        self.root = tk.Tk()
        # Версия прямо в заголовке окна: владелец просил, чтобы было видно,
        # обновилась программа или нет. Заголовок виден всегда — и в окне,
        # и на панели задач, — в отличие от вкладки, куда надо зайти.
        self.root.title(f"{APP_TITLE} — {app_version.short()}")
        # Окно шире: сверху панель управления, снизу строка сохранения и
        # выхода, вкладок полтора десятка. Владелец просил не растягивать
        # окно каждый раз — значит по умолчанию оно должно быть достаточным.
        self.root.geometry("1040x720")
        self.root.minsize(900, 600)

        self._apply_theme()

        self.stop_event = None
        self.bot_thread = None
        # Бот ДОЛЖЕН работать (не путать с «работает сейчас»): по этому
        # флагу сторож отличает нажатый человеком «Стоп» от смерти цикла.
        self._bot_should_run = False
        self.tray_icon = None
        self._dashboard_started = False
        self.chat_history = []

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Первый запуск: сами ставим советники и сервис в MetaTrader, чтобы
        # пользователю не приходилось запускать отдельные установщики.
        self.root.after(1200, self._auto_install_into_mt5_once)
        # Мост для советников — тоже сам, если включён.
        self.root.after(1500, self._start_bridge_if_enabled)
        # Календарь дожимается теми же повторами: терминал бывает не готов
        self.root.after(4000, self._ensure_news_source)
        # Обновление при запуске. Момент выбран не случайно: торговля ещё не
        # началась, открытых позиций у бота нет — подменять его безопасно.
        self._auto_update_busy = False
        self._start_bot_waits = 0
        if updater.enabled() and getattr(cfg, "UPDATE_AUTO_APPLY", False):
            self._auto_update_busy = True
            self.root.after(300, self._auto_apply_update)
        elif getattr(cfg, "UPDATE_CHECK_ON_START", True) and updater.enabled():
            # Без галочки автоустановки — только смотрим и спрашиваем.
            self.root.after(6000, lambda: self.check_updates(silent=True))

        # ПОВТОРНАЯ проверка обновлений во время работы.
        #
        # Проверки «при запуске» хватало, пока программу открывали и закрывали
        # каждый день. На сервере она работает неделями и не перезапускается —
        # там «при запуске» означает «никогда», и новые сборки не доезжают.
        self._schedule_update_check()

        if cfg.USE_WEB_DASHBOARD and not self._dashboard_started:
            try:
                bot_engine.start_dashboard_thread()
                self._dashboard_started = True
            except Exception as e:
                log.exception("Не удалось поднять веб-дашборд: %s", e)

        if TRAY_AVAILABLE:
            self._start_tray()

        # Лента событий прошлого запуска: если программа остановилась ночью,
        # причина должна быть видна утром, а не потеряна вместе с окном.
        try:
            runtime_events.load()
        except Exception:  # noqa: BLE001
            pass

        self._refresh_loop()
        # Календарь подтягивает новости сам — владелец просил, чтобы он
        # обновлялся без нажатий. Первый раз с задержкой: при старте окно
        # ещё рисуется, а связи с терминалом может ещё не быть.
        self.root.after(8000, self._news_auto_refresh)

        # Автозапуск торгового цикла вместе с программой — не нужно нажимать
        # "Старт" руками. Кнопки Старт/Стоп остаются для ручной остановки/
        # перезапуска. Небольшая задержка — чтобы окно успело отрисоваться.
        self.root.after(400, self._start_bot_when_ready)

    # ---- тема оформления --------------------------------------------------
    def _apply_theme(self):
        """Оформление окна. Цвета лежат в ui_theme.py — все в одном месте, с
        проверенным контрастом. Раньше они были вписаны прямо здесь и ещё в
        семи десятках мест по коду, а серый текст на почти чёрном фоне
        читался плохо: «на чёрном фоне очень плохо всё видно»."""
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        self.colors = ui_theme.from_config(cfg)
        ui_theme.apply(self.root, style, self.colors)

    # ---- интерфейс: вкладки -------------------------------------------------
    def _build_ui(self):
        # ---------- Панель управления СВЕРХУ ----------
        # Владелец: «добавь кнопки перезапуск, пауза, старт, и пусть они будут
        # вверху и доступны со всех вкладок». Раньше старт и стоп жили на
        # вкладке «Обзор»: чтобы остановить бота с любой другой вкладки, надо
        # было сначала до неё добраться.
        top = ttk.Frame(self.root)
        top.pack(side="top", fill="x", padx=8, pady=(8, 0))

        self.btn_start = ttk.Button(top, text="▶ Старт", command=self.start_bot)
        self.btn_start.pack(side="left")
        self.btn_stop = ttk.Button(top, text="■ Стоп", command=self.stop_bot,
                                   state="disabled")
        self.btn_stop.pack(side="left", padx=6)
        self.btn_pause = ttk.Button(top, text="⏸ Пауза", command=self.toggle_pause)
        self.btn_pause.pack(side="left", padx=(0, 6))
        self.btn_restart = ttk.Button(top, text="⟳ Перезапуск",
                                      command=self.restart_bot)
        self.btn_restart.pack(side="left")

        # Прежние имена — на те же кнопки: код запуска, остановки и сторожа
        # обращается к ним по-старому, и переписывать его незачем.
        self.start_btn = self.btn_start
        self.stop_btn = self.btn_stop

        self.top_status_var = tk.StringVar(value="Остановлен")
        ttk.Label(top, textvariable=self.top_status_var,
                  font=("Segoe UI Semibold", 10)).pack(side="left", padx=14)

        # ---------- Низ: одна кнопка сохранения и выход ----------
        bottom = ttk.Frame(self.root)
        bottom.pack(side="bottom", fill="x", padx=8, pady=(0, 6))

        # Выход — справа внизу, как просил владелец: «полностью выходит из
        # всего, что запускалось».
        ttk.Button(bottom, text="Выход", command=self.full_exit).pack(side="right")
        ttk.Label(bottom, text="made by Viacheslav.Y.",
                  foreground=self.colors["dim"], font=("Segoe UI", 8)
                  ).pack(side="right", padx=10)

        # Одна главная кнопка сохранения на всю программу. Раньше их было
        # семь, по своей на каждый раздел, и человек не знал, какую нажимать
        # и сохранил ли он вообще всё.
        ttk.Button(bottom, text="💾 Сохранить все настройки",
                   command=self.save_everything).pack(side="left")
        self.save_all_status_var = tk.StringVar(value="")
        ttk.Label(bottom, textvariable=self.save_all_status_var,
                  foreground=self.colors["muted"]).pack(side="left", padx=10)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=6, pady=6)

        # ВЕРХНИХ ВКЛАДОК СЕМЬ, А НЕ ЧЕТЫРНАДЦАТЬ.
        #
        # Владелец: «вкладки, кнопки выходят за границы, приходится
        # увеличивать окно». Четырнадцать вкладок в строку просто не влезали:
        # их полоса шире минимальной ширины окна, поэтому названия обрезались.
        # Ни одна страница при этом не потерялась — они разложены по группам,
        # и внутри группы переключаются вторым, более узким рядом вкладок.
        # Раскладка описана данными в ui_layout.py, и на неё есть тест,
        # который считает ширину в пикселях, а не смотрит на скриншот.
        self.tab_frames = {}     # имя страницы -> её рамка
        self.tab_books = {}      # имя страницы -> блокнот, в котором она лежит
        for группа in ui_layout.group_names():
            страницы = ui_layout.pages(группа)
            if ui_layout.is_single(группа):
                # Одна страница — второго ряда вкладок не нужно.
                рамка = ttk.Frame(self.notebook)
                self.notebook.add(рамка, text=группа)
                self.tab_frames[страницы[0]] = рамка
                self.tab_books[страницы[0]] = self.notebook
                continue
            обёртка = ttk.Frame(self.notebook)
            self.notebook.add(обёртка, text=группа)
            внутри = ttk.Notebook(обёртка)
            внутри.pack(fill="both", expand=True)
            for имя in страницы:
                рамка = ttk.Frame(внутри)
                внутри.add(рамка, text=имя)
                self.tab_frames[имя] = рамка
                self.tab_books[имя] = внутри

        # Вкладка «Сделки» убрана по просьбе владельца: открытые позиции
        # видны на вкладке «Счета», там же их и закрывают. Рамка остаётся
        # (в неё по-прежнему строится таблица позиций), но в окно не
        # добавляется — код сборки таблицы используется экспортом в Excel.
        tab_positions = ttk.Frame(self.notebook)

        self._build_tab_overview(self.tab_frames["Обзор"])
        self._build_tab_broker(self.tab_frames["Брокер"])
        self._build_tab_symbols(self.tab_frames["Символы"])
        self._build_tab_positions(tab_positions)
        self._build_tab_log(self.tab_frames["Лог"])
        self._build_tab_equity(self.tab_frames["Equity"])
        self._build_tab_config(self.tab_frames["Настройка"])
        self._build_tab_schedule(self.tab_frames["Календарь"])
        self._build_tab_sources(self.tab_frames["Источники"])
        self._build_tab_system(self.tab_frames["Система"])
        self._build_tab_telegram(self.tab_frames["Сигналы"])
        self._build_tab_news(self.tab_frames["Новости"])
        self._build_tab_chat(self.tab_frames["Чат"])
        self._build_tab_help(self.tab_frames["Помощь"])
        self._build_tab_accounts(self.tab_frames["Счета"])

        self._check_tab_strip()
        self._apply_ui_mode(initial=True)

    def _check_tab_strip(self):
        """Перемерить полосу вкладок настоящим шрифтом уже открытого окна.

        Оценка в ui_layout.py приблизительная: настоящая ширина знака зависит
        от шрифта системы и масштаба экрана. Здесь она сверяется с
        действительностью. Не помещается — пишем в журнал, а не ломаем окно:
        программа обязана открыться и на нестандартном шрифте."""
        try:
            имена = ui_layout.group_names()
            нужно = ui_layout.measure_strip(имена)
            # Окно ещё не показано, и winfo_width() отдаёт 1. Меряем от
            # минимальной ширины: уже неё окно всё равно не станет.
            есть = max(self.root.winfo_width(), ui_layout.МИН_ШИРИНА_ОКНА)
            if нужно and нужно > есть:
                log.warning("Полоса вкладок шире окна: нужно %s px, есть %s px",
                            нужно, есть)
        except Exception:  # noqa: BLE001
            log.debug("Не удалось перемерить полосу вкладок", exc_info=True)

    # ---- Вкладка "Счета": несколько торговых счетов MT5 ---------------------
    def _build_tab_accounts(self, parent):
        """Список счетов, их состояние и закрытие позиций.

        Вся логика в accounts_tab.py — здесь только подключение, чтобы этот
        файл не разрастался. Ошибка при построении не должна ронять всё окно:
        остальные вкладки продолжат работать.
        """
        try:
            from accounts_tab import AccountsTab
            self.accounts_tab = AccountsTab(parent, self.root)
        except Exception as e:  # noqa: BLE001
            self.accounts_tab = None
            log.exception("Не удалось построить вкладку «Счета»")
            ttk.Label(parent,
                      text=f"Вкладка «Счета» недоступна: {e}",
                      wraplength=600, justify="left").pack(padx=16, pady=16)

    # ---- Простой/Продвинутый режим -----------------------------------------
    def _apply_ui_mode(self, initial: bool = False):
        is_advanced = (self.ui_mode_var.get() == "Продвинутый")
        for name in ADVANCED_TAB_NAMES:
            frame = self.tab_frames.get(name)
            # Страница может лежать во втором ряду вкладок, внутри группы,
            # поэтому прятать её надо в ЕЁ блокноте, а не в главном.
            book = self.tab_books.get(name)
            if frame is None or book is None:
                continue
            shown = str(frame) in book.tabs()
            if is_advanced and not shown:
                book.add(frame, text=name)
            elif not is_advanced and shown:
                book.hide(frame)
        if not initial:
            try:
                _write_config_value("UI_MODE", repr("advanced" if is_advanced else "simple"))
            except Exception:
                pass

    # ---- вкладка "Обзор" ----------------------------------------------------
    def _build_tab_overview(self, parent):
        """Вкладка «Обзор» — одна страница, на которой видно состояние.

        Владелец: «переделай вкладку обзор, убери лишнее», «чтобы ничего не
        повторялось». Здесь было три кнопки управления (Старт, Стоп, Полный
        выход) — теперь они на постоянной панели сверху и внизу, и на всех
        вкладках сразу. Была своя строка состояния — она же есть в верхней
        панели. Счёт и статистика лежали в двух отдельных рамках, хотя это
        одно и то же: про счёт.

        Осталось только то, чего больше нигде нет, и каждое ровно в одном
        месте."""
        parent = self._scrollable(parent)

        # ---------- Шапка: что за программа и какой версии ----------
        head = ttk.Frame(parent)
        head.pack(fill="x", padx=12, pady=(10, 2))
        ttk.Label(head, text=APP_TITLE,
                  font=("Segoe UI", 15, "bold")).pack(side="left")
        ttk.Label(head, text=app_version.full(), foreground=self.colors["muted"],
                  font=("Segoe UI", 9)).pack(side="left", padx=10, pady=(6, 0))

        # Строка состояния здесь НЕ дублируется: она в верхней панели, видна
        # со всех вкладок. Переменная нужна остальному коду — оставляем её,
        # но второй надписи об одном и том же на экране больше нет.
        self.status_var = tk.StringVar(value="Остановлен")

        # ---------- Счёт и статистика: одна рамка вместо двух ----------
        account = ttk.LabelFrame(parent, text=" Счёт ")
        account.pack(fill="x", padx=12, pady=6)
        inner = ttk.Frame(account)
        inner.pack(fill="x", padx=8, pady=6)
        self.info_var = tk.StringVar(value="Бот ещё не запускался.")
        ttk.Label(inner, textvariable=self.info_var, justify="left"
                  ).grid(row=0, column=0, sticky="nw")
        self.stats_var = tk.StringVar(value="—")
        ttk.Label(inner, textvariable=self.stats_var, justify="left",
                  foreground=self.colors["muted"]
                  ).grid(row=0, column=1, sticky="nw", padx=(30, 0))
        inner.columnconfigure(1, weight=1)

        # ---------- Что мешает торговать ----------
        # Одна рамка на все предупреждения: разрешение на торговлю, причины
        # молчания по парам и недавние происшествия. Раньше это было
        # разбросано, а часть не показывалась вовсе.
        self.trade_warning_var = tk.StringVar(value="")
        # Место под предупреждение зарезервировано всегда — так рамка при
        # появлении встаёт РОВНО СЮДА, а не в конец страницы. Сама рамка
        # показывается только когда есть что сказать (см. _refresh_loop):
        # пустая рамка «Внимание» на пол-экрана пугает без причины.
        self.trade_warning_slot = ttk.Frame(parent)
        self.trade_warning_slot.pack(fill="x")
        self.trade_warning_frame = ttk.LabelFrame(self.trade_warning_slot,
                                                  text=" Внимание ")
        # НЕБОЛЬШОЕ ОКНО С ПОЛЗУНКОМ, А НЕ РАСТУЩАЯ НАДПИСЬ. Владелец прислал
        # снимок: рамка «Внимание» заняла ПОЛ-ЭКРАНА и была целиком красной —
        # потому что в неё попал список из 497 отобранных пар. Из-за этого
        # настоящие предупреждения терялись, а красный цвет перестал что-либо
        # значить: красным было всё подряд.
        #
        # Высота фиксирована, длинное уезжает под ползунок. Красным теперь
        # выделяется ТОЛЬКО важное — см. _warning_severity().
        box = ttk.Frame(self.trade_warning_frame)
        box.pack(fill="x", padx=8, pady=6)
        self.trade_warning_text = tk.Text(
            box, height=6, wrap="word", relief="flat", borderwidth=0,
            background=self.colors["bg"], foreground=self.colors["fg"],
            highlightthickness=0)
        scroll = ttk.Scrollbar(box, orient="vertical",
                               command=self.trade_warning_text.yview)
        self.trade_warning_text.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        self.trade_warning_text.pack(side="left", fill="both", expand=True)
        self.trade_warning_text.tag_configure("важно",
                                              foreground=self.colors["loss"])
        self.trade_warning_text.tag_configure("обычное",
                                              foreground=self.colors["fg"])
        self.trade_warning_text.configure(state="disabled")

        # ---------- Действия ----------
        actions = ttk.LabelFrame(parent, text=" Действия ")
        actions.pack(fill="x", padx=12, pady=6)
        row = ttk.Frame(actions)
        row.pack(fill="x", padx=8, pady=6)
        buttons = [
            ("Экспорт в Excel", self.export_excel),
            ("Открыть логи", self.open_logs),
            ("Открыть config.py", self.open_config),
            ("Дашборд для телефона", self.open_dashboard),
        ]
        for column, (text, command) in enumerate(buttons):
            ttk.Button(row, text=text, command=command).grid(
                row=0, column=column, padx=(0, 6))

        # ---------- Что приехало с GitHub ----------
        # Название НЕ «Обновление»: установкой занимается вкладка «Система»,
        # раздел «Обновление из GitHub». Здесь только посмотреть, что нового,
        # — два одинаково названных раздела в разных местах и были той самой
        # путаницей, которую владелец просил убрать.
        sync = ttk.LabelFrame(parent, text=" Что нового в программе ")
        sync.pack(fill="x", padx=12, pady=6)
        sync_row = ttk.Frame(sync)
        sync_row.pack(fill="x", padx=8, pady=6)
        ttk.Button(sync_row, text="Проверить изменения",
                   command=self.sync_from_cloud).grid(row=0, column=0)
        ttk.Button(sync_row, text="Что нового",
                   command=self.show_changes).grid(row=0, column=1, padx=6)
        self.sync_status_var = tk.StringVar(value="")
        ttk.Label(sync, textvariable=self.sync_status_var,
                  foreground=self.colors["muted"], wraplength=900,
                  justify="left").pack(anchor="w", padx=8, pady=(0, 6))
        self._refresh_sync_status()

        # ---------- Как показывать программу ----------
        view = ttk.LabelFrame(parent, text=" Вид и запуск ")
        view.pack(fill="x", padx=12, pady=(6, 12))
        view_row = ttk.Frame(view)
        view_row.pack(fill="x", padx=8, pady=6)
        ttk.Label(view_row, text="Режим интерфейса:").grid(row=0, column=0)
        initial_mode = ("Продвинутый"
                        if getattr(cfg, "UI_MODE", "simple") == "advanced"
                        else "Простой")
        self.ui_mode_var = tk.StringVar(value=initial_mode)
        mode_combo = ttk.Combobox(view_row, textvariable=self.ui_mode_var,
                                  values=["Простой", "Продвинутый"],
                                  state="readonly", width=15)
        mode_combo.grid(row=0, column=1, padx=(6, 20))
        mode_combo.bind("<<ComboboxSelected>>", lambda e: self._apply_ui_mode())

        # ВЫБОР ТЕМЫ. Раньше тема менялась только правкой UI_THEME в файле
        # настроек — то есть для владельца её не существовало.
        ttk.Label(view_row, text="Тема:").grid(row=0, column=2)
        self._theme_titles = {key: title for key, title in ui_theme.choices()}
        self._theme_keys = {title: key for key, title in ui_theme.choices()}
        current_theme = str(getattr(cfg, "UI_THEME", ui_theme.DEFAULT) or "").lower()
        self.theme_var = tk.StringVar(
            value=self._theme_titles.get(current_theme,
                                         self._theme_titles[ui_theme.DEFAULT]))
        theme_combo = ttk.Combobox(view_row, textvariable=self.theme_var,
                                   values=[t for _, t in ui_theme.choices()],
                                   state="readonly", width=16)
        theme_combo.grid(row=0, column=3, padx=(6, 20))
        theme_combo.bind("<<ComboboxSelected>>", lambda e: self._save_theme_choice())

        self.autostart_var = tk.BooleanVar(value=_is_autostart_enabled())
        ttk.Checkbutton(view_row, text="Запускать вместе с Windows",
                        variable=self.autostart_var,
                        command=self._toggle_autostart).grid(row=0, column=2)

        ttk.Label(view, foreground=self.colors["muted"], justify="left",
                  text=f"С телефона по Wi-Fi: "
                       f"http://<IP-компьютера>:{cfg.DASHBOARD_PORT}"
                  ).pack(anchor="w", padx=8, pady=(0, 6))

    # ---- вкладка "Брокер" ----------------------------------------------------
    def _build_tab_broker(self, parent):
        # Боковой ползунок: владелец просил не растягивать окно
        # каждый раз — блоков здесь больше, чем помещается.
        parent = self._scrollable(parent)
        pad = {"padx": 10, "pady": 6}

        ttk.Label(parent, text="Подключение к брокеру (любой MT5-брокер)",
                  font=("Segoe UI", 12, "bold")).pack(**pad)
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=680, justify="left", text=
                  "Впиши сервер/логин/пароль торгового счёта — как при входе в MetaTrader 5. "
                  "Терминал MT5 всё равно должен быть УСТАНОВЛЕН на этом компьютере, но держать его "
                  "открытым и залогиненным заранее уже не обязательно — программа сделает это сама."
                  ).pack(**pad)

        self.use_existing_var = tk.BooleanVar(value=not bool(getattr(cfg, "MT5_LOGIN", 0)))
        ttk.Checkbutton(parent, text="Использовать уже открытый и залогиненный терминал (без пароля в файле)",
                        variable=self.use_existing_var, command=self._toggle_broker_fields).pack(anchor="w", **pad)

        form = ttk.Frame(parent)
        form.pack(fill="x", **pad)

        ttk.Label(form, text="Сервер брокера:").grid(row=0, column=0, sticky="w", pady=4)
        self.server_var = tk.StringVar(value=str(getattr(cfg, "MT5_SERVER", "") or ""))
        self.server_entry = ttk.Entry(form, textvariable=self.server_var, width=34)
        self.server_entry.grid(row=0, column=1, pady=4, sticky="w")

        ttk.Label(form, text="Логин (номер счёта):").grid(row=1, column=0, sticky="w", pady=4)
        self.login_var = tk.StringVar(value=str(getattr(cfg, "MT5_LOGIN", "") or ""))
        self.login_entry = ttk.Entry(form, textvariable=self.login_var, width=34)
        self.login_entry.grid(row=1, column=1, pady=4, sticky="w")

        ttk.Label(form, text="Пароль:").grid(row=2, column=0, sticky="w", pady=4)
        self.password_var = tk.StringVar(value=str(getattr(cfg, "MT5_PASSWORD", "") or ""))
        self.password_entry = ttk.Entry(form, textvariable=self.password_var, width=34, show="*")
        self.password_entry.grid(row=2, column=1, pady=4, sticky="w")

        ttk.Label(form, text="Путь к terminal64.exe (необязательно):").grid(row=3, column=0, sticky="w", pady=4)
        self.term_path_var = tk.StringVar(value=str(getattr(cfg, "MT5_TERMINAL_PATH", "") or ""))
        self.term_path_entry = ttk.Entry(form, textvariable=self.term_path_var, width=34)
        self.term_path_entry.grid(row=3, column=1, pady=4, sticky="w")
        ttk.Button(form, text="Обзор...", command=self._browse_terminal_path).grid(row=3, column=2, padx=6)

        btn_frame = ttk.Frame(parent)
        btn_frame.pack(**pad)
        self.test_conn_btn = ttk.Button(btn_frame, text="Проверить подключение", command=self.test_connection)
        self.test_conn_btn.grid(row=0, column=1, padx=5)

        self.broker_status_var = tk.StringVar(value="")
        ttk.Label(parent, textvariable=self.broker_status_var, wraplength=680, justify="left").pack(**pad)

        self._toggle_broker_fields()

    def _toggle_broker_fields(self):
        state = "disabled" if self.use_existing_var.get() else "normal"
        for w in (self.server_entry, self.login_entry, self.password_entry):
            w.config(state=state)

    def _browse_terminal_path(self):
        path = filedialog.askopenfilename(
            title="Выбери terminal64.exe",
            filetypes=[("MetaTrader 5", "terminal64.exe"), ("Все файлы", "*.*")],
        )
        if path:
            self.term_path_var.set(path)

    def save_broker_settings(self, silent: bool = False):
        try:
            if self.use_existing_var.get():
                _write_config_value("MT5_LOGIN", "0")
                _write_config_value("MT5_PASSWORD", '""')
                _write_config_value("MT5_SERVER", '""')
            else:
                login_text = self.login_var.get().strip()
                server_text = self.server_var.get().strip()
                if not login_text.isdigit():
                    messagebox.showerror(APP_TITLE, "Логин должен быть числом (номер торгового счёта).")
                    return
                if not server_text:
                    messagebox.showerror(APP_TITLE, "Укажи сервер брокера (как в MetaTrader 5 при входе).")
                    return
                _write_config_value("MT5_LOGIN", login_text)
                # Пароль брокера храним в config.py зашифрованным паролем
                # входа (см. secure_store.py) — на диске он не читается
                # открытым текстом, даже если файл скопировать на другой
                # компьютер.
                raw_password = self.password_var.get()
                pw = control.get_session_password()
                salt = getattr(cfg, "SECURITY_SALT", "")
                # protect_secret сам решает, шифровать или класть открытым
                # текстом (приватный режим) — см. secure_store.private_mode()
                stored_password = secure_store.protect_secret(raw_password, pw, salt)
                _write_config_value("MT5_PASSWORD", repr(stored_password))
                _write_config_value("MT5_SERVER", repr(server_text))
            _write_config_value("MT5_TERMINAL_PATH", repr(self.term_path_var.get().strip()))
            _reload_cfg()
            if not silent:
                messagebox.showinfo(APP_TITLE, "Настройки подключения сохранены.")
        except Exception as e:
            log.exception("Не удалось сохранить настройки брокера: %s", e)
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить: {e}")

    def test_connection(self):
        if self.bot_thread and self.bot_thread.is_alive():
            messagebox.showinfo(APP_TITLE, "Останови бота перед проверкой подключения.")
            return
        self.test_conn_btn.config(state="disabled")
        self.broker_status_var.set("Проверяю подключение...")
        threading.Thread(target=self._test_connection_worker, daemon=True).start()

    def _test_connection_worker(self):
        try:
            acc = bot_engine.mt5c.connect()
            msg = f"Подключено: счёт {acc.login} ({acc.server}), баланс {acc.balance:.2f} {acc.currency}"
            bot_engine.mt5c.disconnect()
            self.root.after(0, lambda: self.broker_status_var.set(msg))
        except Exception as e:
            err = str(e)
            self.root.after(0, lambda: self.broker_status_var.set(f"Ошибка: {err}"))
        finally:
            self.root.after(0, lambda: self.test_conn_btn.config(state="normal"))

    # ---- вкладка "Символы" ---------------------------------------------------
    def _build_tab_symbols(self, parent):
        # ПОЛЯ «ДОБАВИТЬ СИМВОЛ» ЗДЕСЬ БОЛЬШЕ НЕТ. Список пар программа берёт
        # у брокера сама и сама отбирает подходящие (symbol_picker.py), так
        # что вписывать пары руками стало не нужно — а поле ввода, которое
        # ничего не решает, только путает.
        #
        # А вот таблица нужна СИЛЬНЕЕ прежнего: пар теперь не четыре, а
        # десятки, и это единственное место, где видно, что именно выбрано и
        # почему пара сейчас не торгует (колонка «Отказ»).
        ttk.Label(parent, text="Пары программа выбирает у брокера сама. Двойной клик по «Вкл» — "
                               "выключить пару вручную. По «Лот» — задать фиксированный лот. "
                               "По «Символ» — мини-график цены.",
                  foreground=self.colors["muted"], wraplength=800, justify="left").pack(padx=10, pady=6, anchor="w")

        self.available_symbols_var = tk.StringVar(
            value="Отбор пар ещё не выполнен (появится, когда бот запущен и подключён к MT5)."
        )
        ttk.Label(parent, textvariable=self.available_symbols_var,
                  foreground=self.colors["muted"], wraplength=800,
                  justify="left").pack(padx=10, pady=(0, 6), anchor="w")

        self.symbols_columns = ("enabled", "symbol", "lot", "buy", "sell", "regime", "ai", "custom", "multi", "learn", "paused", "reject", "risk")
        headings = ("Вкл", "Символ", "Лот", "BUY", "SELL", "Режим", "AI", "Своя стратегия", "Индикаторы", "Автообучение", "Пауза", "Отказ", "Риск лота")
        self.symbols_tree = ttk.Treeview(parent, columns=self.symbols_columns, show="headings", height=12)
        for col, head in zip(self.symbols_columns, headings):
            self.symbols_tree.heading(col, text=head)
            self.symbols_tree.column(col, width=90, anchor="center")
        self.symbols_tree.column("reject", width=220, anchor="w")
        self.symbols_tree.column("symbol", width=90, anchor="w")
        # "Риск лота" — минимальный лот брокера рискует больше настроенного
        # процента (частая причина "очень много убытка" на маленьком
        # депозите: ниже минимального лота опуститься нельзя, чем бы ни был
        # задан риск на сделку). Красным — чтобы не потерялось среди колонок.
        self.symbols_tree.column("risk", width=260, anchor="w")
        self.symbols_tree.tag_configure("risk_over", foreground=self.colors["loss"])
        self.symbols_tree.pack(fill="both", expand=True, padx=10, pady=6)
        self.symbols_tree.bind("<Double-1>", self._on_symbols_double_click)

    def _on_symbols_double_click(self, event):
        region = self.symbols_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.symbols_tree.identify_row(event.y)
        col_id = self.symbols_tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        try:
            col_index = int(col_id.replace("#", "")) - 1
            col_name = self.symbols_columns[col_index]
        except (ValueError, IndexError):
            return
        sym = row_id
        if col_name == "enabled":
            control.set_symbol_enabled(sym, not control.is_symbol_enabled(sym))
        elif col_name == "lot":
            current = control.get_lot_override(sym) or 0
            new_lot = simpledialog.askfloat(
                "Лот", f"Фиксированный лот для {sym} (0 = авторасчёт по риску):",
                initialvalue=current, minvalue=0, parent=self.root,
            )
            if new_lot is not None:
                control.set_lot_override(sym, new_lot)
        elif col_name == "symbol":
            self._open_price_chart(sym)
        self._refresh_symbols_tab()

    def _open_price_chart(self, symbol: str):
        win = tk.Toplevel(self.root)
        win.title(f"График цены — {symbol}")
        win.geometry("520x320")
        win.configure(bg=self.colors["bg"])
        canvas = tk.Canvas(win, bg=self.colors["bg"], highlightthickness=0)
        canvas.pack(fill="both", expand=True, padx=10, pady=10)

        def redraw():
            if not win.winfo_exists():
                return
            snap = ds.get_snapshot() or {}
            sy = (snap.get("symbols", {}) or {}).get(symbol, {})
            closes = sy.get("recent_closes", []) or []
            canvas.delete("all")
            w = canvas.winfo_width() or 480
            h = canvas.winfo_height() or 280
            if len(closes) < 2:
                canvas.create_text(10, h // 2, anchor="w", text="Копится история цены...", fill=self.colors["muted"])
            else:
                lo, hi = min(closes), max(closes)
                rng = (hi - lo) or 1e-9
                step_x = w / (len(closes) - 1)
                points = []
                for i, v in enumerate(closes):
                    x = i * step_x
                    y = h - ((v - lo) / rng) * (h - 30) - 15
                    points.extend([x, y])
                canvas.create_line(*points, fill=self.colors["profit"], width=2)
                canvas.create_text(35, h - 10, text=f"мин: {lo:.5f}", fill=self.colors["muted"], anchor="w")
                canvas.create_text(35, 10, text=f"макс: {hi:.5f}", fill=self.colors["muted"], anchor="w")
            win.after(3000, redraw)

        redraw()

    def _refresh_symbols_tab(self):
        snap = ds.get_snapshot()
        symbols = snap.get("symbols", {}) if snap else {}

        # Строка показывает РЕЗУЛЬТАТ ОТБОРА, а не просто «сколько пар у
        # брокера»: раз пары выбирает программа, человеку важно видеть, из
        # скольких выбрано и сколько сейчас в работе.
        available = (snap.get("available_symbols", []) if snap else []) or []
        if available != getattr(self, "_available_symbols_cache", None) or symbols:
            self._available_symbols_cache = available
            if symbols and available:
                # Владелец: «записывай там, на каких парах мы работаем».
                # Полный список — строками таблицы ниже, а здесь их перечень
                # одной строкой: его удобно прочитать целиком и скопировать.
                self.available_symbols_var.set(
                    f"В работе {len(symbols)} пар — отобраны программой из "
                    f"{len(available)} доступных у брокера:\n"
                    + ", ".join(sorted(symbols)))
            elif symbols:
                self.available_symbols_var.set(f"В работе {len(symbols)} пар.")
            else:
                self.available_symbols_var.set(
                    "Отбор пар ещё не выполнен (появится, когда бот запущен и подключён к MT5)."
                )

        for item in self.symbols_tree.get_children():
            self.symbols_tree.delete(item)
        for sym, sy in symbols.items():
            enabled_txt = "✓" if sy.get("enabled", True) else "✗"
            lot_txt = sy.get("lot_override") or "авто"
            ai_txt = f"{sy.get('ai_direction') or '-'} ({round((sy.get('ai_confidence') or 0) * 100)}%)"
            risk_warning = sy.get("risk_warning", "") or ""
            self.symbols_tree.insert("", "end", iid=sym,
                tags=("risk_over",) if risk_warning else (),
                values=(
                    enabled_txt, sym, lot_txt,
                    round(sy.get("buy_score", 0), 1), round(sy.get("sell_score", 0), 1),
                    sy.get("regime", "-"), ai_txt, round(sy.get("custom_score", 0), 1),
                    round(sy.get("multi_indicator_score", 0), 1),
                    sy.get("learning_status", "-"),
                    sy.get("paused_until") or "-", sy.get("reject_reason", "-"),
                    risk_warning or "-",
                ))

    # ---- вкладка "Сделки" ------------------------------------------------------
    def _build_tab_positions(self, parent):
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=800, justify="left", text=
                  "Показаны ВСЕ открытые позиции счёта — и этого бота, и открытые "
                  "вручную в терминале MT5 (колонка «Источник»)."
                  ).pack(padx=10, pady=(8, 2), anchor="w")
        cols = ("ticket", "symbol", "type", "volume", "price_open", "price_current", "sl", "tp",
                "profit", "open_time", "source")
        headings = ("Тикет", "Символ", "Тип", "Лот", "Вход", "Текущая", "SL", "TP", "Профит",
                    "Время открытия", "Источник")
        self.positions_tree = ttk.Treeview(parent, columns=cols, show="headings", height=12)
        for col, head in zip(cols, headings):
            self.positions_tree.heading(col, text=head)
            self.positions_tree.column(col, width=85, anchor="center")
        self.positions_tree.pack(fill="both", expand=True, padx=10, pady=6)

        btn_row = ttk.Frame(parent)
        btn_row.pack(fill="x", padx=10, pady=6)
        ttk.Button(btn_row, text="Закрыть выбранную сделку", command=self.close_selected_position).pack(
            side="left")
        ttk.Button(btn_row, text="Закрыть ВСЕ сделки", command=self.close_all_positions).pack(
            side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Закрыть прибыльные", command=self.close_profitable_positions).pack(
            side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Закрыть убыточные", command=self.close_losing_positions).pack(
            side="left", padx=(8, 0))

    def _refresh_positions_tab(self):
        snap = ds.get_snapshot()
        positions = snap.get("positions", []) if snap else []
        for item in self.positions_tree.get_children():
            self.positions_tree.delete(item)
        for p in positions:
            source = "Бот" if p.get("is_bot", True) else "Ручная"
            self.positions_tree.insert("", "end", iid=str(p["ticket"]), values=(
                p["ticket"], p["symbol"], p["type"], p["volume"],
                p["price_open"], p.get("price_current", "-"), p["sl"], p["tp"],
                round(p["profit"], 2), p.get("open_time", "-"), source,
            ))

    def close_selected_position(self):
        sel = self.positions_tree.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "Выбери сделку в таблице.")
            return
        ticket = int(sel[0])
        if messagebox.askyesno(APP_TITLE, f"Закрыть позицию #{ticket} по рынку?"):
            control.request_close(ticket)

    def close_all_positions(self):
        count = len(self.positions_tree.get_children())
        if count == 0:
            messagebox.showinfo(APP_TITLE, "Нет открытых сделок.")
            return
        if messagebox.askyesno(
            APP_TITLE,
            f"Закрыть АБСОЛЮТНО ВСЕ открытые позиции счёта ({count} шт.) по рынку? "
            f"Это касается и сделок бота, и открытых вручную в терминале MT5. "
            f"Действие нельзя отменить.",
        ):
            control.request_close_all()

    def _count_positions_by_profit(self, want_profitable: bool) -> int:
        count = 0
        for iid in self.positions_tree.get_children():
            values = self.positions_tree.item(iid, "values")
            try:
                profit = float(values[8])  # колонка "profit", см. cols в _build_tab_positions
            except (IndexError, ValueError):
                continue
            if (profit >= 0) == want_profitable:
                count += 1
        return count

    def close_profitable_positions(self):
        count = self._count_positions_by_profit(want_profitable=True)
        if count == 0:
            messagebox.showinfo(APP_TITLE, "Нет прибыльных сделок сейчас.")
            return
        if messagebox.askyesno(
            APP_TITLE,
            f"Закрыть ВСЕ прибыльные сейчас позиции счёта ({count} шт.) по рынку? "
            f"Это касается и сделок бота, и открытых вручную в терминале MT5.",
        ):
            control.request_close_profitable()

    def close_losing_positions(self):
        count = self._count_positions_by_profit(want_profitable=False)
        if count == 0:
            messagebox.showinfo(APP_TITLE, "Нет убыточных сделок сейчас.")
            return
        if messagebox.askyesno(
            APP_TITLE,
            f"Закрыть ВСЕ убыточные сейчас позиции счёта ({count} шт.) по рынку? "
            f"Это касается и сделок бота, и открытых вручную в терминале MT5. "
            f"Действие зафиксирует текущий убыток по ним.",
        ):
            control.request_close_losing()

    # ---- вкладка "Лог" -----------------------------------------------------------
    def _build_tab_log(self, parent):
        ttk.Label(parent, text="Журнал этого бота", font=("Segoe UI", 11, "bold")).pack(
            padx=10, pady=(8, 2), anchor="w")
        cols = ("Time", "Event", "Symbol", "Direction", "Price", "SL", "TP", "Lot", "Score", "Profit")
        self.log_tree = ttk.Treeview(parent, columns=cols, show="headings", height=10)
        for col in cols:
            self.log_tree.heading(col, text=col)
            self.log_tree.column(col, width=80, anchor="center")
        self.log_tree.pack(fill="both", expand=False, padx=10, pady=(0, 6))

        ttk.Separator(parent, orient="horizontal").pack(fill="x", padx=10, pady=6)

        top_row = ttk.Frame(parent)
        top_row.pack(fill="x", padx=10)
        ttk.Label(top_row, text="Синхронизация с MetaTrader (все сделки счёта, вкл. ручные)",
                  font=("Segoe UI", 11, "bold")).pack(side="left")
        ttk.Button(top_row, text="Обновить", command=self._refresh_log_tab).pack(side="right")

        self.mt5_history_stats_var = tk.StringVar(value="Синхронизация ещё не выполнялась...")
        ttk.Label(parent, textvariable=self.mt5_history_stats_var, foreground=self.colors["muted"],
                  wraplength=800, justify="left").pack(padx=10, pady=(2, 6), anchor="w")

        mt5_cols = ("ticket", "time", "symbol", "type", "volume", "price", "profit", "source")
        mt5_headings = ("Тикет", "Время", "Символ", "Тип", "Лот", "Цена", "Профит", "Источник")
        self.mt5_history_tree = ttk.Treeview(parent, columns=mt5_cols, show="headings", height=10)
        for col, head in zip(mt5_cols, mt5_headings):
            self.mt5_history_tree.heading(col, text=head)
            self.mt5_history_tree.column(col, width=85, anchor="center")
        self.mt5_history_tree.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _refresh_log_tab(self):
        rows = []
        try:
            path = cfg.LOG_CSV_PATH
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f:
                    reader = list(csv.reader(f, delimiter=";"))
                if len(reader) > 1:
                    data = reader[1:]
                    rows = list(reversed(data[-50:]))
        except Exception:
            pass
        for item in self.log_tree.get_children():
            self.log_tree.delete(item)
        for row in rows:
            self.log_tree.insert("", "end", values=row)

        snap = ds.get_snapshot() or {}
        history = snap.get("mt5_history", [])
        stats = snap.get("mt5_history_stats", {})
        for item in self.mt5_history_tree.get_children():
            self.mt5_history_tree.delete(item)
        for d in history[:100]:
            source = "Бот" if d.get("is_bot", True) else "Ручная"
            self.mt5_history_tree.insert("", "end", values=(
                d["ticket"], d["time"], d["symbol"], d["type"], d["volume"], d["price"],
                d["profit"], source,
            ))
        if stats:
            self.mt5_history_stats_var.set(
                f"За последние {stats.get('days', '-')} дн.: сделок {stats.get('total_trades', 0)}, "
                f"винрейт {stats.get('win_rate', 0)}%, профит-фактор {stats.get('profit_factor', 0)}, "
                f"вал. прибыль {stats.get('gross_profit', 0)}, вал. убыток {stats.get('gross_loss', 0)} "
                f"— данные напрямую из истории MT5 (100% совпадает с брокером)."
            )
        elif not history:
            self.mt5_history_stats_var.set(
                "История MT5 ещё пуста или синхронизация ещё не прошла (обновляется раз в минуту)."
            )

    # ---- вкладка "Equity" ----------------------------------------------------------
    def _build_tab_equity(self, parent):
        self.equity_canvas = tk.Canvas(parent, bg=self.colors["bg"], highlightthickness=0)
        self.equity_canvas.pack(fill="both", expand=True, padx=10, pady=10)

    def _redraw_equity_canvas(self):
        c = self.equity_canvas
        c.delete("all")
        w = c.winfo_width() or 600
        h = c.winfo_height() or 300
        hist = control.get_equity_history()
        if len(hist) < 2:
            c.create_text(10, h // 2, anchor="w", text="Копится история...", fill=self.colors["muted"])
            return
        values = [p["equity"] for p in hist]
        lo, hi = min(values), max(values)
        rng = (hi - lo) or 1
        step_x = w / (len(values) - 1)
        points = []
        for i, v in enumerate(values):
            x = i * step_x
            y = h - ((v - lo) / rng) * (h - 30) - 15
            points.extend([x, y])
        if len(points) >= 4:
            c.create_line(*points, fill=self.colors["profit"], width=2, smooth=False)
        c.create_text(35, h - 10, text=f"мин: {lo:.2f}", fill=self.colors["muted"], anchor="w")
        c.create_text(35, 10, text=f"макс: {hi:.2f}", fill=self.colors["muted"], anchor="w")

    # ---- вкладка "Настройка" (всегда видима — быстрые настройки + ВСЕ параметры) ----
    def _build_tab_config(self, parent):
        """Единая вкладка "Настройка" — видна и в простом, и в продвинутом
        режиме (не прячется), чтобы настройки было невозможно "не найти".
        Сверху — быстрые переключатели (профиль/режим/пауза/звук), ниже —
        полный список input-параметров (как в MQL5-советнике) с прокруткой."""
        parent = self._scrollable(parent)   # боковой ползунок
        sub = ttk.Notebook(parent)
        sub.pack(fill="both", expand=True, padx=4, pady=4)

        quick = ttk.Frame(sub)
        sub.add(quick, text="Быстрый старт")
        self._build_quick_setup(quick)

        # Точная настройка: каждый раздел — своя подвкладка
        self.param_vars = {}
        for title, groups in CONFIG_SECTIONS:
            frame = ttk.Frame(sub)
            sub.add(frame, text=title)
            self._build_tab_params(frame, only_groups=groups)

        extra = ttk.Frame(sub)
        sub.add(extra, text="Профили")
        self._build_tab_profiles_and_context(extra)

    # ---- Быстрая настройка: выбрал режим и торгуешь ------------------------
    QUICK_PRESETS = [
        ("Осторожный", "conservative", "scalping",
         "Меньше сделок, но каждая тщательно отобрана.",
         "риск 0.3% · 1 сделка · совокупный риск 0.5% · порог сигнала высокий"),
        ("Сбалансированный", "balanced", "scalping",
         "Золотая середина. С него стоит начинать.",
         "риск 0.7% · до 2 сделок · совокупный риск 1.8% · порог сигнала средний"),
        ("Активный", "aggressive", "both",
         "Больше сделок и больше риск. Только на демо, пока не проверите.",
         "риск 1.2% · до 5 сделок · совокупный риск 6.5% · ловит и новостные пробои"),
    ]

    def _build_quick_setup(self, parent):
        """Три карточки-режима: один щелчок вместо 118 параметров."""
        ttk.Label(parent, text="Выберите режим торговли",
                  font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=12, pady=(14, 2))
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=820, justify="left",
                  text="Режим задаёт сразу всё: риск на сделку, число одновременных "
                       "сделок, потолок совокупного риска и строгость отбора сигналов. "
                       "Точная настройка каждого параметра — на соседних вкладках, "
                       "но начинать с неё не обязательно.\n\n"
                       "Торговля не останавливается и не ждёт. Выключены ВСЕ паузы: "
                       "дневной порог убытка, лимит просадки, пауза после серии "
                       "убытков, ожидание перед разворотом, пауза вокруг полуночи "
                       "брокера и пауза рядом с новостями. Бот отрабатывает всё "
                       "торговое время.\n\n"
                       "Вместо ожидания работают проверки КАЖДОЙ сделки — они ничего "
                       "не ждут, а просто не берут плохую сделку: обязательный "
                       "стоп-лосс не ближе спреда и шума, риск на сделку, потолок "
                       "совокупного риска, фильтр широкого спреда, защита от скачка "
                       "волатильности и снижение объёма по мере серии неудач. "
                       "Вернуть любую паузу — вкладка «Настройка», разделы «Риск» и "
                       "«Защита»."
                  ).pack(anchor="w", padx=12, pady=(0, 12))

        cards = ttk.Frame(parent)
        cards.pack(fill="x", padx=8)

        self.quick_choice = tk.StringVar(value="balanced")
        current = (control.get_risk_profile() or cfg.RISK_PROFILE).value
        for _, value, _, _, _ in self.QUICK_PRESETS:
            if value == current:
                self.quick_choice.set(value)

        for title, value, mode, summary, details in self.QUICK_PRESETS:
            card = ttk.LabelFrame(cards, text=title)
            card.pack(side="left", fill="both", expand=True, padx=6, pady=4)
            ttk.Radiobutton(card, text="Выбрать", value=value,
                            variable=self.quick_choice).pack(anchor="w", padx=8, pady=(6, 2))
            ttk.Label(card, text=summary, wraplength=230, justify="left").pack(
                anchor="w", padx=8, pady=(0, 4))
            ttk.Label(card, text=details, wraplength=230, justify="left",
                      foreground=self.colors["muted"], font=("Segoe UI", 8)).pack(
                anchor="w", padx=8, pady=(0, 8))

        # ---- Стратегия: чем именно торгуем ----
        ttk.Label(parent, text="Стратегия", font=("Segoe UI", 12, "bold")).pack(
            anchor="w", padx=12, pady=(16, 2))
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=820, justify="left",
                  text="Режим выше задаёт РИСК, стратегия — ЧТО считать сигналом. "
                       "Ни одна стратегия не отключает защиты счёта: обязательный "
                       "стоп-лосс, риск на сделку, потолок совокупного риска и "
                       "снижение объёма по серии убытков работают всегда."
                  ).pack(anchor="w", padx=12, pady=(0, 6))

        strat_row = ttk.Frame(parent)
        strat_row.pack(fill="x", padx=12)
        self.strategy_combo = ttk.Combobox(strat_row, values=strategies_mod.titles(),
                                           state="readonly", width=28)
        current_key = getattr(cfg, "ACTIVE_STRATEGY", "") or ""
        current_strategy = strategies_mod.by_key(current_key) or strategies_mod.STRATEGIES[0]
        self.strategy_combo.set(current_strategy.title)
        self.strategy_combo.pack(side="left")
        self.strategy_combo.bind("<<ComboboxSelected>>", self._on_strategy_pick)
        ttk.Button(strat_row, text="Применить стратегию",
                   command=self._apply_strategy).pack(side="left", padx=8)

        self.strategy_desc = ttk.Label(parent, foreground=self.colors["muted"], wraplength=820,
                                       justify="left", text="")
        self.strategy_desc.pack(anchor="w", padx=12, pady=(6, 0))
        self._on_strategy_pick()

        actions = ttk.Frame(parent)
        actions.pack(fill="x", padx=12, pady=(14, 6))
        ttk.Button(actions, text="Применить режим",
                   command=self._apply_quick_preset).pack(side="left")
        ttk.Button(actions, text="Применить и запустить бота",
                   command=self._apply_quick_and_start).pack(side="left", padx=8)

        self.quick_status = ttk.Label(parent, text="", foreground=self.colors["profit"],
                                      wraplength=820, justify="left")
        self.quick_status.pack(anchor="w", padx=12, pady=(4, 10))

        ttk.Separator(parent, orient="horizontal").pack(fill="x", padx=10, pady=8)

        # Ниже — прежние быстрые переключатели (профиль, режим, пауза, звук)
        self._build_tab_settings(parent)

    def _on_strategy_pick(self, _event=None):
        strategy = strategies_mod.by_title(self.strategy_combo.get())
        if strategy is not None:
            self.strategy_desc.configure(text=strategies_mod.describe(strategy))

    def _apply_strategy(self):
        """Записывает параметры стратегии в config.py и применяет на лету."""
        strategy = strategies_mod.by_title(self.strategy_combo.get())
        if strategy is None:
            return
        params = strategies_mod.safe_params(strategy)
        if not messagebox.askyesno(
            APP_TITLE,
            f"Применить стратегию «{strategy.title}»?\n\n"
            f"Будет изменено параметров: {len(params)}.\n"
            "Настройки риска (размер сделки, лимиты убытка) НЕ затрагиваются."
        ):
            return
        try:
            for key, value in params.items():
                literal = str(value) if isinstance(value, bool) else repr(value)
                _write_config_value(key, literal)
            # Ключ активной стратегии: по нему торговый цикл выбирает,
            # чьё мнение подмешивать к оценке сигнала
            _write_config_value("ACTIVE_STRATEGY", repr(strategy.key))
            _write_config_value("USE_STRATEGY_SIGNAL", "True")
            _reload_cfg()
        except Exception as e:  # noqa: BLE001
            log.exception("Не удалось применить стратегию")
            messagebox.showerror(APP_TITLE, f"Не удалось применить стратегию: {e}")
            return
        self.quick_status.configure(
            text=f"Стратегия «{strategy.title}» применена: {strategy.idea}. "
                 f"Изменено параметров: {len(params)}.", foreground=self.colors["profit"])
        messagebox.showinfo(APP_TITLE,
                            f"Стратегия «{strategy.title}» применена.\n\n"
                            "Значения видны на вкладках точной настройки — "
                            "их можно донастроить вручную.")

    def _quick_preset_by_value(self, value):
        for preset in self.QUICK_PRESETS:
            if preset[1] == value:
                return preset
        return self.QUICK_PRESETS[1]

    def _apply_quick_preset(self):
        """Ставит профиль риска и режим торговли одним действием."""
        title, value, mode, _, details = self._quick_preset_by_value(self.quick_choice.get())
        try:
            for label, pv in PROFILE_OPTIONS:
                if pv == value:
                    self.profile_combo.set(label)
            self._apply_profile()
            for label, mv in MODE_OPTIONS:
                if mv == mode:
                    self.mode_combo.set(label)
            self._apply_mode()
        except Exception as e:  # noqa: BLE001
            log.exception("Не удалось применить быстрый режим")
            messagebox.showerror(APP_TITLE, f"Не удалось применить режим: {e}")
            return False
        self.quick_status.configure(
            text=f"Режим «{title}» применён: {details}", foreground=self.colors["profit"])
        return True

    def _apply_quick_and_start(self):
        if not self._apply_quick_preset():
            return
        try:
            self.start_bot()
            self.quick_status.configure(
                text=self.quick_status.cget("text") + "  ·  бот запущен",
                foreground=self.colors["profit"])
        except Exception as e:  # noqa: BLE001
            log.exception("Не удалось запустить бота из быстрой настройки")
            messagebox.showerror(APP_TITLE, f"Режим применён, но бот не запустился: {e}")

    def _build_tab_settings(self, parent):
        parent = self._scrollable(parent)   # боковой ползунок
        pad = {"padx": 10, "pady": 8}

        ttk.Label(parent, text="Профиль риска", font=("Segoe UI", 11, "bold")).pack(anchor="w", **pad)
        self.profile_combo = ttk.Combobox(parent, values=[label for label, _ in PROFILE_OPTIONS],
                                           state="readonly", width=30)
        self.profile_combo.pack(anchor="w", padx=10)
        current_profile = (control.get_risk_profile() or cfg.RISK_PROFILE).value
        for label, value in PROFILE_OPTIONS:
            if value == current_profile:
                self.profile_combo.set(label)
        ttk.Button(parent, text="Применить профиль", command=self._apply_profile).pack(anchor="w", **pad)

        ttk.Label(parent, text="Режим торговли", font=("Segoe UI", 11, "bold")).pack(anchor="w", **pad)
        self.mode_combo = ttk.Combobox(parent, values=[label for label, _ in MODE_OPTIONS],
                                        state="readonly", width=30)
        self.mode_combo.pack(anchor="w", padx=10)
        current_mode = (control.get_trading_mode() or cfg.TRADING_MODE).value
        for label, value in MODE_OPTIONS:
            if value == current_mode:
                self.mode_combo.set(label)
        ttk.Button(parent, text="Применить режим", command=self._apply_mode).pack(anchor="w", **pad)

        self.pause_btn = ttk.Button(parent, text="Пауза (новые сделки)", command=self._toggle_pause)
        self.pause_btn.pack(anchor="w", **pad)

        self.sound_var = tk.BooleanVar(value=getattr(cfg, "USE_SOUND_NOTIFICATIONS", True))
        ttk.Checkbutton(parent, text="Звук + всплывающие уведомления о сделках", variable=self.sound_var,
                        command=self._toggle_sound).pack(anchor="w", **pad)

    def _toggle_sound(self):
        _write_config_value("USE_SOUND_NOTIFICATIONS", str(self.sound_var.get()))
        try:
            _reload_cfg()
        except Exception:
            pass

    def _apply_profile(self):
        label = self.profile_combo.get()
        value = next((v for l, v in PROFILE_OPTIONS if l == label), None)
        if value:
            control.set_risk_profile(cfg.RiskProfile(value))
            messagebox.showinfo(APP_TITLE, f"Профиль риска изменён на «{label}».")

    def _apply_mode(self):
        label = self.mode_combo.get()
        value = next((v for l, v in MODE_OPTIONS if l == label), None)
        if value:
            control.set_trading_mode(cfg.TradingMode(value))
            messagebox.showinfo(APP_TITLE, f"Режим торговли изменён на «{label}».")

    def _toggle_pause(self):
        control.set_paused(not control.is_paused())

    # ---- вкладка "Параметры" (продвинутый режим — все input-параметры, как в советнике) ----
    def _build_tab_profiles_and_context(self, parent):
        """Редакторы профилей риска и корреляций — на отдельной подвкладке.

        Раньше они были внутри списка параметров. После разбивки настроек на
        разделы список строится несколько раз, и редакторы задваивались бы.
        """
        ttk.Label(parent, text="Профили риска и связи между инструментами",
                  font=("Segoe UI", 12, "bold")).pack(padx=10, pady=(10, 6), anchor="w")

        outer = ttk.Frame(parent)
        outer.pack(fill="both", expand=True, padx=6, pady=4)
        canvas = tk.Canvas(outer, bg=self.colors["bg"], highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # ---- Профили риска — детальные параметры (RISK_PROFILES) ----
        profile_box = ttk.LabelFrame(inner, text="Профиль риска — детальные параметры")
        profile_box.pack(fill="x", padx=6, pady=(4, 10))
        ttk.Label(profile_box, text="Профиль:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        self.param_profile_combo = ttk.Combobox(
            profile_box, values=[label for label, _ in PROFILE_OPTIONS], state="readonly", width=24)
        self.param_profile_combo.set(PROFILE_OPTIONS[1][0])
        self.param_profile_combo.grid(row=0, column=1, sticky="w", padx=6, pady=4)
        self.param_profile_combo.bind("<<ComboboxSelected>>", lambda e: self._load_profile_fields())

        self.profile_field_vars = {}
        profile_fields_frame = ttk.Frame(profile_box)
        profile_fields_frame.grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=4)
        self._build_profile_fields(profile_fields_frame)
        self._load_profile_fields()


        # ---- Контекст рынка — корреляции по инструментам (MARKET_CONTEXT) ----
        context_box = ttk.LabelFrame(inner, text="Контекст рынка — корреляции по инструментам (до 3 на символ)")
        context_box.pack(fill="x", padx=6, pady=(4, 10))
        ttk.Label(context_box, foreground=self.colors["muted"], wraplength=760, justify="left", text=
                  "Для каждого торгуемого символа можно задать до 3 коррелирующих инструментов "
                  "(например, индекс доллара для золота). Пустое поле = слот не используется. "
                  "Работает, только если включён общий флаг «Учитывать коррелирующие инструменты» "
                  "в разделе «Контекст рынка» ниже."
                  ).pack(anchor="w", padx=6, pady=(4, 6))
        self.context_symbol_vars = {}
        for sym in cfg.SYMBOLS:
            row = ttk.Frame(context_box)
            row.pack(fill="x", padx=6, pady=2)
            ttk.Label(row, text=sym, width=12).pack(side="left")
            slots = list(cfg.MARKET_CONTEXT.get(sym, []))
            slot_vars = []
            for slot_i in range(3):
                sym_var = tk.StringVar(value=slots[slot_i][0] if slot_i < len(slots) else "")
                corr_var = tk.StringVar(value=slots[slot_i][1] if slot_i < len(slots) else "positive")
                ttk.Entry(row, textvariable=sym_var, width=10).pack(side="left", padx=(4, 2))
                ttk.Combobox(row, textvariable=corr_var, values=["positive", "negative"],
                             state="readonly", width=9).pack(side="left", padx=(0, 8))
                slot_vars.append((sym_var, corr_var))
            self.context_symbol_vars[sym] = slot_vars

        # ---- Плоские параметры, сгруппированные по разделам (как в config.py) ----

    def _scrollable(self, parent):
        """Возвращает рамку с вертикальной прокруткой. Складывать содержимое
        нужно уже в неё. Нужна там, где блоков заведомо больше, чем помещается
        в окно: без прокрутки нижние блоки не видны вовсе, а не «обрезаны»."""
        outer = ttk.Frame(parent)
        outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer, bg=self.colors["bg"], highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        window = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        # Содержимое тянется по ширине окна: иначе блоки с fill="x" остались бы
        # шириной в один символ.
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(window, width=e.width))
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        return inner

    def _build_tab_params(self, parent, only_groups=None):
        """only_groups=None — все параметры; список — только эти группы."""
        title = "Точная настройка" if only_groups is None else "Параметры раздела"
        ttk.Label(parent, text=title,
                  font=("Segoe UI", 12, "bold")).pack(padx=10, pady=(10, 2), anchor="w")
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=800, justify="left", text=
                  "Здесь можно вручную выставить КАЖДЫЙ параметр торговой логики — так же, "
                  "как input-параметры MQL5-советника. «Сохранить» применяет изменения сразу, "
                  "бот подхватит их на лету, без перезапуска. Единственное, что сюда не входит — "
                  "корреляции MARKET_CONTEXT редактируются отдельно ниже, а брокер/новости/AI-ключи "
                  "— на своих вкладках."
                  ).pack(padx=10, pady=(0, 8), anchor="w")

        outer = ttk.Frame(parent)
        outer.pack(fill="both", expand=True, padx=6, pady=4)
        canvas = tk.Canvas(outer, bg=self.colors["bg"], highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        groups = {}
        for key, ptype, group, label, choices in ADVANCED_PARAMS:
            if only_groups is not None and group not in only_groups:
                continue
            groups.setdefault(group, []).append((key, ptype, label, choices))

        # param_vars общий на все подвкладки: кнопка "Сохранить" на любой из
        # них записывает ВСЕ изменённые поля, а не только своего раздела
        if not hasattr(self, "param_vars"):
            self.param_vars = {}
        for group, items in groups.items():
            box = ttk.LabelFrame(inner, text=group)
            box.pack(fill="x", padx=6, pady=4)
            for row_i, (key, ptype, label, choices) in enumerate(items):
                ttk.Label(box, text=label, wraplength=340, justify="left").grid(
                    row=row_i, column=0, sticky="w", padx=6, pady=3)
                current = param_current_value(key)
                if ptype == "bool":
                    var = tk.BooleanVar(value=bool(current))
                    ttk.Checkbutton(box, variable=var).grid(row=row_i, column=1, sticky="w", padx=6, pady=3)
                elif ptype == "choice":
                    var = tk.StringVar(value=str(current))
                    ttk.Combobox(box, textvariable=var, values=choices, state="readonly", width=14).grid(
                        row=row_i, column=1, sticky="w", padx=6, pady=3)
                elif ptype == "secret":
                    # Ключ в интерфейс не выводим: показываем заглушку, если он
                    # уже сохранён. Пустое поле при сохранении = не менять ключ.
                    var = tk.StringVar(value=SECRET_PLACEHOLDER if current else "")
                    ttk.Entry(box, textvariable=var, width=44, show="*").grid(
                        row=row_i, column=1, sticky="w", padx=6, pady=3)
                else:
                    var = tk.StringVar(value=str(current))
                    ttk.Entry(box, textvariable=var, width=18).grid(row=row_i, column=1, sticky="w", padx=6, pady=3)
                self.param_vars[key] = (ptype, var)

                # Кнопка «?» рядом с каждым параметром: что делает, значение по
                # умолчанию и что будет, если изменить. Справка лежит в
                # param_help.py, здесь только показ.
                if param_help.has_help(key):
                    ttk.Button(box, text="?", width=3,
                               command=lambda k=key, l=label: self._show_param_help(k, l)).grid(
                        row=row_i, column=2, sticky="w", padx=(2, 6), pady=3)

        btn_frame = ttk.Frame(inner)
        btn_frame.pack(fill="x", padx=6, pady=14)
        ttk.Button(btn_frame, text="Обновить из файла", command=self.reload_advanced_params).grid(
            row=0, column=1, padx=4)

    def _show_param_help(self, key: str, label: str):
        """Окно справки по одному параметру."""
        window = tk.Toplevel(self.root)
        window.title(f"Справка: {key}")
        window.geometry("640x520")
        window.configure(bg=self.colors["bg"])

        ttk.Label(window, text=label, font=("Segoe UI", 11, "bold"),
                  wraplength=600, justify="left").pack(anchor="w", padx=14, pady=(12, 2))
        ttk.Label(window, text=key, foreground=self.colors["muted"]).pack(anchor="w", padx=14)

        item = param_help.entry(key)
        # Полоса прокрутки обязательна: у части параметров есть ещё и блок
        # «Внимание», и без прокрутки он оказывался за нижним краем окна —
        # то есть предупреждение было не видно именно там, где оно важнее всего.
        body_frame = ttk.Frame(window)
        body_frame.pack(fill="both", expand=True, padx=14, pady=10)
        body = tk.Text(body_frame, wrap="word", height=16, bg=self.colors["card"], fg=self.colors["fg"],
                       relief="flat", padx=10, pady=8)
        body_scroll = ttk.Scrollbar(body_frame, command=body.yview)
        body.configure(yscrollcommand=body_scroll.set)
        body.pack(side="left", fill="both", expand=True)
        body_scroll.pack(side="right", fill="y")

        body.tag_configure("head", foreground=self.colors["accent"], spacing1=6)
        body.tag_configure("warn", foreground=self.colors["warning"])
        body.insert("end", item["what"] + "\n\n")

        default = param_help.default_of(key)
        if default is not None:
            body.insert("end", "Значение по умолчанию\n", "head")
            body.insert("end", f"{default!r}\n\n")
        if item["more"]:
            body.insert("end", "Если увеличить / включить\n", "head")
            body.insert("end", item["more"] + "\n\n")
        if item["less"]:
            body.insert("end", "Если уменьшить / выключить\n", "head")
            body.insert("end", item["less"] + "\n\n")
        if item["warn"]:
            body.insert("end", "Внимание\n", "head")
            body.insert("end", item["warn"] + "\n", "warn")
        body.configure(state="disabled")

        ttk.Button(window, text="Закрыть", command=window.destroy).pack(pady=(0, 12))

    def _build_profile_fields(self, parent):
        for row_i, (field, ftype, label) in enumerate(RISK_PROFILE_FIELD_DEFS):
            ttk.Label(parent, text=label, wraplength=260, justify="left").grid(
                row=row_i, column=0, sticky="w", padx=6, pady=2)
            if ftype == "bool":
                var = tk.BooleanVar(value=False)
                ttk.Checkbutton(parent, variable=var).grid(row=row_i, column=1, sticky="w", padx=6, pady=2)
            else:
                var = tk.StringVar(value="")
                ttk.Entry(parent, textvariable=var, width=18).grid(row=row_i, column=1, sticky="w", padx=6, pady=2)
            self.profile_field_vars[field] = (ftype, var)

    def _current_profile_enum(self):
        label = self.param_profile_combo.get()
        value = next((v for l, v in PROFILE_OPTIONS if l == label), "balanced")
        return cfg.RiskProfile(value)

    def _load_profile_fields(self):
        profile_enum = self._current_profile_enum()
        params = cfg.RISK_PROFILES.get(profile_enum, {})
        for field, (ftype, var) in self.profile_field_vars.items():
            value = params.get(field, "")
            if ftype == "bool":
                var.set(bool(value))
            else:
                var.set(str(value))

    def save_profile_fields(self, silent: bool = False):
        profile_enum = self._current_profile_enum()
        new_params = {}
        errors = []
        for field, (ftype, var) in self.profile_field_vars.items():
            raw = var.get()
            try:
                if ftype == "bool":
                    new_params[field] = bool(raw)
                elif ftype == "int":
                    new_params[field] = int(raw)
                elif ftype == "float":
                    new_params[field] = float(raw)
                else:
                    new_params[field] = str(raw)
            except (TypeError, ValueError):
                errors.append(field)
        if errors:
            messagebox.showerror(APP_TITLE, "Некорректные значения: " + ", ".join(errors))
            return
        try:
            all_profiles = dict(cfg.RISK_PROFILES)
            all_profiles[profile_enum] = new_params
            _write_config_block("RISK_PROFILES", _format_risk_profiles(all_profiles))
            _reload_cfg()
            if not silent:
                messagebox.showinfo(APP_TITLE, f"Профиль «{new_params.get('name', profile_enum.value)}» сохранён.")
        except Exception as e:
            log.exception("Не удалось сохранить профиль риска: %s", e)
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить: {e}")

    def save_market_context(self, silent: bool = False):
        try:
            new_context = dict(cfg.MARKET_CONTEXT)
            for sym, slot_vars in self.context_symbol_vars.items():
                slots = []
                for sym_var, corr_var in slot_vars:
                    ctx_sym = sym_var.get().strip()
                    if ctx_sym:
                        slots.append((ctx_sym, corr_var.get()))
                new_context[sym] = slots
            _write_config_block("MARKET_CONTEXT", repr(new_context))
            _reload_cfg()
            if not silent:
                messagebox.showinfo(APP_TITLE, "Контекст рынка сохранён.")
        except Exception as e:
            log.exception("Не удалось сохранить контекст рынка: %s", e)
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить: {e}")

    def save_advanced_params(self, silent: bool = False):
        new_values = {}
        errors = []
        for key, (ptype, var) in self.param_vars.items():
            raw = var.get()
            # Пустое числовое поле — не ошибка ввода, а параметр, которого ещё
            # не было в вашем config.py. Подставляем значение по умолчанию из
            # config.py.example вместо того, чтобы отказывать в сохранении
            # ВСЕХ настроек из-за одного нового поля. Секреты не трогаем:
            # там пустота означает «ключ не меняли».
            if ptype in ("int", "float") and str(raw).strip() == "":
                default = param_help.default_of(key)
                if default is not None:
                    raw = default
            try:
                if ptype == "bool":
                    new_values[key] = bool(raw)
                elif ptype == "int":
                    new_values[key] = int(raw)
                elif ptype == "float":
                    new_values[key] = float(raw)
                else:
                    new_values[key] = str(raw)
            except (TypeError, ValueError):
                errors.append(key)
        if errors:
            messagebox.showerror(
                APP_TITLE,
                "Некорректные значения в полях: " + ", ".join(errors) +
                "\n\nВ этих полях ждут число. Нажмите «?» рядом с полем — там "
                "написано, что означает параметр и какое значение стандартное.")
            return

        tf_rank = {"M1": 1, "M5": 2, "M15": 3, "M30": 4, "H1": 5, "H4": 6, "D1": 7}
        tf = new_values.get("TIMEFRAME")
        trend_tf = new_values.get("TREND_TIMEFRAME")
        if tf and trend_tf and tf_rank.get(trend_tf, 0) <= tf_rank.get(tf, 0):
            messagebox.showerror(APP_TITLE, "Старший таймфрейм тренда должен быть СТАРШЕ рабочего таймфрейма.")
            return

        try:
            session_pw = control.get_session_password() or ""
            salt = getattr(cfg, "SECURITY_SALT", "") or ""
            for key, value in new_values.items():
                ptype = self.param_vars.get(key, ("", None))[0]
                if ptype == "secret":
                    text = str(value).strip()
                    if not text:
                        continue  # поле оставили пустым — старый ключ не трогаем
                    if text == SECRET_PLACEHOLDER:
                        continue  # ключ не меняли, в поле стоит заглушка
                    text = secure_store.protect_secret(text, session_pw, salt)
                    _write_config_value(key, repr(text))
                    continue
                literal = str(value) if isinstance(value, bool) else repr(value)
                _write_config_value(key, literal)
            _reload_cfg()
            if not silent:
                messagebox.showinfo(APP_TITLE, "Параметры сохранены и применены.")
        except Exception as e:
            log.exception("Не удалось сохранить расширенные параметры: %s", e)
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить: {e}")

    def reload_advanced_params(self):
        for key, (ptype, var) in self.param_vars.items():
            current = param_current_value(key)
            if ptype == "bool":
                var.set(bool(current))
            elif ptype == "secret":
                # Ключ в поле не возвращаем — только заглушку, как при
                # построении вкладки. Иначе «Обновить из файла» выкладывал бы
                # секрет на экран.
                var.set(SECRET_PLACEHOLDER if current else "")
            else:
                var.set(str(current))
        self._load_profile_fields()
        for sym, slot_vars in self.context_symbol_vars.items():
            slots = list(cfg.MARKET_CONTEXT.get(sym, []))
            for slot_i, (sym_var, corr_var) in enumerate(slot_vars):
                sym_var.set(slots[slot_i][0] if slot_i < len(slots) else "")
                corr_var.set(slots[slot_i][1] if slot_i < len(slots) else "positive")

    # ---- вкладка "Новости" ---------------------------------------------------------
    def _build_tab_news(self, parent):
        # Боковой ползунок: владелец просил не растягивать окно
        # каждый раз — блоков здесь больше, чем помещается.
        parent = self._scrollable(parent)
        pad = {"padx": 10, "pady": 6}

        ttk.Label(parent, text="Предстоящие события", font=("Segoe UI", 12, "bold")).pack(**pad)
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=800, justify="left", text=
                  "Полный список новостей. Включить и выключить источники — на вкладке "
                  "«Источники». Расписание работы бота — на вкладке «Календарь»."
                  ).pack(anchor="w", padx=10)

        btn_frame = ttk.Frame(parent)
        btn_frame.pack(anchor="w", **pad)
        ttk.Button(btn_frame, text="Обновить календарь",
                   command=self.refresh_news_tab).grid(row=0, column=0)
        ttk.Button(btn_frame, text="Проверить и починить источник",
                   command=self.fix_news_source).grid(row=0, column=1, padx=6)
        ttk.Button(btn_frame, text="Почему нет новостных сделок",
                   command=self.explain_news_trading).grid(row=0, column=2)

        # Состояние всей цепочки новостей одной строкой. Без неё понять, почему
        # новостной режим молчит, было нельзя: бот писал «свежего пробоя нет»
        # и когда пробоя правда не было, и когда календаря не существует вовсе.
        self.news_source_var = tk.StringVar(value="Источник новостей: проверяю...")
        ttk.Label(parent, textvariable=self.news_source_var, wraplength=800,
                  justify="left", font=("Segoe UI", 9, "bold")).pack(anchor="w", **pad)

        self.news_status_var = tk.StringVar(value="")
        ttk.Label(parent, textvariable=self.news_status_var, foreground=self.colors["muted"], wraplength=800,
                  justify="left").pack(anchor="w", **pad)

        cols = ("time", "left", "currency", "event", "impact", "actual", "estimate", "prev")
        headings = ("Время", "Осталось", "Валюта", "Событие", "Важность", "Факт", "Прогноз", "Пред.")
        self.news_tree = ttk.Treeview(parent, columns=cols, show="headings", height=10)
        for col, head in zip(cols, headings):
            self.news_tree.heading(col, text=head)
            self.news_tree.column(col, width=90, anchor="center")
        self.news_tree.column("event", width=220, anchor="w")
        self.news_tree.pack(fill="both", expand=True, padx=10, pady=6)

    # ---- синхронизация с облаком (главная страница) ----------------------------------
    def _refresh_sync_status(self):
        """Что стоит сейчас — до всякого обращения к сети."""
        if not updater.enabled():
            self.sync_status_var.set(
                "Выключено. Включить и указать репозиторий — на вкладке «Система».")
            return
        installed = updater.current_version()
        repo = updater.repo() or "репозиторий не задан"
        self.sync_status_var.set(
            f"Источник: {repo}. Установленная версия: {installed or 'ещё не отмечена'}.")

    def sync_from_cloud(self):
        """Кнопка на главной: проверить, что изменилось, и предложить обновить.

        Отдельная от вкладки «Система» только точкой входа — работу делает тот
        же updater. Дублировать логику ради второй кнопки нельзя: разошлись бы
        со временем."""
        if not updater.enabled():
            messagebox.showinfo(
                APP_TITLE,
                "Синхронизация выключена.\n\nВключите её на вкладке «Система» и "
                "укажите репозиторий — тогда программа будет забирать изменения "
                "сама, без переустановки.")
            return
        self.sync_status_var.set("Смотрю, что изменилось в облаке...")
        self.check_updates(silent=False)

    def show_changes(self):
        """Показывает последние изменения в репозитории — просто посмотреть,
        ничего не устанавливая."""
        if not updater.enabled() or not updater.repo():
            messagebox.showinfo(APP_TITLE,
                                "Сначала укажите репозиторий на вкладке «Система».")
            return

        def worker():
            entries, error = updater.recent_changes()
            self.root.after(0, lambda: self._show_changes_window(entries, error))

        self.sync_status_var.set("Загружаю список изменений...")
        threading.Thread(target=worker, daemon=True, name="cloud-changes").start()

    def _show_changes_window(self, entries, error):
        if error:
            self.sync_status_var.set(error)
            messagebox.showwarning(APP_TITLE, error)
            return

        self._refresh_sync_status()
        window = tk.Toplevel(self.root)
        window.title("Что нового в облаке")
        window.geometry("820x460")

        ttk.Label(window, text="Последние изменения в коде",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(10, 2))

        installed = updater.current_version()
        ttk.Label(window, foreground=self.colors["muted"], wraplength=780, justify="left",
                  text=("Зелёным помечено то, что уже установлено у вас. "
                        "Выше него — изменения, которых пока нет.")
                  ).pack(anchor="w", padx=12, pady=(0, 6))

        cols = ("when", "what", "rev")
        tree = ttk.Treeview(window, columns=cols, show="headings", height=16)
        for col, head, width in (("when", "Когда", 130), ("what", "Что изменилось", 560),
                                 ("rev", "Версия", 100)):
            tree.heading(col, text=head)
            tree.column(col, width=width, anchor="w" if col == "what" else "center")
        tree.pack(fill="both", expand=True, padx=12, pady=6)
        tree.tag_configure("installed", foreground=self.colors["profit"])

        for item in entries:
            tag = ("installed",) if item["revision"] == installed else ()
            tree.insert("", "end", tags=tag,
                        values=(item["date"], item["message"], item["revision"]))

        if not entries:
            self.sync_status_var.set("Список изменений пуст.")

        ttk.Button(window, text="Закрыть", command=window.destroy).pack(pady=(0, 10))

    # ---- вкладка "Система" -----------------------------------------------------------
    def export_history(self):
        """Выгрузить настоящие свечи брокера в файлы — для проверки стратегии.

        Работает только при подключённом терминале: свечи отдаёт он. Тяжёлую
        часть уносим в отдельный поток, иначе окно замрёт на минуту-другую и
        человек решит, что программа повисла."""
        self.history_export_var.set("Выгружаю... это может занять минуту.")

        def работа():
            try:
                import history_export
                отчёты = history_export.export_all(
                    progress=lambda t: self.root.after(
                        0, lambda: self.history_export_var.set(t)))
                текст = history_export.describe(отчёты)
            except Exception as e:  # noqa: BLE001
                log.exception("Выгрузка истории не удалась")
                текст = (f"Не получилось: {e}. Проверьте, что терминал "
                         f"подключён (вкладка «Брокер»).")
            self.root.after(0, lambda: self.history_export_var.set(текст))

        threading.Thread(target=работа, daemon=True).start()

    def show_mt5_terminal(self):
        """Вернуть окно терминала на экран.

        Нужна обязательно: спрятанное окно исчезает и с панели задач, мышью
        его не достать. Без этой кнопки человек остался бы без MetaTrader
        совсем и решил бы, что программа его сломала."""
        count = mt5c.show_terminal()
        self.terminal_window_var.set(
            "Окно терминала показано." if count
            else "Терминал не найден — он запущен?")

    def hide_mt5_terminal(self):
        count = mt5c.hide_terminal()
        self.terminal_window_var.set(
            "Окно терминала скрыто, торговля продолжается." if count
            else "Терминал не найден — он запущен?")

    def _build_tab_system(self, parent):
        """Проверка компьютера, мост для советников и обновление — в одном
        месте. Всё, что отвечает на вопрос «почему не работает» и «как
        получить свежую версию»."""
        pad = {"padx": 12, "pady": 4}

        # Блоков на этой вкладке больше, чем помещается в окно: без прокрутки
        # нижние (журнал в облаке, таблица проверок) просто не видны, и человек
        # решает, что их нет.
        parent = self._scrollable(parent)

        ttk.Label(parent, text="Состояние системы",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", **pad)

        # ---------- Окно терминала ----------
        term = ttk.LabelFrame(parent, text=" Терминал MetaTrader ")
        term.pack(fill="x", padx=12, pady=(6, 4))
        ttk.Label(term, foreground=self.colors["muted"], wraplength=780,
                  justify="left", text=
                  "Программа сама запускает терминал и сама входит в счёт — "
                  "вводить что-либо в MetaTrader не нужно. Его окно спрятано, "
                  "чтобы не мешало; терминал при этом работает полностью.\n"
                  "Совсем без терминала работать нельзя: именно он держит связь "
                  "с брокером и исполняет приказы, а его протокол закрытый."
                  ).pack(anchor="w", padx=8, pady=(4, 2))
        row = ttk.Frame(term)
        row.pack(anchor="w", padx=8, pady=(0, 6))
        ttk.Button(row, text="Показать терминал",
                   command=self.show_mt5_terminal).pack(side="left")
        ttk.Button(row, text="Спрятать терминал",
                   command=self.hide_mt5_terminal).pack(side="left", padx=6)
        self.terminal_window_var = tk.StringVar(value="")
        ttk.Label(row, textvariable=self.terminal_window_var,
                  foreground=self.colors["muted"]).pack(side="left", padx=8)

        # ---------- Выгрузка истории для проверки ----------
        hist = ttk.LabelFrame(parent, text=" История для проверки стратегии ")
        hist.pack(fill="x", padx=12, pady=(6, 4))
        ttk.Label(hist, foreground=self.colors["muted"], wraplength=780,
                  justify="left", text=
                  "Чтобы проверить стратегию на прошлом, нужны настоящие свечи "
                  "вашего брокера. Программа выгрузит их сама — нажмите кнопку "
                  "и подождите. Файлы лягут в папку history рядом с программой.\n"
                  "Выгружается EURUSD и XAUUSD на M5. Последняя свеча в файл не "
                  "попадает: она ещё не закрыта, и в расчёт идти не может."
                  ).pack(anchor="w", padx=8, pady=(4, 2))
        hrow = ttk.Frame(hist)
        hrow.pack(anchor="w", padx=8, pady=(0, 6))
        ttk.Button(hrow, text="Выгрузить историю",
                   command=self.export_history).pack(side="left")
        self.history_export_var = tk.StringVar(value="")
        ttk.Label(hrow, textvariable=self.history_export_var,
                  foreground=self.colors["muted"], wraplength=520,
                  justify="left").pack(side="left", padx=8)

        # ---------- Мост для советников ----------
        br = ttk.LabelFrame(parent, text=" Мост для советников MetaTrader ")
        br.pack(fill="x", padx=12, pady=(6, 4))

        ttk.Label(br, foreground=self.colors["muted"], wraplength=780, justify="left", text=
                  "Советники спрашивают у моста режим рынка. Мост встроен в программу — "
                  "отдельно запускать нечего. Слушает только 127.0.0.1, наружу не "
                  "открывается. Его ответ может лишь УМЕНЬШИТЬ объём сделки."
                  ).grid(row=0, column=0, columnspan=3, sticky="w", padx=8, pady=(4, 2))

        self.bridge_enabled_var = tk.BooleanVar(value=getattr(cfg, "BRIDGE_ENABLED", False))
        ttk.Checkbutton(br, text="Включить мост", variable=self.bridge_enabled_var).grid(
            row=1, column=0, sticky="w", padx=8, pady=3)

        ttk.Label(br, text="Порт:").grid(row=1, column=1, sticky="e", padx=4)
        self.bridge_port_var = tk.StringVar(value=str(getattr(cfg, "BRIDGE_PORT", 8080)))
        ttk.Entry(br, textvariable=self.bridge_port_var, width=8).grid(
            row=1, column=2, sticky="w", padx=4)

        self.bridge_status_var = tk.StringVar(value="")
        ttk.Label(br, textvariable=self.bridge_status_var, foreground=self.colors["muted"],
                  wraplength=780, justify="left").grid(
            row=2, column=0, columnspan=3, sticky="w", padx=8, pady=(2, 6))

        # ---------- Обновление ----------
        up = ttk.LabelFrame(parent, text=" Обновление из GitHub ")
        up.pack(fill="x", padx=12, pady=(6, 4))

        ttk.Label(up, foreground=self.colors["muted"], wraplength=780, justify="left", text=
                  "Правки кода приезжают с GitHub сами — переустанавливать программу не "
                  "нужно. Кнопка «Обновить всё» скачивает и ставит: советники в "
                  "MetaTrader (сразу, без перезапуска), файлы самой программы и новые "
                  "настройки. Перезапуск нужен только самой программе.\n"
                  "Ваши данные не трогаются никогда: настройки, счета, пароли, журналы "
                  "и сессия Telegram остаются на месте."
                  ).grid(row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(4, 2))

        # Версия — САМОЙ ЗАМЕТНОЙ строкой в этом блоке. Владелец просил,
        # чтобы было видно, обновилось или нет; раньше это можно было понять
        # только косвенно, по изменившемуся поведению.
        self.version_var = tk.StringVar(value=f"Версия: {app_version.full()}")
        ttk.Label(up, textvariable=self.version_var, font=("Segoe UI", 10, "bold"),
                  wraplength=780, justify="left").grid(
            row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(2, 6))

        self.update_enabled_var = tk.BooleanVar(value=getattr(cfg, "UPDATE_ENABLED", False))
        ttk.Checkbutton(up, text="Проверять обновления",
                        variable=self.update_enabled_var).grid(
            row=2, column=0, columnspan=2, sticky="w", padx=8, pady=(8, 2))

        ttk.Label(up, text="Репозиторий:").grid(row=3, column=0, sticky="w", padx=8, pady=3)
        self.update_repo_var = tk.StringVar(value=getattr(cfg, "UPDATE_REPO", ""))
        ttk.Entry(up, textvariable=self.update_repo_var, width=44).grid(
            row=3, column=1, sticky="w", padx=8)
        ttk.Label(up, foreground=self.colors["dim"], text="владелец/название, например simafon1-cyber/repp"
                  ).grid(row=4, column=1, sticky="w", padx=8)

        ttk.Label(up, text="Ветка:").grid(row=5, column=0, sticky="w", padx=8, pady=3)
        self.update_branch_var = tk.StringVar(value=getattr(cfg, "UPDATE_BRANCH", ""))
        ttk.Entry(up, textvariable=self.update_branch_var, width=44).grid(
            row=5, column=1, sticky="w", padx=8)
        ttk.Label(up, foreground=self.colors["dim"],
                  text="пусто = программа сама возьмёт главную ветку репозитория"
                  ).grid(row=6, column=1, sticky="w", padx=8)

        ttk.Label(up, text="Токен GitHub:").grid(row=7, column=0, sticky="w", padx=8, pady=3)
        self.update_token_var = tk.StringVar(value=getattr(cfg, "UPDATE_TOKEN", ""))
        ttk.Entry(up, textvariable=self.update_token_var, width=44, show="*").grid(
            row=7, column=1, sticky="w", padx=8)
        ttk.Label(up, foreground=self.colors["dim"], wraplength=420, justify="left",
                  text="Для закрытого репозитория. Права: Contents: Read-only — "
                       "обычное обновление; Actions: Read and write — если хотите, "
                       "чтобы программа сама заказывала сборку .exe"
                  ).grid(row=8, column=1, sticky="w", padx=8)

        self.update_auto_var = tk.BooleanVar(value=getattr(cfg, "UPDATE_AUTO_APPLY", False))
        ttk.Checkbutton(up, variable=self.update_auto_var,
                        text="Ставить обновление само при запуске (не спрашивая)").grid(
            row=9, column=0, columnspan=2, sticky="w", padx=8, pady=(6, 0))
        ttk.Label(up, foreground=self.colors["dim"], wraplength=520, justify="left",
                  text="При старте торговля ещё не началась и открытых позиций у бота "
                       "нет — подменять его в этот момент безопасно. Посреди работы "
                       "обновление не ставится никогда, даже с этой галочкой.").grid(
            row=10, column=0, columnspan=2, sticky="w", padx=28, pady=(0, 2))

        self.update_build_var = tk.BooleanVar(value=getattr(cfg, "UPDATE_REQUEST_BUILD", False))
        ttk.Checkbutton(up, variable=self.update_build_var,
                        text="Если готовой сборки нет — заказать её на GitHub самому").grid(
            row=11, column=0, columnspan=2, sticky="w", padx=8, pady=(2, 0))
        ttk.Label(up, foreground=self.colors["dim"], wraplength=520, justify="left",
                  text="Раньше это делалось руками: вкладка Actions -> Run workflow. "
                       "Токену нужно право Actions: Read and write.").grid(
            row=12, column=0, columnspan=2, sticky="w", padx=28, pady=(0, 4))

        upbtn = ttk.Frame(up)
        upbtn.grid(row=13, column=0, columnspan=2, sticky="w", padx=8, pady=6)
        ttk.Button(upbtn, text="Обновить всё сейчас",
                   command=self.update_everything_now).grid(row=0, column=1, padx=6)
        ttk.Button(upbtn, text="Проверить обновления",
                   command=self.check_updates).grid(row=0, column=2)
        ttk.Button(upbtn, text="Собрать новую версию",
                   command=self.request_build_now).grid(row=0, column=3, padx=6)

        self.update_status_var = tk.StringVar(value="")
        ttk.Label(up, textvariable=self.update_status_var, foreground=self.colors["muted"],
                  wraplength=780, justify="left").grid(
            row=14, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

        # ---------- Журнал сделок в облаке ----------
        jr = ttk.LabelFrame(parent, text=" Журнал сделок в облаке ")
        jr.pack(fill="x", padx=12, pady=(6, 4))

        ttk.Label(jr, foreground=self.colors["muted"], wraplength=780, justify="left", text=
                  "История сделок выкладывается в папку journal/ вашего ЗАКРЫТОГО "
                  "репозитория GitHub: можно открыть с телефона или показать, когда "
                  "компьютер выключен. Три файла — журнал бота, реальные закрытые "
                  "сделки из MetaTrader и разбор словами (винрейт, средний плюс/минус, "
                  "сколько сделок умерло за секунды, какая пара даёт минус).\n"
                  "Пароли, ключи и токены туда НЕ попадают — только сделки.\n"
                  "Этот же репозиторий и токен используются для резервной копии "
                  "списка счетов (кнопки «Сохранить/Восстановить из облака» на "
                  "вкладке «Счета») — заводить для неё отдельные настройки не нужно."
                  ).grid(row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(4, 2))

        self.journal_enabled_var = tk.BooleanVar(
            value=getattr(cfg, "JOURNAL_CLOUD_ENABLED", False))
        ttk.Checkbutton(jr, text="Выкладывать журнал в облако",
                        variable=self.journal_enabled_var).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=8, pady=2)

        ttk.Label(jr, text="Репозиторий:").grid(row=2, column=0, sticky="w", padx=8, pady=3)
        self.journal_repo_var = tk.StringVar(value=getattr(cfg, "JOURNAL_REPO", ""))
        ttk.Entry(jr, textvariable=self.journal_repo_var, width=44).grid(
            row=2, column=1, sticky="w", padx=8)
        ttk.Label(jr, foreground=self.colors["dim"],
                  text="пусто = тот же, что для обновлений").grid(
            row=3, column=1, sticky="w", padx=8)

        ttk.Label(jr, text="Токен GitHub (запись):").grid(row=4, column=0, sticky="w", padx=8, pady=3)
        self.journal_token_var = tk.StringVar(
            value=SECRET_PLACEHOLDER if getattr(cfg, "JOURNAL_TOKEN", "") else "")
        ttk.Entry(jr, textvariable=self.journal_token_var, width=44, show="*").grid(
            row=4, column=1, sticky="w", padx=8)
        ttk.Label(jr, foreground=self.colors["dim"], wraplength=420, justify="left",
                  text="Fine-grained token на этот репозиторий, права "
                       "Contents: Read and write. Пустое поле = не менять").grid(
            row=5, column=1, sticky="w", padx=8)

        ttk.Label(jr, text="Как часто, минут:").grid(row=6, column=0, sticky="w", padx=8, pady=3)
        self.journal_minutes_var = tk.StringVar(
            value=str(getattr(cfg, "JOURNAL_UPLOAD_MINUTES", 15)))
        ttk.Entry(jr, textvariable=self.journal_minutes_var, width=10).grid(
            row=6, column=1, sticky="w", padx=8)

        jrbtn = ttk.Frame(jr)
        jrbtn.grid(row=7, column=0, columnspan=2, sticky="w", padx=8, pady=6)
        ttk.Button(jrbtn, text="Выложить сейчас",
                   command=self.upload_journal_now).grid(row=0, column=1, padx=6)
        ttk.Button(jrbtn, text="Открыть в браузере",
                   command=self.open_journal_in_browser).grid(row=0, column=2)

        self.journal_status_var = tk.StringVar(value="")
        ttk.Label(jr, textvariable=self.journal_status_var, foreground=self.colors["muted"],
                  wraplength=780, justify="left").grid(
            row=8, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

        # ---------- Проверка компьютера ----------
        ttk.Label(parent, text="Что установлено на этом компьютере",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(8, 2))

        ttk.Button(parent, text="Проверить заново",
                   command=self.refresh_diagnostics).pack(anchor="w", padx=12)

        cols = ("what", "state", "detail")
        self.diag_tree = ttk.Treeview(parent, columns=cols, show="headings", height=12)
        for col, head, width in (("what", "Что", 190), ("state", "Состояние", 90),
                                 ("detail", "Подробности", 480)):
            self.diag_tree.heading(col, text=head)
            self.diag_tree.column(col, width=width,
                                  anchor="center" if col == "state" else "w")
        self.diag_tree.pack(fill="both", expand=True, padx=12, pady=(4, 4))
        self.diag_tree.tag_configure("fail", foreground=self.colors["loss"])
        self.diag_tree.tag_configure("warn", foreground=self.colors["warning"])
        self.diag_tree.tag_configure("ok", foreground=self.colors["profit"])

        self.diag_summary_var = tk.StringVar(value="")
        ttk.Label(parent, textvariable=self.diag_summary_var, wraplength=800,
                  justify="left").pack(anchor="w", padx=12, pady=(0, 8))

        self.refresh_diagnostics()
        self._refresh_bridge_status()

    # Сколько раз пробовать поднять мост и календарь при запуске. Владелец:
    # «мост тоже не всегда сразу включается», «календарь MT5 показывает, что
    # не всегда работает». Обе вещи зависят от терминала MT5, который в
    # момент старта программы может быть ещё не готов: Windows поднимает
    # автозапуск пачкой, и терминал нередко доходит до рабочего состояния
    # позже. Одна попытка при старте — это лотерея.
    STARTUP_RETRIES = 5
    STARTUP_RETRY_MS = 20000

    def _start_bridge_if_enabled(self, attempt: int = 1):
        """Поднять мост. Не вышло — попробовать ещё через 20 секунд.

        Раньше попытка была ровно одна: если терминал в этот момент ещё не
        поднялся, мост оставался выключенным до перезапуска программы."""
        started = False
        try:
            if not bridge_host.enabled():
                return
            problem = bridge_host.start()
            if problem:
                log.warning("Мост (попытка %d): %s", attempt, problem)
            else:
                started = True
            self._refresh_bridge_status()
        except Exception as e:  # noqa: BLE001
            log.warning("Мост не запущен (попытка %d): %s", attempt, e)

        if started:
            if attempt > 1:
                runtime_events.record("мост", f"поднялся с {attempt}-й попытки")
            return
        if attempt < self.STARTUP_RETRIES:
            self.root.after(self.STARTUP_RETRY_MS,
                            lambda: self._start_bridge_if_enabled(attempt + 1))
        else:
            runtime_events.record(
                "мост", f"не поднялся за {self.STARTUP_RETRIES} попыток — "
                        f"проверьте вкладку «Система»")

    def _ensure_news_source(self, attempt: int = 1):
        """Дожать источник новостей: поставить и собрать сервис календаря,
        а если он поставлен, но файл календаря не появляется — повторить
        проверку. Терминал после запуска Windows готов не сразу, и первая
        проверка часто попадает в этот промежуток."""
        ready = False
        try:
            state = news_autostart.ensure_ready(force=True)
            ready = bool(state.get("ready"))
            if not ready and state.get("news_mode"):
                log.info("Календарь (попытка %d): %s", attempt,
                         news_autostart.describe(state))
        except Exception as e:  # noqa: BLE001
            log.warning("Проверка календаря не прошла (попытка %d): %s", attempt, e)

        if ready:
            if attempt > 1:
                runtime_events.record("календарь",
                                      f"заработал с {attempt}-й попытки")
            return
        if attempt < self.STARTUP_RETRIES:
            self.root.after(self.STARTUP_RETRY_MS,
                            lambda: self._ensure_news_source(attempt + 1))

    def refresh_diagnostics(self):
        """Проверки трогают диск и ищут терминалы — делаем это в фоне."""
        def worker():
            try:
                results = diagnostics.run_all()
            except Exception as e:
                log.exception("Проверка системы не удалась: %s", e)
                results = []
            self.root.after(0, lambda: self._apply_diagnostics(results))

        self.diag_summary_var.set("Проверяю...")
        threading.Thread(target=worker, daemon=True, name="diagnostics").start()

    def _apply_diagnostics(self, results):
        for item in self.diag_tree.get_children():
            self.diag_tree.delete(item)
        labels = {diagnostics.OK: "есть", diagnostics.WARN: "внимание",
                  diagnostics.FAIL: "НЕТ"}
        for r in results:
            detail = r["detail"]
            if r["fix"]:
                detail += f"  ->  {r['fix']}"
            self.diag_tree.insert("", "end", tags=(r["level"],),
                                  values=(r["name"], labels.get(r["level"], ""), detail))
        self.diag_summary_var.set(diagnostics.summary(results) if results
                                  else "Не удалось выполнить проверку.")

    def _refresh_bridge_status(self):
        try:
            st = bridge_host.status()
            self.bridge_status_var.set(
                st["detail"] + (f"  Запросов от советников: {st['requests']}."
                                if st["requests"] else ""))
        except Exception:
            pass

    def upload_journal_now(self):
        """Выложить журнал сделок в облако прямо сейчас.

        Сеть — в фоне: окно не должно замирать, пока GitHub принимает файлы."""
        ok, reason = cloud_journal.ready()
        if not ok:
            messagebox.showwarning(APP_TITLE, reason)
            return
        if not messagebox.askyesno(
            APP_TITLE,
            f"Выложить историю сделок в репозиторий {cloud_journal.repo()} "
            f"(ветка {cloud_journal.branch()}, папка {cloud_journal.folder()}/)?\n\n"
            f"Уйдут только сделки. Пароли, ключи и токены не отправляются.",
        ):
            return

        self.journal_status_var.set("Отправляю...")

        def worker():
            try:
                result = cloud_journal.upload(ds.get_snapshot())
            except Exception as e:  # noqa: BLE001
                log.exception("Журнал в облако: %s", e)
                result = {"ok": False, "error": cloud_journal.explain_error(e),
                          "files": [], "analysis": {}}
            self.root.after(0, lambda: self._apply_journal_result(result))

        threading.Thread(target=worker, daemon=True, name="cloud-journal").start()

    def _apply_journal_result(self, result):
        # Вкладка «Система» могла быть не построена (простой режим интерфейса) —
        # тогда показывать статус негде, но выгрузка всё равно нужна.
        if not hasattr(self, "journal_status_var"):
            return
        if not result.get("ok"):
            self.journal_status_var.set("Не отправлено: " + result.get("error", ""))
            return
        analysis = result.get("analysis") or {}
        self.journal_status_var.set(
            f"Готово: {len(result.get('files', []))} файла в {cloud_journal.repo()} "
            f"({time.strftime('%H:%M:%S')}). Сделок в разборе: "
            f"{analysis.get('trades', 0)}, итог {analysis.get('net', 0)}. "
            f"Открыть: {cloud_journal.web_url()}")

    def open_journal_in_browser(self):
        url = cloud_journal.web_url()
        if not url:
            messagebox.showwarning(APP_TITLE, "Сначала укажите репозиторий.")
            return
        webbrowser.open(url)

    def save_system_settings(self, silent: bool = False):
        try:
            port = int(self.bridge_port_var.get().strip() or 8080)
        except ValueError:
            messagebox.showwarning(APP_TITLE, "Порт должен быть числом.")
            return
        if not (1024 <= port <= 65535):
            messagebox.showwarning(APP_TITLE, "Порт должен быть в диапазоне 1024-65535.")
            return

        try:
            journal_minutes = int(float(self.journal_minutes_var.get().strip() or 15))
        except ValueError:
            messagebox.showwarning(APP_TITLE, "«Как часто, минут» — должно быть числом.")
            return
        if journal_minutes < 1:
            journal_minutes = 1

        pw = control.get_session_password()
        salt = getattr(cfg, "SECURITY_SALT", "")
        # Токен GitHub — такой же секрет, как ключи API
        token = secure_store.protect_secret(self.update_token_var.get().strip(), pw, salt)

        _write_config_value("BRIDGE_ENABLED", repr(bool(self.bridge_enabled_var.get())))
        _write_config_value("BRIDGE_PORT", repr(port))
        _write_config_value("UPDATE_ENABLED", repr(bool(self.update_enabled_var.get())))
        _write_config_value("UPDATE_REPO", repr(self.update_repo_var.get().strip()))
        # Пустое поле НЕ подменяем на "main" насильно: у репозитория может не
        # быть такой ветки вовсе (обычное дело, пока никто не влил ветку в
        # главную) — тогда каждый файл отвечал бы "не найден" без понятной
        # причины. Пустая строка означает "программа сама узнает у GitHub
        # главную ветку репозитория" (см. updater.repo_default_branch()).
        _write_config_value("UPDATE_BRANCH", repr(self.update_branch_var.get().strip()))
        _write_config_value("UPDATE_TOKEN", repr(token))
        _write_config_value("UPDATE_AUTO_APPLY", repr(bool(self.update_auto_var.get())))
        _write_config_value("UPDATE_REQUEST_BUILD", repr(bool(self.update_build_var.get())))

        _write_config_value("JOURNAL_CLOUD_ENABLED", repr(bool(self.journal_enabled_var.get())))
        _write_config_value("JOURNAL_REPO", repr(self.journal_repo_var.get().strip()))
        _write_config_value("JOURNAL_UPLOAD_MINUTES", repr(journal_minutes))
        # Пустое поле или заглушка = токен не меняли. Иначе затирали бы
        # сохранённый токен каждый раз, когда сохраняются любые другие
        # настройки на этой вкладке.
        journal_token = self.journal_token_var.get().strip()
        if journal_token and journal_token != SECRET_PLACEHOLDER:
            _write_config_value(
                "JOURNAL_TOKEN",
                repr(secure_store.protect_secret(journal_token, pw, salt)))
            self.journal_token_var.set(SECRET_PLACEHOLDER)
        try:
            _reload_cfg()
        except Exception:
            pass

        # Мост перезапускаем сразу: ждать перезапуска программы ради галочки незачем
        bridge_host.stop()
        if bridge_host.enabled():
            problem = bridge_host.start()
            if problem:
                messagebox.showwarning(APP_TITLE, problem)
        self._refresh_bridge_status()
        if not silent:
            messagebox.showinfo(APP_TITLE, "Настройки системы сохранены.")
        self.refresh_diagnostics()

    # ---- обновление при запуске ---------------------------------------------
    # Верхняя граница ожидания: даже если GitHub не отвечает и обновление
    # висит, торговля должна начаться. Лучше работать на старой версии, чем не
    # работать вовсе.
    START_BOT_MAX_WAIT_TICKS = 240        # 240 x 500 мс = 2 минуты

    def _start_bot_when_ready(self):
        """Ждём, пока доставится обновление, и только потом стартуем торговлю.

        Запускать бота посреди замены файлов нельзя: часть модулей была бы
        старой, часть новой."""
        if not getattr(self, "_auto_update_busy", False):
            self.start_bot()
            return
        self._start_bot_waits += 1
        if self._start_bot_waits >= self.START_BOT_MAX_WAIT_TICKS:
            log.warning("Обновление при запуске затянулось — начинаю торговлю "
                        "на текущей версии.")
            self._auto_update_busy = False
            self.start_bot()
            return
        self.root.after(500, self._start_bot_when_ready)

    def _schedule_update_check(self):
        """Проверять обновления не только при запуске, но и во время работы.

        Владелец: «чтобы программа сама проверяла обновления и загружала
        нововведения». На сервере программа не перезапускается неделями, и
        проверка при запуске туда просто не доходит.

        Ставим следующую проверку ЗАРАНЕЕ, до самой проверки: если она упадёт
        (нет сети, GitHub не отвечает), цепочка не оборвётся и следующая
        попытка всё равно состоится."""
        minutes = float(getattr(cfg, "UPDATE_CHECK_MINUTES", 0) or 0)
        if minutes <= 0 or not updater.enabled():
            return
        self.root.after(int(minutes * 60_000), self._periodic_update_check)

    def open_positions_count(self) -> int:
        """Сколько сделок сейчас открыто, по данным торгового цикла.

        Не спрашиваем MetaTrader напрямую: окно вообще никогда не ходит в
        терминал, иначе оно подвисало бы вместе с ним. Берём последний
        снимок, который выкладывает торговый цикл."""
        try:
            snap = ds.get_snapshot() or {}
            rows = snap.get("positions") or []
            return len(rows)
        except Exception:           # noqa: BLE001
            # Не смогли выяснить — считаем, что сделки ЕСТЬ. Ошибаться нужно
            # в сторону «не трогать программу», а не наоборот.
            return 1

    def _periodic_update_check(self):
        self._schedule_update_check()
        if getattr(self, "_auto_update_busy", False):
            return
        try:
            if not getattr(cfg, "UPDATE_AUTO_APPLY", False):
                self.check_updates(silent=True)
                return

            # ПОДМЕНЯТЬ ПРОГРАММУ, ПОКА ОТКРЫТЫ СДЕЛКИ, НЕЛЬЗЯ.
            #
            # Автоустановка задумывалась только для момента запуска — там
            # торговля ещё не началась, и открытых позиций у бота нет (см.
            # комментарий в _build_ui). Периодическую проверку я добавил
            # позже и этого условия не перенёс: с включённой автоустановкой
            # она подменила бы файлы посреди ведения сделок — а сделку ведут
            # трейлинг и безубыток, и остаться без них с открытой позицией
            # хуже, чем обновиться на несколько часов позже.
            #
            # Поэтому: есть открытые сделки — просто ждём следующей проверки.
            # Она всё равно повторится, и рано или поздно застанет момент,
            # когда сделок нет.
            open_now = self.open_positions_count()
            if open_now > 0:
                log.info("Обновление отложено: открыто сделок %d", open_now)
                return

            self._auto_update_busy = True
            self._auto_apply_update()
        except Exception as e:      # noqa: BLE001
            log.warning("Проверка обновлений не удалась: %s", e)

    def _auto_apply_update(self):
        """Скачать и поставить обновление без вопросов — только при запуске."""
        self.status_var.set("Проверяю обновления...")

        def worker():
            summary = {"errors": [], "lines": [], "restart_needed": False}
            try:
                check = updater.check()
                if check.get("error"):
                    summary["errors"].append(check["error"])
                elif check.get("available"):
                    summary = updater.update_everything()
                    summary["revision"] = check.get("revision", "")
                else:
                    summary["lines"].append("Установлена последняя версия.")
            except Exception as e:  # noqa: BLE001
                log.warning("Автообновление не удалось: %s", e)
                summary["errors"].append(updater.explain_error(e))
            self.root.after(0, lambda: self._after_auto_update(summary))

        threading.Thread(target=worker, daemon=True, name="update-auto").start()

    def _after_auto_update(self, summary: dict):
        self._auto_update_busy = False
        errors = summary.get("errors") or []
        if errors:
            # Нет связи с GitHub — не повод не торговать. Пишем в журнал и
            # работаем на той версии, что есть.
            log.warning("Автообновление: %s", "; ".join(errors[:3]))
            self.status_var.set("Готово (обновиться не удалось)")
            return

        if summary.get("revision"):
            try:
                updater.remember_revision(summary["revision"], _write_config_value)
            except Exception:
                pass
        try:
            config_migrate.sync()
        except Exception as e:  # noqa: BLE001
            log.warning("Не удалось дописать новые настройки: %s", e)

        if not summary.get("restart_needed"):
            return

        # Файлы на диске уже новые, а в памяти — старые модули. Работать в
        # таком виде нельзя: части программы разошлись бы по версиям.
        self._show_toast("Обновление установлено",
                         "Программа перезапускается в новой версии.")
        self.root.after(1500, self._restart_after_update)

    def _restart_after_update(self):
        try:
            self.stop_bot()
        except Exception:
            pass
        problem = updater.restart_program()
        if problem:
            messagebox.showwarning(
                APP_TITLE,
                problem + "\n\nЗакройте и откройте программу вручную — "
                          "обновление встанет при следующем запуске.")

    def _bot_is_busy(self) -> str:
        """Почему сейчас нельзя подменять программу. Пустая строка — можно.

        Обновление посреди работы — это подмена торгового робота под открытыми
        сделками: позиции останутся без того, кто ведёт их стоп и цель."""
        if self.bot_thread and self.bot_thread.is_alive():
            snap = ds.get_snapshot() or {}
            open_count = len(snap.get("positions", []))
            if open_count:
                return (f"Сейчас открыто сделок: {open_count}. Обновление подменяет "
                        f"программу — сначала закройте сделки или остановите бота.")
            return ("Бот работает. Обновление лучше ставить на остановленном: "
                    "нажмите «Стоп» на вкладке «Обзор».")
        return ""

    def _apply_update_fields(self) -> list:
        """Записать в config.py то, что СЕЙЧАС набрано в полях обновления.

        Обновление читает настройки из сохранённого config.py (updater.repo(),
        updater.branch()), а не из полей на экране. Человек, вписавший ветку и
        сразу нажавший «Обновить всё сейчас», справедливо ждёт, что использована
        будет вписанная ветка — а получал ошибку про старое сохранённое
        значение и никакого намёка, что надо было сперва нажать «Сохранить».
        Поэтому перед любым действием с GitHub сами дописываем изменённое.

        Возвращает список того, что поменялось — чтобы сказать об этом вслух."""
        changed = []
        pairs = [
            ("UPDATE_ENABLED", bool(self.update_enabled_var.get())),
            ("UPDATE_REPO", self.update_repo_var.get().strip()),
            ("UPDATE_BRANCH", self.update_branch_var.get().strip()),
        ]
        for name, value in pairs:
            if getattr(cfg, name, None) == value:
                continue
            _write_config_value(name, repr(value))
            changed.append(name)

        # Токен: пустое поле и заглушка означают «не менять», иначе затирали бы
        # сохранённый токен при каждом нажатии кнопки обновления.
        token_text = self.update_token_var.get().strip()
        if token_text and token_text != SECRET_PLACEHOLDER:
            if token_text != getattr(cfg, "UPDATE_TOKEN", ""):
                pw = control.get_session_password()
                salt = getattr(cfg, "SECURITY_SALT", "")
                _write_config_value(
                    "UPDATE_TOKEN",
                    repr(secure_store.protect_secret(token_text, pw, salt)))
                changed.append("UPDATE_TOKEN")
        elif not token_text and getattr(cfg, "UPDATE_TOKEN", ""):
            # Поле очистили осознанно — значит токен больше не нужен
            # (например, репозиторий открытый). Это ОТЛИЧАЕТСЯ от заглушки.
            _write_config_value("UPDATE_TOKEN", repr(""))
            changed.append("UPDATE_TOKEN (очищен)")

        if changed:
            try:
                _reload_cfg()
            except Exception as e:  # noqa: BLE001
                log.warning("Не удалось перечитать настройки обновления: %s", e)
            updater.reset_caches()
        return changed

    def update_everything_now(self):
        """Одна кнопка: советники + сама программа + новые настройки."""
        applied = self._apply_update_fields()
        if applied:
            self.update_status_var.set(
                "Сохранены изменённые настройки: " + ", ".join(applied))
        # Называем РОВНО то, чего не хватает. Раньше здесь стояла одна фраза
        # на оба случая: «включите галочку И впишите репозиторий». Владелец
        # получил её при вписанном репозитории и снятой галочке — и пошёл
        # искать несуществующую проблему с репозиторием.
        missing = []
        if not updater.enabled():
            missing.append(
                "поставьте галочку «Проверять обновления» (сейчас она снята)")
        if not (updater.repo() and "/" in updater.repo()):
            missing.append(
                "впишите репозиторий в виде владелец/название, "
                "например simafon1-cyber/repp")
        if missing:
            messagebox.showwarning(
                APP_TITLE,
                "Обновление не запущено. Что сделать:\n\n• "
                + "\n• ".join(missing)
                + "\n\nЗатем нажмите «Сохранить все настройки» и повторите.")
            return

        busy = self._bot_is_busy()
        if busy and not messagebox.askyesno(
                APP_TITLE, busy + "\n\nВсё равно продолжить?"):
            return
        if not messagebox.askyesno(
                APP_TITLE,
                f"Скачать и поставить свежую версию из {updater.repo()} "
                f"(ветка {updater.branch()})?\n\n"
                "Обновятся советники в MetaTrader и файлы программы.\n"
                "Настройки, счета, пароли и журналы НЕ трогаются."):
            return

        self.update_status_var.set("Обновляю...")

        def worker():
            try:
                summary = updater.update_everything(
                    progress=lambda t: self.root.after(
                        0, lambda: self.update_status_var.set(t)))
            except Exception as e:  # noqa: BLE001
                log.exception("Обновление не удалось: %s", e)
                summary = {"errors": [updater.explain_error(e)], "lines": [],
                           "restart_needed": False}
            self.root.after(0, lambda: self._after_update_everything(summary))

        threading.Thread(target=worker, daemon=True, name="update-all").start()

    def _after_update_everything(self, summary: dict):
        errors = summary.get("errors") or []
        lines = summary.get("lines") or []

        if errors:
            text = "Обновление не завершено:\n- " + "\n- ".join(errors[:4])
            self.update_status_var.set(text.replace("\n", " "))
            # Готовой сборки нет — можем заказать её прямо отсюда
            if (any("Собрать новую версию" in e for e in errors)
                    and getattr(cfg, "UPDATE_REQUEST_BUILD", False)):
                self._request_build_worker(auto=True)
                return
            messagebox.showwarning(APP_TITLE, text)
            return

        # Токен был, GitHub его не принял, но репозиторий открытый и всё
        # получилось без токена. Молчать нельзя: настройка человека фактически
        # не сработала, и он должен об этом узнать.
        if updater.token_was_ignored():
            lines.append(
                "Токен GitHub не принят и был пропущен — репозиторий открытый, "
                "токен для обновления не нужен. Если он вам не нужен и дальше, "
                "очистите поле «Токен GitHub», чтобы это сообщение не повторялось.")

        # Вписанной ветки в репозитории не оказалось, программа подобрала
        # существующую сама. Молчать нельзя по той же причине: настройка
        # человека фактически не сработала.
        fixed_branch = updater.branch_was_fixed()
        if fixed_branch:
            lines.append(
                f"Ветки, вписанной в поле «Ветка», в репозитории нет — "
                f"обновление взято из «{fixed_branch}». Впишите её в поле и "
                f"нажмите «Сохранить», либо очистите поле совсем: тогда "
                f"программа всегда берёт главную ветку репозитория сама.")

        # Новые настройки — тем же нажатием: иначе после обновления кода
        # программа искала бы параметры, которых нет в вашем config.py.
        try:
            added = config_migrate.sync()
            if added:
                lines.append(f"Добавлены новые настройки: {', '.join(added)}.")
                _reload_cfg()
        except Exception as e:  # noqa: BLE001
            log.warning("Не удалось дописать новые настройки: %s", e)

        revision = ""
        try:
            check = updater.check()
            revision = check.get("revision", "")
            if revision:
                updater.remember_revision(revision, _write_config_value)
                _reload_cfg()
        except Exception:
            pass

        text = "\n".join(lines) if lines else "Всё уже свежее — обновлять нечего."
        if summary.get("restart_needed"):
            text += "\n\nПерезапустите программу, чтобы новая версия заработала."
        self.update_status_var.set(text.replace("\n", " "))
        self._refresh_version_line()
        messagebox.showinfo(APP_TITLE, text)

    def _refresh_version_line(self):
        """Обновить строку версии: та ли она, что лежит в GitHub.

        Запрос к сети идёт в отдельном потоке — окно не должно замирать
        из-за строчки текста."""
        def worker():
            try:
                text = updater.version_status()
            except Exception as e:  # noqa: BLE001
                text = f"Версия: {app_version.full()} (проверить не удалось: {e})"
            self.root.after(0, lambda: self.version_var.set(text))

        threading.Thread(target=worker, daemon=True, name="version-check").start()

    def request_build_now(self):
        self._apply_update_fields()
        if not messagebox.askyesno(
                APP_TITLE,
                "Попросить GitHub собрать новую версию программы?\n\n"
                "Сборка идёт 5-10 минут на серверах GitHub — ваш компьютер не "
                "нагружается. Когда закончится, нажмите «Обновить всё сейчас»."):
            return
        self._request_build_worker(auto=False)

    def _request_build_worker(self, auto: bool):
        self.update_status_var.set("Прошу GitHub собрать новую версию...")

        def worker():
            problem = updater.request_build()
            status = updater.build_status() if not problem else {}
            self.root.after(0, lambda: self._after_build_request(problem, status, auto))

        threading.Thread(target=worker, daemon=True, name="update-build").start()

    def _after_build_request(self, problem: str, status: dict, auto: bool):
        if problem:
            self.update_status_var.set(problem)
            messagebox.showwarning(APP_TITLE, problem)
            return
        text = ("Сборка заказана. " + (status.get("text", "") or
                "Обычно занимает 5-10 минут.") +
                "\nКогда закончится — нажмите «Обновить всё сейчас».")
        self.update_status_var.set(text.replace("\n", " "))
        if not auto:
            messagebox.showinfo(APP_TITLE, text)

    def check_updates(self, silent: bool = False):
        if not silent:
            # Ручная проверка: используем то, что набрано в полях сейчас.
            # Автопроверка при запуске (silent) полей не трогает — окно ещё
            # только открылось, менять настройки из-за неё нельзя.
            self._apply_update_fields()

        def worker():
            result = updater.check()
            self.root.after(0, lambda: self._after_update_check(result, silent))

        self.update_status_var.set("Проверяю GitHub...")
        threading.Thread(target=worker, daemon=True, name="update-check").start()

    def _after_update_check(self, result: dict, silent: bool):
        if result.get("error"):
            self.update_status_var.set(result["error"])
            if not silent:
                messagebox.showwarning(APP_TITLE, result["error"])
            return

        if not result.get("available"):
            self.update_status_var.set(result.get("message") or "Обновлений нет.")
            if not silent:
                messagebox.showinfo(APP_TITLE, "Установлена последняя версия.")
            return

        revision = result["revision"]
        text = f"Есть новая версия ({revision}): {result.get('message', '')}"
        self.update_status_var.set(text)

        # Спрашиваем ВСЕГДА, даже при автопроверке: подменять торгового робота
        # без ведома человека, пока у него открыты позиции, недопустимо.
        if not messagebox.askyesno(
                APP_TITLE,
                f"{text}\n\nОбновить советники в MetaTrader сейчас?\n\n"
                "Программа при этом продолжит работать в текущей версии — "
                "она обновляется отдельно, при перезапуске."):
            return
        self._apply_update(revision)

    def _apply_update(self, revision: str):
        def worker():
            report = updater.update_advisors(
                progress=lambda t: self.root.after(0, lambda: self.update_status_var.set(t)))
            self.root.after(0, lambda: self._after_update_applied(report, revision))

        self.update_status_var.set("Скачиваю обновление...")
        threading.Thread(target=worker, daemon=True, name="update-apply").start()

    def _after_update_applied(self, report: dict, revision: str):
        if report.get("errors"):
            text = "Обновление не завершено: " + "; ".join(report["errors"][:3])
            self.update_status_var.set(text)
            messagebox.showwarning(APP_TITLE, text)
            return
        updater.remember_revision(revision, _write_config_value)
        try:
            _reload_cfg()
        except Exception:
            pass
        text = f"Советники обновлены до версии {revision}. {report.get('installed', '')}"
        self.update_status_var.set(text)
        messagebox.showinfo(APP_TITLE, text)

    # ---- вкладка "Источники" ---------------------------------------------------------
    def _build_tab_sources(self, parent):
        """Одно место, где всё включается и выключается.

        Раньше настройки календаря жили на вкладке «Новости», а Telegram — на
        «Сигналы TG», и чтобы что-то отключить, надо было помнить, где именно
        оно лежит. Теперь выключатели собраны здесь, а те вкладки показывают
        только данные."""
        # Боковой ползунок: владелец просил не растягивать окно
        # каждый раз — блоков здесь больше, чем помещается.
        parent = self._scrollable(parent)
        pad = {"padx": 12, "pady": 4}

        ttk.Label(parent, text="Откуда брать данные",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", **pad)
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=820, justify="left", text=
                  "Снимите галочку — источник перестанет использоваться. Календарь и "
                  "сигналы Telegram независимы: выключение одного не трогает другое."
                  ).pack(anchor="w", **pad)

        # ---------- Установка в MetaTrader ----------
        mt = ttk.LabelFrame(parent, text=" Установка в MetaTrader 5 ")
        mt.pack(fill="x", padx=12, pady=(8, 4))

        ttk.Label(mt, foreground=self.colors["muted"], wraplength=780, justify="left", text=
                  "Советники и сервис календаря копируются в терминал и собираются "
                  "автоматически — жать F7 в MetaEditor не нужно. Делается один раз "
                  "при первом запуске; кнопка ниже — если терминал переустановили или "
                  "добавили новый."
                  ).grid(row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(4, 2))

        ttk.Button(mt, text="Установить в MetaTrader",
                   command=self.install_into_mt5).grid(row=1, column=0, sticky="w",
                                                       padx=8, pady=4)
        self.mt_install_var = tk.StringVar(value="")
        ttk.Label(mt, textvariable=self.mt_install_var, foreground=self.colors["muted"],
                  wraplength=600, justify="left").grid(row=1, column=1, sticky="w", padx=8)

        # ---------- Календарь новостей ----------
        cal = ttk.LabelFrame(parent, text=" Календарь новостей ")
        cal.pack(fill="x", padx=12, pady=(8, 4))

        ttk.Label(cal, foreground=self.colors["muted"], wraplength=780, justify="left", text=
                  "Опрашиваются по порядку — берётся первый, который ответил."
                  ).grid(row=0, column=0, columnspan=3, sticky="w", padx=8, pady=(4, 2))

        self.news_chain_vars = {}
        self.src_status_vars = {}
        current_chain = list(news_calendar.news_source_chain())
        row = 1
        for name in news_providers.PROVIDERS.keys():
            var = tk.BooleanVar(value=name in current_chain)
            self.news_chain_vars[name] = var
            title = news_providers.PROVIDER_TITLES.get(name, name)
            ttk.Checkbutton(cal, text=title, variable=var).grid(
                row=row, column=0, sticky="w", padx=8, pady=2)
            status = tk.StringVar(value="")
            self.src_status_vars[name] = status
            ttk.Label(cal, textvariable=status, foreground=self.colors["muted"]).grid(
                row=row, column=1, sticky="w", padx=8)
            row += 1

        ttk.Label(cal, text="Ключ Finnhub:").grid(row=row, column=0, sticky="w", padx=8, pady=4)
        keys = getattr(cfg, "NEWS_API_KEYS", {}) or {}
        self.news_api_key_var = tk.StringVar(value=keys.get("finnhub", ""))
        ttk.Entry(cal, textvariable=self.news_api_key_var, width=44, show="*").grid(
            row=row, column=1, sticky="w", padx=8, pady=4)

        # ---------- Сигналы Telegram ----------
        tg = ttk.LabelFrame(parent, text=" Сигналы из Telegram ")
        tg.pack(fill="x", padx=12, pady=(8, 4))

        self.tg_enabled_var = tk.BooleanVar(value=getattr(cfg, "TELEGRAM_ENABLED", False))
        ttk.Checkbutton(tg, text="Читать сигналы из Telegram", variable=self.tg_enabled_var).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(6, 2))

        ttk.Label(tg, foreground=self.colors["muted"], wraplength=780, justify="left", text=
                  "Читается входом под ВАШИМ аккаунтом: Telegram запрещает ботам видеть "
                  "сообщения других ботов. Каналы читаются так же, как боты, но на "
                  "закрытый канал нужно быть подписанным."
                  ).grid(row=1, column=0, columnspan=2, sticky="w", padx=8, pady=2)

        ttk.Label(tg, text="Каналы и боты:").grid(row=2, column=0, sticky="w", padx=8, pady=3)
        self.tg_sources_var = tk.StringVar(
            value=", ".join(str(x) for x in (getattr(cfg, "TELEGRAM_SOURCES", []) or [])))
        ttk.Entry(tg, textvariable=self.tg_sources_var, width=54).grid(
            row=2, column=1, sticky="w", padx=8, pady=3)
        ttk.Label(tg, foreground=self.colors["dim"], text="через запятую, например:  @signals_channel, @my_crypto_signalsbot"
                  ).grid(row=3, column=1, sticky="w", padx=8)

        ttk.Label(tg, text="api_id:").grid(row=4, column=0, sticky="w", padx=8, pady=3)
        self.tg_api_id_var = tk.StringVar(value=str(getattr(cfg, "TELEGRAM_API_ID", 0) or ""))
        ttk.Entry(tg, textvariable=self.tg_api_id_var, width=22).grid(
            row=4, column=1, sticky="w", padx=8, pady=3)

        ttk.Label(tg, text="api_hash:").grid(row=5, column=0, sticky="w", padx=8, pady=3)
        self.tg_api_hash_var = tk.StringVar(value=getattr(cfg, "TELEGRAM_API_HASH", ""))
        ttk.Entry(tg, textvariable=self.tg_api_hash_var, width=54, show="*").grid(
            row=5, column=1, sticky="w", padx=8, pady=3)
        ttk.Label(tg, foreground=self.colors["dim"], text="бесплатно на my.telegram.org -> API development tools"
                  ).grid(row=6, column=1, sticky="w", padx=8)

        ttk.Label(tg, text="Что сигнал может:").grid(row=7, column=0, sticky="nw", padx=8, pady=(6, 3))
        self.tg_role_var = tk.StringVar(value=tgs.role())
        role_box = ttk.Frame(tg)
        role_box.grid(row=7, column=1, sticky="w", padx=8, pady=(6, 3))
        for i, (value, title) in enumerate((
                (tgs.ROLE_SHOW, "Только показывать — на торговлю не влияет"),
                (tgs.ROLE_VETO, "Может запретить вход"),
                (tgs.ROLE_SCORE, "Может запретить вход и добавить баллы"))):
            ttk.Radiobutton(role_box, text=title, value=value,
                            variable=self.tg_role_var).grid(row=i, column=0, sticky="w")

        ttk.Label(tg, foreground=self.colors["warning"], wraplength=780, justify="left", text=
                  "Ни в одном режиме чужой сигнал не может открыть сделку, увеличить лот "
                  "или риск, отодвинуть стоп-лосс и обойти лимиты риска."
                  ).grid(row=8, column=0, columnspan=2, sticky="w", padx=8, pady=(2, 6))

        self.tg_src_status_var = tk.StringVar(value="")
        ttk.Label(tg, textvariable=self.tg_src_status_var, foreground=self.colors["muted"],
                  wraplength=780, justify="left").grid(
            row=9, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

        # ---------- Кнопки ----------
        btns = ttk.Frame(parent)
        btns.pack(anchor="w", **pad)
        ttk.Button(btns, text="Войти в Telegram",
                   command=self.telegram_login).grid(row=0, column=1, padx=6)
        ttk.Button(btns, text="Проверить источники",
                   command=lambda: self.refresh_sources_tab(apply_fields=True)
                   ).grid(row=0, column=2)

        self.refresh_sources_tab()

    def install_into_mt5(self, silent: bool = False):
        """Установка советников и сервиса в терминал.

        Идёт в фоновом потоке: копирование и компиляция занимают секунды, но
        MetaEditor может задуматься, и замораживать окно на это время нельзя.

        silent=True — автоматический запуск при первом старте: без окошка
        «готово», только строка состояния. Молча ставить что-то в чужую
        программу и ещё и рапортовать об этом поверх экрана — перебор."""
        def worker():
            report = mt5_install.install_all(
                progress=lambda t: self.root.after(0, lambda: self._set_mt_status(t)))
            self.root.after(0, lambda: self._after_mt_install(report, silent))

        self._set_mt_status("Устанавливаю в MetaTrader...")
        threading.Thread(target=worker, daemon=True, name="mt5-install").start()

    def _set_mt_status(self, text: str):
        var = getattr(self, "mt_install_var", None)
        if var is not None:
            try:
                var.set(text)
            except Exception:
                pass

    def _after_mt_install(self, report: dict, silent: bool):
        self._set_mt_status(report.get("text", ""))
        log.info("Установка в MetaTrader: %s", report.get("text", ""))
        if silent:
            return
        if report.get("terminals"):
            messagebox.showinfo(APP_TITLE, report.get("text", "Готово."))
        else:
            messagebox.showwarning(APP_TITLE, report.get("text", "Не удалось."))

    def _auto_install_into_mt5_once(self):
        """Первый запуск: ставим всё сами, без вопросов.

        Смысл в том, чтобы пользователю не пришлось ставить ничего отдельно.
        Повторно не лезем: если сервис календаря уже на месте, считаем, что
        установка была."""
        try:
            if not mt5_install.sources_available():
                return
            if mt5_install.is_installed():
                self._set_mt_status("Уже установлено в MetaTrader.")
                return
            self.install_into_mt5(silent=True)
        except Exception as e:
            log.warning("Автоустановка в MetaTrader не выполнена: %s", e)

    def _apply_source_fields(self) -> list:
        """Записать в config.py то, что СЕЙЧАС набрано на вкладке «Источники».

        Живой случай: владелец поставил галочку «Читать сигналы из Telegram»,
        вписал api_id и api_hash, нажал «Войти в Telegram» — а программа
        ответила «Telegram выключен в настройках (TELEGRAM_ENABLED = False)».

        Причина ровно та же, что была у обновления с веткой: кнопки читали
        СОХРАНЁННЫЙ config.py, а не поля на экране. Пока не нажата «Сохранить
        всё», для программы галочка не поставлена, а api_id пуст — и она
        честно про это писала, только человек-то видел заполненные поля.

        Возвращает список того, что изменилось (для лога и объяснения)."""
        changed = []
        try:
            api_id = int(self.tg_api_id_var.get().strip() or 0)
        except ValueError:
            api_id = int(getattr(cfg, "TELEGRAM_API_ID", 0) or 0)

        pw = control.get_session_password()
        salt = getattr(cfg, "SECURITY_SALT", "")
        sources = [s.strip() for s in self.tg_sources_var.get().split(",") if s.strip()]

        pairs = [
            ("TELEGRAM_ENABLED", bool(self.tg_enabled_var.get())),
            ("TELEGRAM_API_ID", api_id),
            ("TELEGRAM_SOURCES", sources),
            ("TELEGRAM_ROLE", self.tg_role_var.get()),
        ]
        for name, value in pairs:
            if getattr(cfg, name, None) == value:
                continue
            _write_config_value(name, repr(value))
            changed.append(name)

        # api_hash — секрет: пустое поле и заглушка означают «не менять»,
        # иначе одно нажатие затирало бы уже сохранённый ключ.
        hash_text = self.tg_api_hash_var.get().strip()
        if hash_text and hash_text != SECRET_PLACEHOLDER:
            _write_config_value(
                "TELEGRAM_API_HASH",
                repr(secure_store.protect_secret(hash_text, pw, salt)))
            changed.append("TELEGRAM_API_HASH")

        if changed:
            try:
                _reload_cfg()
            except Exception:  # noqa: BLE001
                pass
        return changed

    def save_sources(self, silent: bool = False):
        """Сохраняет обе группы разом — это одна кнопка на всю вкладку."""
        chain = [name for name, var in self.news_chain_vars.items() if var.get()]
        tg_on = bool(self.tg_enabled_var.get())
        sources = [s.strip() for s in self.tg_sources_var.get().split(",") if s.strip()]

        if not chain and not tg_on:
            if not messagebox.askyesno(
                    APP_TITLE,
                    "Выключены ВСЕ источники: и календарь новостей, и Telegram.\n\n"
                    "Бот продолжит торговать, но перестанет останавливаться перед "
                    "выходом важных новостей.\n\nВсё равно сохранить?"):
                return

        try:
            api_id = int(self.tg_api_id_var.get().strip() or 0)
        except ValueError:
            messagebox.showwarning(APP_TITLE, "api_id должен быть числом.")
            return

        pw = control.get_session_password()
        salt = getattr(cfg, "SECURITY_SALT", "")

        def protect(value):
            """Как секрет ляжет в config.py: зашифрованным или открытым
            текстом в приватном режиме — решает secure_store."""
            return secure_store.protect_secret(value, pw, salt)

        keys = dict(getattr(cfg, "NEWS_API_KEYS", {}) or {})
        keys["finnhub"] = protect(self.news_api_key_var.get().strip())

        _write_config_value("NEWS_PROVIDER_CHAIN", repr(chain))
        _write_config_value("NEWS_API_KEYS", repr(keys))
        _write_config_value("TELEGRAM_ENABLED", repr(tg_on))
        _write_config_value("TELEGRAM_API_ID", repr(api_id))
        _write_config_value("TELEGRAM_API_HASH", repr(protect(self.tg_api_hash_var.get().strip())))
        _write_config_value("TELEGRAM_SOURCES", repr(sources))
        _write_config_value("TELEGRAM_ROLE", repr(self.tg_role_var.get()))
        try:
            _reload_cfg()
        except Exception:
            pass

        # Выключили Telegram — останавливаем чтение сразу, а не до перезапуска
        if not tg_on:
            tgr.stop()

        if not silent:
            messagebox.showinfo(APP_TITLE, "Источники сохранены.")
        self.refresh_sources_tab()
        self.refresh_news_tab()

    def refresh_sources_tab(self, apply_fields: bool = False):
        """Показывает по каждому источнику, работает он или нет и почему.

        apply_fields=True — сначала применить набранное на экране. Так
        работает кнопка «Проверить источники»: проверять СОХРАНЁННЫЕ
        значения, когда человек только что заполнил поля, — значит показать
        ему заведомо неверный ответ («Telegram выключен», хотя галочка на
        экране стоит)."""
        if apply_fields:
            try:
                self._apply_source_fields()
            except Exception as e:  # noqa: BLE001
                log.warning("Не удалось применить поля источников: %s", e)
        chain = list(news_calendar.news_source_chain())
        keys = getattr(cfg, "NEWS_API_KEYS", {}) or {}
        for name, var in self.src_status_vars.items():
            if name not in chain:
                var.set("выключен")
                continue
            if name == "mt5":
                try:
                    path = news_providers.mt5_calendar_path()
                    if os.path.exists(path):
                        age = (time.time() - os.path.getmtime(path)) / 60.0
                        var.set(f"файл обновлён {age:.0f} мин назад"
                                if age <= news_providers.MT5_CALENDAR_MAX_AGE_SECONDS / 60
                                else f"файл устарел ({age:.0f} мин) — сервис остановлен?")
                    else:
                        var.set("файла нет — запустите сервис CalendarExport")
                except Exception:
                    # Текст исключения тут не нужен: он повторяет то же самое
                    # своими словами и вылезает за край окна.
                    var.set("нет связи с терминалом MT5")
            elif name == "finnhub":
                var.set("готов" if keys.get("finnhub") else "нужен ключ")
            else:
                var.set("включён")

        if not tgs.enabled():
            self.tg_src_status_var.set("Выключено.")
        else:
            problem = tgr.preflight()
            if problem:
                self.tg_src_status_var.set(problem)
            else:
                if not tgr.is_running():
                    tgr.start()
                st = tgs.status()
                last = st.get("last_message")
                tail = f" Последнее сообщение: {last.strftime('%H:%M:%S')}." if last else ""
                self.tg_src_status_var.set(st.get("detail", "") + tail)

    # ---- вкладка "Сигналы TG" --------------------------------------------------------
    def _build_tab_telegram(self, parent):
        pad = {"padx": 10, "pady": 4}

        ttk.Label(parent, text="Сигналы из Telegram",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", **pad)
        ttk.Label(parent, foreground=self.colors["muted"], wraplength=800, justify="left", text=
                  "Что распознано из каналов и ботов. Подключение и выключатель — "
                  "на вкладке «Источники»."
                  ).pack(anchor="w", **pad)

        btns = ttk.Frame(parent)
        btns.pack(anchor="w", **pad)
        ttk.Button(btns, text="Обновить",
                   command=self.refresh_telegram_tab).grid(row=0, column=0)

        self.tg_status_var = tk.StringVar(value="")
        ttk.Label(parent, textvariable=self.tg_status_var, foreground=self.colors["muted"],
                  wraplength=800, justify="left").pack(anchor="w", **pad)

        ttk.Label(parent, text="Последние распознанные сигналы",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10)
        cols = ("time", "instrument", "direction", "text")
        headings = ("Время", "Инструмент", "Направление", "Сообщение")
        self.tg_tree = ttk.Treeview(parent, columns=cols, show="headings", height=12)
        widths = {"time": 80, "instrument": 100, "direction": 110, "text": 430}
        for col, head in zip(cols, headings):
            self.tg_tree.heading(col, text=head)
            self.tg_tree.column(col, width=widths[col],
                                anchor="w" if col == "text" else "center")
        self.tg_tree.pack(fill="both", expand=True, padx=10, pady=(2, 8))

        self.refresh_telegram_tab()

    def telegram_login(self):
        """Одноразовый вход по номеру телефона.

        Сам вход идёт в ФОНОВОМ потоке, а вопросы (код, пароль) задаются в
        потоке интерфейса через root.after. Раньше вход выполнялся прямо
        здесь, в потоке интерфейса: окно замирало на всё время входа, и
        диалог ввода кода мог не показаться вовсе — со стороны это выглядело
        как «программа зависла» или «войти не получается»."""
        # Сначала применяем то, что набрано на экране: иначе вход шёл бы по
        # СТАРЫМ значениям, а человек видел бы свои свежие в полях.
        applied = self._apply_source_fields()
        if applied:
            log.info("Источники: применены поля с экрана перед входом (%s)",
                     ", ".join(applied))

        problem = tgr.login_preflight()
        if problem:
            messagebox.showwarning(APP_TITLE, problem)
            return

        phone = simpledialog.askstring(
            APP_TITLE, "Номер телефона в международном формате (например +79991234567):",
            parent=self.root)
        if not phone:
            return

        def ask_in_gui(prompt: str, secret: bool = False) -> str:
            """Задать вопрос из фонового потока и дождаться ответа.

            Диалоги Tk можно открывать только в потоке интерфейса, поэтому
            вопрос ставится в его очередь, а фоновый поток ждёт на событии."""
            answer = {}
            done = threading.Event()

            def ask():
                try:
                    answer["value"] = simpledialog.askstring(
                        APP_TITLE, prompt, parent=self.root,
                        show="*" if secret else None) or ""
                except Exception:
                    answer["value"] = ""
                finally:
                    done.set()

            self.root.after(0, ask)
            done.wait()
            return answer.get("value", "")

        def worker():
            error = tgr.login(
                phone.strip(),
                lambda: ask_in_gui("Код подтверждения из Telegram:"),
                lambda: ask_in_gui("Облачный пароль Telegram (двухфакторная защита):",
                                   secret=True),
            )
            self.root.after(0, lambda: self._after_telegram_login(error))

        self._set_tg_status("Вход в Telegram: ждём подтверждения...")
        threading.Thread(target=worker, daemon=True, name="telegram-login").start()

    def _set_tg_status(self, text: str):
        """Строка состояния Telegram есть на двух вкладках — обновляем обе,
        какая бы из них ни была построена."""
        for name in ("tg_src_status_var", "tg_status_var"):
            var = getattr(self, name, None)
            if var is not None:
                try:
                    var.set(text)
                except Exception:
                    pass

    def _after_telegram_login(self, error: str):
        if error:
            messagebox.showerror(APP_TITLE, error)
            self._set_tg_status(error)
        else:
            messagebox.showinfo(APP_TITLE, "Вход выполнен. Код больше не понадобится.")
            tgr.start()
        self.refresh_telegram_tab()
        try:
            self.refresh_sources_tab()
        except Exception:
            pass

    def refresh_telegram_tab(self):
        if not tgs.enabled():
            self.tg_status_var.set("Чтение выключено. Поставьте галочку и нажмите «Сохранить».")
        else:
            problem = tgr.preflight()
            if problem:
                self.tg_status_var.set(problem)
            else:
                st = tgs.status()
                if not tgr.is_running():
                    tgr.start()
                    st = tgs.status()
                last = st.get("last_message")
                tail = f" Последнее сообщение: {last.strftime('%H:%M:%S')}." if last else ""
                self.tg_status_var.set(st.get("detail", "") + tail)

        for item in self.tg_tree.get_children():
            self.tg_tree.delete(item)
        for sig in tgs.history():
            self.tg_tree.insert("", "end", values=(
                sig["time"].strftime("%H:%M:%S"),
                sig["instrument"],
                "покупка" if sig["direction"] == tgs.BUY else "продажа",
                sig["text"].replace("\n", " ")[:120],
            ))

    # ---- вкладка "Календарь" ---------------------------------------------------------
    def _build_tab_schedule(self, parent):
        """Расписание работы бота: когда войдёт, когда нет и по какой новости.

        Всё, что здесь показано, считается модулем trading_schedule.py, который
        повторяет РЕАЛЬНЫЕ фильтры входа и берёт те же настройки. Это не
        отдельный «примерный» прогноз — если написано «не входит до 15:30»,
        бот действительно не войдёт."""
        # Боковой ползунок: владелец просил не растягивать окно
        # каждый раз — блоков здесь больше, чем помещается.
        parent = self._scrollable(parent)
        pad = {"padx": 10, "pady": 4}

        ttk.Label(parent, text="Расписание работы бота",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", **pad)

        # --- Карточка "прямо сейчас" ---
        status_box = ttk.Frame(parent)
        status_box.pack(fill="x", **pad)
        self.sched_status_var = tk.StringVar(value="Нажмите «Обновить»")
        self.sched_status_label = ttk.Label(status_box, textvariable=self.sched_status_var,
                                            font=("Segoe UI", 11, "bold"))
        self.sched_status_label.pack(anchor="w")
        self.sched_detail_var = tk.StringVar(value="")
        ttk.Label(status_box, textvariable=self.sched_detail_var, foreground=self.colors["muted"],
                  wraplength=780, justify="left").pack(anchor="w")

        btns = ttk.Frame(parent)
        btns.pack(anchor="w", **pad)
        ttk.Button(btns, text="Обновить", command=self.refresh_news_tab).grid(row=0, column=0)
        self.sched_updated_var = tk.StringVar(value="")
        ttk.Label(btns, textvariable=self.sched_updated_var, foreground=self.colors["dim"]).grid(
            row=0, column=1, padx=10)

        # --- График ---
        ttk.Label(parent, text="Ближайшие 12 часов", font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10)
        self.news_canvas = tk.Canvas(parent, height=NEWS_CHART_HEIGHT, bg=self.colors["card"],
                                     highlightthickness=0)
        self.news_canvas.pack(fill="x", padx=10, pady=(2, 6))
        # Перерисовываем при изменении ширины окна: координаты считаются от
        # фактической ширины, иначе график остался бы обрезанным.
        self.news_canvas.bind("<Configure>", lambda e: self._draw_news_chart())
        self._news_events_cache = []

        # --- Таблица расписания ---
        ttk.Label(parent, text="Что и когда остановит торговлю",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10)
        cols = ("window", "action", "event", "currency", "symbols")
        headings = ("Окно", "Что делает бот", "Новость", "Валюта", "Затронутые пары")
        self.sched_tree = ttk.Treeview(parent, columns=cols, show="headings", height=9)
        widths = {"window": 120, "action": 110, "event": 240, "currency": 70, "symbols": 160}
        for col, head in zip(cols, headings):
            self.sched_tree.heading(col, text=head)
            self.sched_tree.column(col, width=widths[col],
                                   anchor="w" if col in ("event", "symbols") else "center")
        self.sched_tree.pack(fill="both", expand=True, padx=10, pady=(2, 4))
        # Идущее прямо сейчас окно подсвечиваем — его легко не заметить в списке
        self.sched_tree.tag_configure("now", foreground=self.colors["loss"])

        self.sched_free_var = tk.StringVar(value="")
        ttk.Label(parent, textvariable=self.sched_free_var, foreground=self.colors["profit"],
                  wraplength=780, justify="left").pack(anchor="w", padx=10, pady=(0, 8))

    def _watched_symbols(self) -> list:
        """Пары, по которым бот реально работает. Список с дашборда точнее
        конфига: там учтены отключённые вручную символы."""
        try:
            symbols = list(ds.get_snapshot().get("symbols", {}).keys())
        except Exception:
            symbols = []
        return symbols or list(getattr(cfg, "SYMBOLS", []))

    def _apply_schedule(self, events):
        symbols = self._watched_symbols()
        now = datetime.now()

        status = tsched.current_status(symbols, events, now)
        if status["trading"]:
            self.sched_status_var.set("Сейчас: торгует")
            try:
                self.sched_status_label.configure(foreground=self.colors["profit"])
            except Exception:
                pass
        else:
            # Причина уже сформулирована целиком ("Выходной — рынок закрыт"),
            # приписывать к ней ещё одно тире незачем.
            self.sched_status_var.set(f"Сейчас: не входит. {status['reason']}")
            try:
                self.sched_status_label.configure(foreground=self.colors["loss"])
            except Exception:
                pass

        detail = status["detail"]
        nxt = tsched.next_block(symbols, events, now)
        if nxt:
            detail += (f"  Ближайшая пауза: {nxt['start'].strftime('%H:%M')}–"
                       f"{nxt['end'].strftime('%H:%M')} ({nxt['event']}).")
        elif status["trading"]:
            detail += "  Известных пауз по новостям впереди нет."
        if news_providers.looks_like_broken_encoding(events):
            detail += "\n\n" + news_providers.BROKEN_ENCODING_HINT
        self.sched_detail_var.set(detail)

        for item in self.sched_tree.get_children():
            self.sched_tree.delete(item)

        rows = tsched.build_schedule(symbols, events, now, hours_ahead=48)
        impact_ru = {"high": "важная", "medium": "средняя", "low": "слабая"}
        for r in rows:
            window = f"{r['start'].strftime('%d.%m %H:%M')}–{r['end'].strftime('%H:%M')}"
            action = tsched.ACTION_TITLES.get(r["action"], r["action"])
            if r["active_now"]:
                action += " (идёт)"
            self.sched_tree.insert(
                "", "end",
                tags=("now",) if r["active_now"] else (),
                values=(window, action, f"{r['event']} ({impact_ru.get(r['impact'], '')})",
                        r["currency"], ", ".join(r["symbols"])))

        if not rows:
            if not tsched.news_filter_enabled():
                self.sched_free_var.set(
                    "Фильтр новостей выключен (USE_NEWS_FILTER = False) — бот не будет "
                    "останавливаться перед выходом данных.")
            else:
                self.sched_free_var.set(
                    "Новостей, затрагивающих ваши пары, в ближайшие двое суток не найдено.")
        else:
            free = tsched.quiet_windows(symbols, events, now, hours_ahead=12)
            if free:
                parts = [f"{a.strftime('%H:%M')}–{b.strftime('%H:%M')}" for a, b in free[:4]]
                self.sched_free_var.set("Спокойные окна (12 ч): " + ", ".join(parts))
            else:
                self.sched_free_var.set("Спокойных окон в ближайшие 12 часов нет.")

        self.sched_updated_var.set("обновлено " + now.strftime("%H:%M:%S"))

    # ---- график календаря ------------------------------------------------------------
    def _draw_news_chart(self):
        """Временная шкала на NEWS_CHART_HOURS часов вперёд: каждое событие —
        засечка, важные ещё и с заштрихованной зоной, в которой торговля
        блокируется фильтром новостей. Смысл графика именно в этих зонах:
        видно не только "когда новость", но и "когда бот не будет входить"."""
        canvas = getattr(self, "news_canvas", None)
        if canvas is None:
            return
        canvas.delete("all")

        width = canvas.winfo_width()
        if width < 50:          # окно ещё не разложено — рисовать нечего
            return
        h = NEWS_CHART_HEIGHT
        axis_y = h - 18
        left_pad, right_pad = 8, 8
        span = width - left_pad - right_pad

        now = datetime.now()
        horizon = timedelta(hours=NEWS_CHART_HOURS)

        def x_of(t):
            frac = (t - now).total_seconds() / horizon.total_seconds()
            return left_pad + max(0.0, min(1.0, frac)) * span

        # Ось и часовые метки
        canvas.create_line(left_pad, axis_y, width - right_pad, axis_y, fill=self.colors["dim"])
        step_hours = 1 if NEWS_CHART_HOURS <= 12 else 3
        hour = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
        while hour <= now + horizon:
            x = x_of(hour)
            canvas.create_line(x, axis_y - 4, x, axis_y + 4, fill=self.colors["dim"])
            canvas.create_text(x, axis_y + 11, text=hour.strftime("%H:%M"),
                               fill=self.colors["muted"], font=("Segoe UI", 7))
            hour += timedelta(hours=step_hours)

        block_minutes = getattr(cfg, "NEWS_HARD_BLOCK_WINDOW_MIN", 30)
        colors = {"high": self.colors["loss"], "medium": self.colors["warning"], "low": self.colors["accent"]}

        # Показываем ТОЛЬКО события, затрагивающие ваши пары — ровно те же, что
        # попадают в таблицу расписания. Иначе график рисовал бы красную зону
        # блокировки, скажем, на японской статистике, а бот по EURUSD и золоту
        # в этот момент спокойно торговал бы: график обещал бы то, чего не будет.
        watched = self._watched_symbols()
        events = [e for e in self._news_events_cache
                  if now <= e["time"] <= now + horizon and tsched.affected_symbols(watched, e)]
        # Зоны блокировки рисуем ПЕРВЫМ проходом, чтобы прямоугольники не легли
        # поверх засечек соседних событий.
        for e in events:
            if e.get("impact") == "high":
                x1 = x_of(e["time"] - timedelta(minutes=block_minutes))
                x2 = x_of(e["time"] + timedelta(minutes=block_minutes))
                canvas.create_rectangle(x1, 6, x2, axis_y, fill=self.colors["card"], outline="")

        # Последняя занятая координата на каждой высоте — чтобы подписи валют
        # у событий, идущих подряд, не наезжали друг на друга.
        taken = {}
        for e in events:
            impact = e.get("impact", "low")
            colour = colors.get(impact, colors["low"])
            x = x_of(e["time"])

            top = 10 if impact == "high" else (20 if impact == "medium" else 30)
            canvas.create_line(x, top, x, axis_y, fill=colour, width=2 if impact == "high" else 1)

            # Ищем свободную строку выше; если её нет — подпись не рисуем вовсе.
            # Обрезанная у края канвы подпись хуже, чем её отсутствие: засечка
            # события в любом случае на месте, а детали видно в таблице ниже.
            label_y = top - 4
            while label_y >= 6:
                if taken.get(label_y, -999) <= x - NEWS_LABEL_MIN_GAP_PX:
                    canvas.create_text(x, label_y, text=e.get("currency", ""), fill=colour,
                                       font=("Segoe UI", 7, "bold"))
                    taken[label_y] = x
                    break
                label_y -= 9

        # Отметка "сейчас"
        canvas.create_line(left_pad, 6, left_pad, axis_y, fill=self.colors["profit"], width=2)
        canvas.create_text(left_pad + 4, 8, text="сейчас", anchor="nw",
                           fill=self.colors["profit"], font=("Segoe UI", 7))

        if not events:
            canvas.create_text(width / 2, h / 2 - 6,
                               text=f"Ближайшие {NEWS_CHART_HOURS} ч — событий по вашим парам нет",
                               fill=self.colors["dim"], font=("Segoe UI", 9))
        else:
            canvas.create_text(width - right_pad, 8, anchor="ne",
                               text="красное — торговля заблокирована",
                               fill=self.colors["loss"], font=("Segoe UI", 7))

    def fix_news_source(self):
        """Проверить цепочку новостей и починить то, что чинится само.

        Установка и сборка сервиса делаются программой; единственное, что
        приходится нажать руками, — первый запуск сервиса в Навигаторе
        терминала: снаружи MetaTrader сервисы запускать не даёт."""
        self.news_source_var.set("Проверяю источник новостей...")

        def worker():
            try:
                news_autostart.reset_checks()
                done = news_autostart.repair(
                    progress=lambda t: self.root.after(
                        0, lambda t=t: self.news_source_var.set(t)))
                state = news_autostart.check()
                text = news_autostart.describe(state)
                if done:
                    text = "Сделано: " + "; ".join(done) + ". " + text
            except Exception as e:  # noqa: BLE001
                text = f"Проверить источник не удалось: {e}"
            self.root.after(0, lambda: self.news_source_var.set(text))

        threading.Thread(target=worker, daemon=True, name="news-source").start()

    def explain_news_trading(self):
        """По каждой паре — что сейчас с новостями. Владелец: «мне нужно,
        чтобы работала каждая новость» и «я не заметил за ним этого».

        Заметить было нельзя: если входа не случилось, программа молчала.
        Здесь она проходит те же проверки, что и при торговле, и говорит,
        на какой именно остановилась."""
        self.news_source_var.set("Проверяю новостную торговлю...")

        def worker():
            lines = []
            try:
                for symbol in list(getattr(cfg, "SYMBOLS", []) or [])[:8]:
                    try:
                        lines.append(f"{symbol}: "
                                     + news_calendar.explain_news_entry(symbol))
                    except Exception as e:  # noqa: BLE001
                        lines.append(f"{symbol}: проверить не удалось ({e})")
            except Exception as e:  # noqa: BLE001
                lines = [f"Проверка не прошла: {e}"]
            text = "\n\n".join(lines) or "Инструменты не заданы."
            self.root.after(0, lambda: messagebox.showinfo(
                APP_TITLE, "Новостная торговля сейчас:\n\n" + text))
            self.root.after(0, self.refresh_news_source_line)

        threading.Thread(target=worker, daemon=True, name="news-explain").start()

    def refresh_news_source_line(self):
        """Показать состояние источника, ничего не меняя (при открытии вкладки)."""
        def worker():
            try:
                text = news_autostart.describe(news_autostart.check())
            except Exception as e:  # noqa: BLE001
                text = f"Состояние источника неизвестно: {e}"
            self.root.after(0, lambda: self.news_source_var.set(text))

        threading.Thread(target=worker, daemon=True, name="news-source-line").start()

    def _news_auto_refresh(self):
        """Сам обновляет календарь раз в NEWS_AUTO_REFRESH_MINUTES минут.

        Владелец: «и календарь сам обновлялся с новостями». Раньше события
        подтягивались только по кнопке — если вкладку не открывать, бот
        работал на календаре, загруженном при запуске, и мог не знать про
        новость, вышедшую час назад.

        Заодно перепроверяется сам источник (news_autostart): сервис в
        терминале могли остановить, и календарь тихо протух бы."""
        minutes = float(getattr(cfg, "NEWS_AUTO_REFRESH_MINUTES", 15) or 0)
        if minutes <= 0:
            return          # 0 = обновлять само не нужно
        try:
            news_autostart.ensure_ready()
            self.refresh_news_tab()
        except Exception as e:  # noqa: BLE001
            log.warning("Автообновление календаря не прошло: %s", e)
        finally:
            # Следующий раз планируем ВСЕГДА, даже после ошибки: временный
            # сбой сети не должен навсегда отключать автообновление.
            self.root.after(int(minutes * 60_000), self._news_auto_refresh)

    def refresh_news_tab(self):
        self.news_status_var.set("Загружаю...")
        self.refresh_news_source_line()
        threading.Thread(target=self._refresh_news_worker, daemon=True).start()

    def _refresh_news_worker(self):
        events, used, error = news_calendar.get_events_with_source()
        now = datetime.now()
        horizon = now + timedelta(days=3)
        rank = {"low": 0, "medium": 1, "high": 2}

        # Таблица "Новости" — только предстоящие события.
        upcoming = [e for e in events
                    if rank.get(e["impact"], 0) >= 1 and now <= e["time"] <= horizon]
        # Расписание — ВСЕ события, включая только что прошедшие: окно блокировки
        # вокруг новости, вышедшей 10 минут назад, ещё идёт, и именно его важнее
        # всего показать. Отфильтруй мы их здесь — строка "идёт сейчас" не
        # появилась бы никогда.
        recent = now - timedelta(hours=6)
        for_schedule = [e for e in events if recent <= e["time"] <= horizon]

        self.root.after(0, lambda: self._apply_news_result(upcoming, for_schedule, used, error))

    def _apply_news_result(self, events, for_schedule, used, error):
        for item in self.news_tree.get_children():
            self.news_tree.delete(item)

        now = datetime.now()
        impact_ru = {"high": "важная", "medium": "средняя", "low": "слабая"}
        for e in events:
            left = e["time"] - now
            hours, rem = divmod(int(left.total_seconds()), 3600)
            left_txt = f"{hours} ч {rem // 60} мин" if hours else f"{rem // 60} мин"
            self.news_tree.insert("", "end", values=(
                e["time"].strftime("%d.%m %H:%M"), left_txt, e["currency"], e["event"],
                impact_ru.get(e["impact"], e["impact"]),
                e.get("actual", ""), e.get("estimate", ""), e.get("prev", ""),
            ))

        self._news_events_cache = for_schedule
        self._draw_news_chart()
        try:
            self._apply_schedule(for_schedule)
        except Exception as e:
            # Расписание — справочная вкладка; её поломка не должна ронять
            # вкладку "Новости" вместе с настройками источников.
            log.exception("Не удалось построить расписание: %s", e)

        source_txt = news_providers.PROVIDER_TITLES.get(used, used) if used else ""
        if error:
            self.news_status_var.set(error)
        elif news_providers.looks_like_broken_encoding(events):
            # Названия пришли как "??????" — файл записан старой версией
            # сервиса. Молчать нельзя: человек будет думать, что дело в шрифте.
            self.news_status_var.set(news_providers.BROKEN_ENCODING_HINT)
        elif not events:
            self.news_status_var.set(
                f"Источник: {source_txt}. Важных событий в ближайшие 3 дня нет.")
        else:
            self.news_status_var.set(f"Источник: {source_txt}. Событий: {len(events)}")

    # ---- вкладка "Chat AI" ------------------------------------------------------------
    def _build_tab_chat(self, parent):
        ttk.Label(parent, text="Чат с Claude", font=("Segoe UI", 12, "bold")).pack(padx=10, pady=6, anchor="w")

        text_frame = ttk.Frame(parent)
        text_frame.pack(fill="both", expand=True, padx=10, pady=6)
        self.chat_text = tk.Text(text_frame, wrap="word", state="disabled", bg=self.colors["card"], fg=self.colors["fg"],
                                  insertbackground=self.colors["fg"])
        self.chat_text.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(text_frame, command=self.chat_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.chat_text.config(yscrollcommand=scrollbar.set)

        input_frame = ttk.Frame(parent)
        input_frame.pack(fill="x", padx=10, pady=6)
        self.chat_entry = ttk.Entry(input_frame)
        self.chat_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self.chat_entry.bind("<Return>", lambda e: self.send_chat_message())
        self.chat_send_btn = ttk.Button(input_frame, text="Отправить", command=self.send_chat_message)
        self.chat_send_btn.pack(side="left")

    def _append_chat(self, who: str, text: str):
        self.chat_text.config(state="normal")
        self.chat_text.insert("end", f"{who}: {text}\n\n")
        self.chat_text.see("end")
        self.chat_text.config(state="disabled")

    def send_chat_message(self):
        msg = self.chat_entry.get().strip()
        if not msg:
            return
        self.chat_entry.delete(0, "end")
        self._append_chat("Ты", msg)
        self.chat_history.append({"role": "user", "content": msg})
        self.chat_send_btn.config(state="disabled")
        threading.Thread(target=self._chat_worker, daemon=True).start()

    def _chat_worker(self):
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=cfg.ANTHROPIC_API_KEY)
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=800,
                messages=self.chat_history[-20:],
            )
            answer = resp.content[0].text
        except Exception as e:
            answer = f"[Ошибка обращения к Claude: {e}]"
        self.chat_history.append({"role": "assistant", "content": answer})
        self.root.after(0, lambda: self._append_chat("Claude", answer))
        self.root.after(0, lambda: self.chat_send_btn.config(state="normal"))

    # ---- вкладка "Как пользоваться" (всегда видна) --------------------------------
    def _build_tab_help(self, parent):
        # Боковой ползунок: владелец просил не растягивать окно
        # каждый раз — блоков здесь больше, чем помещается.
        outer = ttk.Frame(parent)
        outer.pack(fill="both", expand=True, padx=6, pady=6)
        text = tk.Text(outer, wrap="word", bg=self.colors["bg"], fg=self.colors["fg"], insertbackground=self.colors["fg"],
                        relief="flat", padx=14, pady=10, font=("Segoe UI", 10))
        scrollbar = ttk.Scrollbar(outer, command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)
        text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        text.tag_configure("h1", font=("Segoe UI", 13, "bold"), foreground=self.colors["profit"], spacing3=6)
        text.tag_configure("h2", font=("Segoe UI", 11, "bold"), foreground=self.colors["fg"], spacing1=10, spacing3=4)
        text.tag_configure("body", font=("Segoe UI", 10), foreground=self.colors["fg"], spacing3=4)

        def h1(s):
            text.insert("end", s + "\n", "h1")

        def h2(s):
            text.insert("end", s + "\n", "h2")

        def p(s):
            text.insert("end", s + "\n", "body")

        h1("AI Scalper Pro — как пользоваться программой")
        p("Подробная инструкция по всем вкладкам и функциям, по порядку.")

        h2("1. Экран входа")
        p("Логин и пароль — те же, что заданы в настройках (DASHBOARD_LOGIN/пароль). "
          "Галочка «Запомнить пароль» — сохраняет пароль на ЭТОМ компьютере (Windows "
          "сам защищает файл, привязка к учётной записи Windows). При следующем "
          "запуске пароль будет уже подставлен в поле — просто нажми «Войти». Число "
          "попыток входа не ограничено.")

        h2("2. Обзор")
        p("Главная вкладка. Переключатель «Простой/Продвинутый» — простой режим "
          "показывает только самое нужное (Обзор/Символы/Сделки/Лог/Настройка/Как "
          "пользоваться), продвинутый добавляет ещё Брокер/Equity/Новости/Chat AI. "
          "Кнопки «Старт»/«Стоп» управляют торговым циклом вручную (хотя он и "
          "так запускается сам сразу при входе). «Полный выход» — гарантированно "
          "закрывает программу целиком (не путать с обычным закрытием окна, которое "
          "просто сворачивает программу в трей). Здесь же видно баланс, эквити, "
          "профиль риска, режим торговли и краткую статистику.")

        h2("3. Брокер (продвинутый режим)")
        p("Данные подключения к MT5: сервер/логин/пароль твоего брокера. Если "
          "оставить поля пустыми — программа подключается к уже открытому и "
          "залогиненному терминалу MT5. Если заполнить — программа сама запускает "
          "терминал и логинится этими данными при каждом старте. Кнопка «Проверить "
          "подключение» временно недоступна, пока бот работает (чтобы не дёргать MT5 "
          "из двух мест одновременно).")

        h2("4. Символы")
        p("Список торгуемых инструментов. Вписывать пары руками БОЛЬШЕ НЕ НУЖНО: "
          "при запуске программа берёт весь список брокера и сама отбирает "
          "подходящие — по карману ли минимальный лот, какую долю движения "
          "съедает спред, есть ли движение, разрешена ли торговля. Сверху "
          "написано, сколько пар в работе и из скольких они выбраны.\n\n"
          "Двойной клик по названию символа открывает мини-график цены, по «Вкл» — "
          "выключает пару вручную. Видно score BUY/SELL, режим рынка (тренд/флэт), "
          "сигнал AI, статус автообучения и причину, по которой последний раз не "
          "открылась сделка.\n\n"
          "Колонка «Риск лота» — предупреждение на маленьком депозите: минимальный "
          "лот брокера иногда рискует БОЛЬШЕ, чем разрешает настроенный процент "
          "риска (ниже минимального лота опуститься нельзя, сколько ни настраивай "
          "риск). Строка подсвечивается красным, а в колонке — точная цифра: "
          "«минимальный лот X рискует Y — в Z раз больше настроенного риска».")

        h2("5. Счета")
        p("Несколько торговых счетов сразу: у каждого свой логин, сервер, "
          "инструменты и лимиты риска. Кнопки запуска/остановки, «Закрыть все "
          "позиции», только прибыльные, только убыточные. Красная кнопка "
          "«Закрыть всё на всех счетах» — для аварийной ситуации.\n\n"
          "Если строка счёта помечена значком 🔒 и подсвечена красным — пароль "
          "счёта не расшифрован ТЕКУЩИМ паролем входа (обычно потому что "
          "программа открылась без экрана входа). Пароль НЕ потерян: войдите "
          "с правильным паролем и откройте вкладку заново.\n\n"
          "Кнопки «☁ Сохранить счета в облако» / «☁ Восстановить из облака» — "
          "резервная копия списка счетов в том же закрытом репозитории, что и "
          "журнал сделок (вкладка «Система»). Список счетов НЕ входит в git и "
          "не трогается обновлением программы — без резервной копии он не "
          "переживёт переустановку или перенос на другой компьютер.")

        h2("6. Сделки")
        p("Все ОТКРЫТЫЕ позиции на счёте — не только те, что открыл этот бот, но и "
          "открытые вручную в терминале MT5 (колонка «Источник»: Бот/Ручная). Кнопка "
          "«Закрыть выбранную сделку» закрывает позицию по рынку — работает для любой "
          "строки в таблице.")

        h2("7. Лог")
        p("История сделок. Верхняя таблица — журнал этого бота (что и когда бот сам "
          "открывал/закрывал, с score). Нижняя таблица — синхронизированная история "
          "из MT5 (раздел «Синхронизация с MetaTrader»): подтягивается напрямую из "
          "брокера раз в минуту, включает вообще ВСЕ закрытые сделки за последние "
          "30 дней (в т.ч. открытые вручную), и статистика (винрейт, профит-фактор) "
          "там всегда 100% совпадает с историей у брокера.")

        h2("8. Equity (продвинутый режим)")
        p("График изменения эквити счёта с момента запуска программы.")

        h2("9. Настройка (видна всегда)")
        p("Сверху — быстрые переключатели: профиль риска (Консервативный/"
          "Сбалансированный/Агрессивный/Истеричка), режим торговли (Скальпинг/"
          "Новости/Оба), пауза новых сделок, звук+всплывающие уведомления.\n"
          "Ниже — ПОЛНЫЙ список входных параметров торговой логики, один в один "
          "как окно «Inputs» в MQL5-советнике: индикаторы, price action, объём, "
          "автонастройка под инструмент, защитные фильтры, AI-сигнал, режим/"
          "контекст рынка, score-фильтр, стопы/TP, Break Even, трейлинг, Profit "
          "Lock, просадка/серии убытков, издержки, новости, автообучение, "
          "автообновление, часы торговли, лот, частичное закрытие — плюс отдельный "
          "редактор каждого профиля риска и корреляций контекста рынка по символам. "
          "«Сохранить» применяет изменения сразу, без перезапуска программы.\n\n"
          "Рядом с КАЖДЫМ полем есть кнопка «?». Она открывает описание: что "
          "параметр делает, какое значение стандартное и что изменится, если "
          "его увеличить или уменьшить. Если сомневаетесь — верните "
          "стандартное значение, оно написано в той же справке.\n\n"
          "Настройки, появившиеся в новой версии программы, дописываются в ваш "
          "файл настроек сами, со стандартными значениями. Ваши собственные "
          "значения при этом не трогаются.")

        h2("10. Новости (продвинутый режим)")
        p("Источник экономического календаря (провайдер + API-ключ) и таблица "
          "предстоящих новостей. Пока ключ не задан — фильтр по новостям просто "
          "не влияет на торговлю (безопасное поведение по умолчанию).")

        h2("11. Chat AI (продвинутый режим)")
        p("Обычный чат с Claude прямо в программе — можно спросить что угодно про "
          "рынок или настройки, не переключаясь на другое окно.")

        h2("12. Система: обновление без переустановки")
        p("Вкладка «Система» → «Обновление из GitHub». Репозиторий и проверка "
          "обновлений уже включены по умолчанию — программа сама обновляется "
          "из того же репозитория, откуда приехала, вписывать ничего не "
          "нужно. Если репозиторий закрытый — впишите токен (см. подсказку "
          "под полем) и нажмите «Обновить всё сейчас». Одним нажатием "
          "обновятся: советники в MetaTrader (сразу, без перезапуска), файлы "
          "самой программы и новые настройки.\n\n"
          "Если готовой сборки ещё нет — кнопка «Собрать новую версию»: "
          "программа сама попросит GitHub собрать её (5-10 минут на их "
          "серверах, ваш компьютер не нагружается). Раньше для этого надо было "
          "заходить на сайт во вкладку Actions. Токену нужно право "
          "Actions: Read and write.\n\n"
          "Галочка «Ставить обновление само при запуске» — программа всё "
          "сделает и перезапустится сама, ничего не спрашивая. При старте "
          "торговля ещё не началась и открытых сделок нет, поэтому подменять "
          "её в этот момент безопасно.\n\n"
          "Чего обновление НЕ делает: не ставится посреди работы под открытыми "
          "сделками (спросит); не трогает ваши настройки, счета, пароли, "
          "журналы и сессию Telegram; не ставит половину новой версии — если "
          "хоть один файл не скачался, не заменяется ни один, а старые "
          "сохраняются рядом с припиской .bak.")

        h2("13. Система: журнал сделок в облаке")
        p("Вкладка «Система» → «Журнал сделок в облаке». Программа выкладывает "
          "историю сделок в папку journal/ вашего ЗАКРЫТОГО репозитория "
          "GitHub — три файла: журнал бота, реальные закрытые сделки из "
          "MetaTrader (с временем жизни каждой сделки) и разбор обычными "
          "словами: винрейт, средний плюс и средний минус, сколько сделок "
          "умерло за секунды, какая пара даёт минус.\n\n"
          "Зачем: историю можно открыть с телефона или показать, когда рабочий "
          "компьютер выключен. Время жизни сделки — главная улика: сделки, "
          "закрывающиеся за 8-11 секунд, означают стоп внутри рыночного шума, "
          "а не «неудачный вход».\n\n"
          "Выключено по умолчанию. Нужен токен GitHub с правом записи "
          "(Contents: Read and write). В облако уходят ТОЛЬКО сделки: ни "
          "паролей от счёта, ни ключей, ни токенов, ни файла настроек. Номер "
          "счёта можно скрыть галочкой в настройках (JOURNAL_MASK_ACCOUNT).")

        h2("Значок в трее и автозапуск")
        p("Закрытие окна крестиком сворачивает программу в трей, а не закрывает её — "
          "торговля продолжается в фоне. Из меню на значке в трее — Старт/Стоп/"
          "Показать окно/Полный выход. Галочка «Запускать вместе с Windows» "
          "(в «Обзоре») включает автозапуск при включении компьютера.")

        h2("Безопасность")
        p("Пароль MT5 и ключи AI/новостей хранятся в config.py в зашифрованном виде "
          "(не открытым текстом) — расшифровываются только в памяти после входа. "
          "config.py переписывается атомарно с резервными копиями, чтобы сбой "
          "посреди записи (антивирус, отключение питания) не испортил файл.")

        text.config(state="disabled")

    # ---- запуск/остановка торгового движка --------------------------------
    # Сколько секунд без признаков жизни считаем «цикл встал». Круг цикла
    # занимает POLL_SECONDS (обычно 5 с) плюс запросы к терминалу; берём с
    # большим запасом, чтобы медленный ответ брокера не считался смертью.
    WATCHDOG_SILENCE_SECONDS = 180

    def _silence_reasons(self, snap: dict) -> list:
        """Почему сейчас нет входов — словами, без открытия других вкладок.

        Жалоба владельца: «перезапустил программу и начала открывать сделки,
        до этого затишье было». Причина у таких затиший обычно вполне
        конкретная, и программа её знает — просто держала при себе.

        Особый случай, ради которого всё и сделано: запреты по просадке и
        пауза после серии убытков живут ТОЛЬКО В ПАМЯТИ. Перезапуск их
        снимает, и со стороны это выглядит как «программа подвисла, помог
        перезапуск». На самом деле сработала защита, и знать об этом надо."""
        symbols = snap.get("symbols") or []
        counted = {}
        for item in symbols:
            reason = str((item or {}).get("reject_reason", "") or "").strip()
            if not reason or reason.startswith("OK"):
                continue
            counted[reason] = counted.get(reason, 0) + 1
        if not counted:
            return []
        # Если все пары молчат по одной причине — показываем её одной строкой,
        # а не повторяем на каждую пару.
        total = len([i for i in symbols if i])
        lines = []
        for reason, count in sorted(counted.items(), key=lambda kv: -kv[1])[:3]:
            where = "по всем парам" if count >= total and total else f"пар: {count}"
            lines.append(f"{reason} ({where})")
        return lines

    def _watchdog_tick(self):
        """Сторож торгового цикла.

        Жалоба владельца: «работает пару часов и всё, потом надо перезапуск
        приложения». Причина найдена в main.py: вызов монитора позиций стоял
        СНАРУЖИ защиты от ошибок, а внешний перехват ловил только Ctrl+C.
        Любая неожиданная ошибка там молча выводила цикл из работы: поток
        умирал, окно продолжало показывать «Работает», сделки не открывались.

        Одной починки мало: поток может не умереть, а ЗАВИСНУТЬ внутри
        запроса к терминалу — снаружи это выглядит так же. Поэтому смотрим
        на пульс: время последнего пройденного круга (main.last_heartbeat).
        Нет пульса дольше WATCHDOG_SILENCE_SECONDS — поднимаем цикл заново
        сами, не дожидаясь, пока человек заметит и перезапустит программу."""
        if not getattr(self, "_bot_should_run", False):
            return              # человек сам нажал «Стоп» — не мешаем
        alive = bool(self.bot_thread and self.bot_thread.is_alive())
        silent = 0.0
        try:
            silent = bot_engine.seconds_since_heartbeat()
        except Exception:  # noqa: BLE001
            silent = 0.0
        reason = bot_engine.watchdog_reason(
            True, alive, silent, self.WATCHDOG_SILENCE_SECONDS)
        if not reason:
            return

        log.warning("Сторож: %s — перезапускаю торговый цикл сам.", reason)
        runtime_events.record("сторож", f"{reason} — цикл перезапущен программой")
        self.status_var.set(f"Перезапуск: {reason}")
        # Старый цикл (если он завис) просим остановиться и заводим новый.
        # Поток-зомби демонический, программу он держать не будет.
        if self.stop_event:
            self.stop_event.set()
        self.bot_thread = None
        self.start_bot()

    def start_bot(self):
        if self.bot_thread and self.bot_thread.is_alive():
            return
        # Помечаем НАМЕРЕНИЕ: бот должен работать. По нему сторож ниже
        # отличает «человек нажал Стоп» от «цикл умер сам».
        self._bot_should_run = True
        self.stop_event = threading.Event()
        self.bot_thread = threading.Thread(target=self._run_bot, daemon=True)
        self.bot_thread.start()
        self.status_var.set("Запускается...")
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")

    def _run_bot(self):
        """Тело торгового потока.

        Ничего из tkinter отсюда напрямую не трогаем: обновление виджетов из
        чужого потока — само по себе источник зависаний, а раньше здесь ещё
        и открывалось модальное окно с ошибкой, которое НИКТО не мог закрыть,
        пока человек не подойдёт к компьютеру. Всё, что нужно показать,
        передаём в поток окна через root.after."""
        try:
            bot_engine.main(stop_event=self.stop_event, start_dashboard=False)
        except Exception as e:  # noqa: BLE001
            log.exception("Бот остановился с ошибкой: %s", e)
            self.root.after(0, lambda e=e: self.status_var.set(f"Ошибка: {e}"))
        finally:
            def done():
                if getattr(self, "_bot_should_run", False):
                    # Останов не по нашей воле — сторож поднимет цикл заново
                    return
                if "Ошибка" not in self.status_var.get():
                    self.status_var.set("Остановлен")
                self.start_btn.config(state="normal")
                self.stop_btn.config(state="disabled")
            try:
                self.root.after(0, done)
            except Exception:  # noqa: BLE001
                pass

    def toggle_pause(self):
        """Пауза/продолжить. Пауза НЕ убивает торговый цикл: он продолжает
        вести уже открытые сделки (трейлинг, безубыток, частичное закрытие),
        но новых не открывает. Это разное — «постоять» и «выключить»."""
        paused = not control.is_paused()
        control.set_paused(paused)
        runtime_events.record(
            "управление", "пауза включена" if paused else "пауза снята")
        self._refresh_top_bar()

    # Сколько раз проверяем, остановился ли старый цикл, и с каким шагом.
    # 40 раз по 250 мс — это десять секунд: достаточно, чтобы цикл завершил
    # текущий проход и закрыл соединение, и не настолько долго, чтобы человек
    # решил, что кнопка не работает.
    RESTART_CHECKS = 40
    RESTART_STEP_MS = 250

    @staticmethod
    def restart_decision(thread_alive: bool, checks_left: int) -> str:
        """Что делать при перезапуске: «ждать», «запускать» или «сдаться».

        ПОЧЕМУ ЭТО РЕШЕНИЕ ВЫНЕСЕНО ОТДЕЛЬНО. Владелец: «кнопка перезапуск не
        работает». Раньше здесь стояла глухая пауза в 600 мс, после которой
        цикл запускался заново независимо от того, остановился старый или нет.

        Так делать нельзя, и вот почему. Соединение с терминалом ОДНО на всю
        программу. Старый цикл, завершаясь, закрывает его (mt5c.disconnect в
        конце main). Если он успеет сделать это ПОСЛЕ того, как новый цикл уже
        подключился, новый останется без связи — и будет писать «потеряно
        соединение с MT5», а сделок не будет. Пауза в 600 мс превращала это в
        подбрасывание монетки.

        Правило простое: пока старый поток жив — ждём. Не дождались за
        отведённое время — НЕ запускаем второй поверх первого, а говорим
        человеку правду."""
        if not thread_alive:
            return "запускать"
        if checks_left > 0:
            return "ждать"
        return "сдаться"

    def restart_bot(self):
        """Перезапуск торгового цикла — то, что владелец делал руками,
        закрывая и открывая программу. Окно и настройки при этом остаются
        на месте, перезапускается только сам цикл."""
        runtime_events.record("управление", "перезапуск цикла по кнопке")
        self._bot_should_run = False
        try:
            if self.stop_event:
                self.stop_event.set()
        except Exception:  # noqa: BLE001
            pass
        control.set_paused(False)
        self.top_status_var.set("Перезапуск: жду остановки старого цикла...")
        self._restart_when_stopped(self.RESTART_CHECKS)

    def _restart_when_stopped(self, checks_left: int):
        """Дождаться остановки старого цикла и только потом запустить новый."""
        alive = bool(self.bot_thread and self.bot_thread.is_alive())
        решение = self.restart_decision(alive, checks_left)

        if решение == "ждать":
            self.root.after(self.RESTART_STEP_MS,
                            lambda: self._restart_when_stopped(checks_left - 1))
            return

        if решение == "сдаться":
            # Запустить второй цикл поверх работающего было бы хуже всего:
            # два цикла ведут одни и те же сделки и оба двигают стоп-лосс.
            self.top_status_var.set("Перезапуск не удался")
            runtime_events.record(
                "управление", "перезапуск не удался: прежний цикл не "
                              "остановился за 10 секунд")
            messagebox.showwarning(
                APP_TITLE,
                "Прежний торговый цикл не остановился за 10 секунд.\n\n"
                "Второй цикл поверх него не запускаю: два цикла вели бы одни "
                "и те же сделки и оба двигали бы стоп-лосс.\n\n"
                "Нажмите «Стоп», дождитесь надписи «Остановлен», затем "
                "«Старт».")
            return

        self.bot_thread = None
        self.start_bot()

    def _refresh_top_bar(self):
        """Состояние верхней панели: она видна со всех вкладок, и по ней
        человек понимает, работает бот или нет, не переключаясь на «Обзор»."""
        alive = bool(self.bot_thread and self.bot_thread.is_alive())
        paused = control.is_paused()
        if not alive:
            text = "Остановлен"
        elif paused:
            text = "ПАУЗА — новых сделок нет, открытые ведутся"
        else:
            text = "Работает"
        self.top_status_var.set(text)
        try:
            self.btn_pause.config(
                text="▶ Продолжить" if paused else "⏸ Пауза",
                state="normal" if alive else "disabled")
            self.btn_start.config(state="disabled" if alive else "normal")
            self.btn_stop.config(state="normal" if alive else "disabled")
            self.btn_restart.config(state="normal" if alive else "disabled")
        except Exception:  # noqa: BLE001
            pass

    def save_everything(self):
        """ОДНА кнопка сохранения на всю программу.

        Владелец: «сократи кнопки сохранить, сделай одну основную внизу».
        Их было семь — по своей на каждый раздел, — и понять, какую нажимать
        и сохранено ли уже всё, было нельзя. Теперь одна внизу сохраняет
        всё сразу: брокера, параметры, профиль, контекст, источники,
        систему. Каждый раздел сохраняется своим кодом, как и раньше —
        меняется только то, откуда это запускается."""
        savers = [
            ("брокер", getattr(self, "save_broker_settings", None)),
            ("параметры", getattr(self, "save_advanced_params", None)),
            ("профиль риска", getattr(self, "save_profile_fields", None)),
            ("контекст рынка", getattr(self, "save_market_context", None)),
            ("источники", getattr(self, "save_sources", None)),
            ("система", getattr(self, "save_system_settings", None)),
        ]
        done, problems = [], []
        for name, saver in savers:
            if saver is None:
                continue
            try:
                saver(silent=True)
                done.append(name)
            except TypeError:
                # Раздел ещё не умеет тихий режим — сохраняем как есть
                try:
                    saver()
                    done.append(name)
                except Exception as e:  # noqa: BLE001
                    problems.append(f"{name}: {e}")
            except Exception as e:  # noqa: BLE001
                problems.append(f"{name}: {e}")

        try:
            settings_backup.save()
        except Exception as e:  # noqa: BLE001
            problems.append(f"копия настроек: {e}")

        if problems:
            self.save_all_status_var.set("Сохранено с ошибками")
            messagebox.showwarning(
                APP_TITLE,
                "Сохранено: " + ", ".join(done) +
                "\n\nНе сохранилось:\n- " + "\n- ".join(problems))
            return
        self.save_all_status_var.set(
            f"Сохранено ({time.strftime('%H:%M:%S')})")
        messagebox.showinfo(APP_TITLE,
                            "Все настройки сохранены: " + ", ".join(done) + ".")

    def stop_bot(self):
        self._bot_should_run = False
        if self.stop_event:
            self.stop_event.set()
        self.status_var.set("Останавливается...")
        self.stop_btn.config(state="disabled")

    # ---- периодическое обновление всех вкладок --------------------------------
    # Слова, по которым строка считается ВАЖНОЙ и красится красным.
    # Всё остальное — обычные сообщения о ходе работы, и красить их красным
    # значит обесценить сам красный цвет: у владельца на снимке красным было
    # ВСЁ, включая список отобранных пар.
    CRITICAL_WORDS = (
        "ошибк", "сбой", "не удалось", "потеряна связь", "остановлен",
        "запрещ", "отменена", "аварийн", "не запуст", "повреж", "нет доступа",
        "недостаточно средств", "риск", "просадк", "лимит",
    )

    @staticmethod
    def _warning_severity(line: str) -> str:
        """«важно» или «обычное». Решает только цвет, не содержание."""
        text = str(line or "").lower()
        return "важно" if any(w in text for w in App.CRITICAL_WORDS) else "обычное"

    def _show_warnings(self, problems) -> None:
        """Заполнить рамку «Внимание». Длинные строки подрезаются.

        ПОЧЕМУ ПОДРЕЗАЮТСЯ. В рамку однажды попал список из 497 отобранных
        пар одной строкой — и занял пол-экрана, вытеснив всё остальное.
        Место сообщения — рамка, место списка пар — вкладка «Символы»."""
        self.trade_warning_text.configure(state="normal")
        self.trade_warning_text.delete("1.0", "end")
        for line in problems or ():
            text = str(line)
            if len(text) > 300:
                text = text[:300] + "… (полностью — на вкладке «Символы» и в логе)"
            self.trade_warning_text.insert("end", "• " + text + "\n",
                                           self._warning_severity(text))
        self.trade_warning_text.configure(state="disabled")

    def _refresh_loop(self):
        try:
            snap = ds.get_snapshot()
            problems = []
            if snap:
                acc = snap.get("account", {})
                mode_txt = "LIVE" if snap.get("live_trading") else "DRY-RUN"
                self.info_var.set(
                    f"Счёт {acc.get('login', '-')}  ({acc.get('server', '-')})\n"
                    f"Баланс: {acc.get('balance', 0):.2f}   Эквити: {acc.get('equity', 0):.2f}\n"
                    f"Режим: {mode_txt} | Профиль: {snap.get('risk_profile', '-')} | "
                    f"Торговля: {snap.get('trading_mode', '-')} | Сделок сегодня: {snap.get('trades_today', 0)}"
                )
                st = snap.get("stats", {})
                self.stats_var.set(
                    f"Сделок всего: {st.get('total_trades', 0)}   Винрейт: {st.get('win_rate', 0)}%\n"
                    f"P/L за день: {st.get('day_pnl_pct', 0)}%   Просадка: {st.get('drawdown_pct', 0)}%   "
                    f"Открытых позиций: {len(snap.get('positions', []))}"
                )
                # Диагностика "сделки не открываются" — предупреждение видно, только
                # если реально есть проблема с разрешением на торговлю в MT5.
                perm = snap.get("trade_permission", {}) or {}
                problems = perm.get("problems", [])
                # Причины отказа по каждой паре — из того же снимка. Раньше их
                # можно было увидеть только на вкладке «Символы», по одной,
                # мелким текстом. Из-за этого «затишье» выглядело как поломка,
                # хотя программа знала причину и молчала о ней.
                problems = list(problems) + self._silence_reasons(snap)
            # СОБЫТИЯ ПОКАЗЫВАЮТСЯ ДАЖЕ БЕЗ СНИМКА — и это важно.
            # Снимок появляется только после того, как торговый цикл прошёл
            # первый круг, а до него идёт подготовка: подключение к терминалу
            # и отбор пар у брокера. Раньше этот блок стоял ВНУТРИ «если есть
            # снимок», поэтому во время подготовки окно не говорило ничего —
            # и занятая работой программа выглядела зависшей. Владелец так и
            # написал: «нет отклика от программы, виснет».
            events = runtime_events.describe(3)
            if events:
                problems.append("Недавние события:\n  " + events.replace("\n", "\n  "))
            if problems:
                self._show_warnings(problems)
                # Рамка появляется, ТОЛЬКО когда есть что сказать: пустая
                # рамка «Внимание» на пол-экрана пугает без причины.
                if not self.trade_warning_frame.winfo_ismapped():
                    self.trade_warning_frame.pack(fill="x", padx=12, pady=6)
            else:
                self._show_warnings([])
                if self.trade_warning_frame.winfo_ismapped():
                    self.trade_warning_frame.pack_forget()

            if self.bot_thread and self.bot_thread.is_alive():
                pause_txt = " (пауза)" if control.is_paused() else ""
                # Пока первого круга не было, «Работает» вводит в заблуждение:
                # цикл ещё готовится (подключение, отбор пар), сделок нет и не
                # будет несколько секунд. Пишем правду.
                if not snap:
                    self.status_var.set("Подготовка: подключение и отбор пар...")
                else:
                    self.status_var.set("Работает" + pause_txt)
            self._watchdog_tick()
            self._refresh_top_bar()

            self._refresh_symbols_tab()
            self._refresh_positions_tab()
            self._refresh_log_tab()
            self._redraw_equity_canvas()
            self.pause_btn.config(text="Возобновить торговлю" if control.is_paused() else "Пауза (новые сделки)")

            for title, message in control.drain_notifications():
                self._show_toast(title, message)

            self._upload_journal_if_due()
        except Exception:
            log.exception("Ошибка обновления интерфейса")
        self.root.after(3000, self._refresh_loop)

    def _upload_journal_if_due(self):
        """Плановая выгрузка журнала в облако. Сама отправка — в отдельном
        потоке: этот цикл рисует окно, и ждать в нём ответа GitHub нельзя,
        иначе программа замрёт на несколько секунд каждые N минут."""
        if not cloud_journal.enabled():
            return
        if getattr(self, "_journal_uploading", False):
            return
        if (time.time() - cloud_journal.last_upload_ts()
                < cloud_journal.upload_interval_seconds()):
            return

        self._journal_uploading = True
        snapshot = ds.get_snapshot()

        def worker():
            try:
                result = cloud_journal.upload_if_due(snapshot)
            except Exception as e:  # noqa: BLE001
                log.warning("Плановая выгрузка журнала не удалась: %s", e)
                result = None

            def finish():
                self._journal_uploading = False
                if result:
                    self._apply_journal_result(result)
            self.root.after(0, finish)

        threading.Thread(target=worker, daemon=True, name="cloud-journal-auto").start()

    # ---- уведомления --------------------------------------------------------------
    def _show_toast(self, title: str, message: str):
        try:
            if getattr(cfg, "USE_SOUND_NOTIFICATIONS", True) and sys.platform == "win32":
                import winsound
                winsound.MessageBeep()
        except Exception:
            pass
        try:
            toast = tk.Toplevel(self.root)
            toast.overrideredirect(True)
            toast.attributes("-topmost", True)
            toast.configure(bg=self.colors["tab_bg"])
            x = self.root.winfo_x() + max(self.root.winfo_width() - 320, 0)
            y = self.root.winfo_y() + 40
            toast.geometry(f"300x70+{max(x, 0)}+{max(y, 0)}")
            tk.Label(toast, text=title, bg=self.colors["tab_bg"], fg=self.colors["profit"], font=("Segoe UI", 10, "bold"),
                     anchor="w").pack(fill="x", padx=10, pady=(8, 0))
            tk.Label(toast, text=message, bg=self.colors["tab_bg"], fg=self.colors["fg"], anchor="w", wraplength=280,
                     justify="left").pack(fill="x", padx=10)
            toast.after(4500, toast.destroy)
        except Exception:
            pass

    # ---- экспорт в Excel ----------------------------------------------------------
    def export_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(APP_TITLE, "Библиотека openpyxl не установлена.\nВыполни: pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                                             initialfile="ai_scalper_report.xlsx")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws_trades = wb.active
            ws_trades.title = "Сделки"
            if os.path.exists(cfg.LOG_CSV_PATH):
                with open(cfg.LOG_CSV_PATH, encoding="utf-8") as f:
                    for row in csv.reader(f, delimiter=";"):
                        ws_trades.append(row)

            ws_stats = wb.create_sheet("Статистика")
            snap = ds.get_snapshot() or {}
            st = snap.get("stats", {})
            ws_stats.append(["Показатель", "Значение"])
            for k, v in st.items():
                ws_stats.append([k, v])
            acc = snap.get("account", {})
            ws_stats.append([])
            ws_stats.append(["Счёт", acc.get("login", "-")])
            ws_stats.append(["Сервер", acc.get("server", "-")])
            ws_stats.append(["Баланс", acc.get("balance", 0)])
            ws_stats.append(["Эквити", acc.get("equity", 0)])

            # Синхронизация с MetaTrader (п.4): реальная история сделок брокера,
            # включая открытые вручную — не только журнал этого бота выше.
            ws_mt5 = wb.create_sheet("История MT5")
            ws_mt5.append(["Тикет", "Время", "Символ", "Тип", "Лот", "Цена", "Профит", "Источник"])
            for d in snap.get("mt5_history", []):
                ws_mt5.append([
                    d.get("ticket"), d.get("time"), d.get("symbol"), d.get("type"),
                    d.get("volume"), d.get("price"), d.get("profit"),
                    "Бот" if d.get("is_bot", True) else "Ручная",
                ])
            hist_stats = snap.get("mt5_history_stats", {})
            if hist_stats:
                ws_mt5.append([])
                for k, v in hist_stats.items():
                    ws_mt5.append([k, v])

            wb.save(path)
            messagebox.showinfo(APP_TITLE, f"Сохранено: {path}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить: {e}")

    # ---- вспомогательные действия ------------------------------------------
    def open_dashboard(self):
        webbrowser.open(f"http://127.0.0.1:{cfg.DASHBOARD_PORT}")

    def open_logs(self):
        try:
            os.startfile(_app_dir)
        except Exception:
            messagebox.showinfo("Логи", f"Файл лога: {LOG_FILE}")

    def open_config(self):
        config_path = os.path.join(_app_dir, "config.py")
        try:
            os.startfile(config_path)
        except Exception:
            messagebox.showinfo("Настройки", f"Файл настроек: {config_path}")

    def _save_theme_choice(self):
        """Записать выбранную тему и сказать, что нужен перезапуск.

        Перекрашивать уже построенное окно на ходу — отдельная большая работа:
        цвета розданы десяткам виджетов при создании. Обещать мгновенную смену
        и не сделать её хуже, чем честно попросить перезапуск."""
        key = self._theme_keys.get(self.theme_var.get(), ui_theme.DEFAULT)
        try:
            _write_config_value("UI_THEME", repr(key))
            _reload_cfg()
        except Exception as e:  # noqa: BLE001
            messagebox.showwarning(APP_TITLE, f"Не удалось сохранить тему: {e}")
            return
        messagebox.showinfo(
            APP_TITLE,
            f"Тема «{self.theme_var.get()}» сохранена.\n\n"
            "Она применится при следующем запуске программы: цвета "
            "раздаются окну при его построении.")

    def _toggle_autostart(self):
        # Раньше неудача уходила в журнал, а человек видел поставленную
        # галочку и был уверен, что всё получилось. Теперь галочка снимается
        # обратно, и сказано, что именно не вышло.
        if self.autostart_var.get():
            error = _enable_autostart()
            if error:
                self.autostart_var.set(False)
                messagebox.showwarning(APP_TITLE, error)
            else:
                messagebox.showinfo(
                    APP_TITLE,
                    "Программа будет запускаться вместе с Windows.\n\n"
                    f"Запускаться будет этот файл:\n{_exe_path()}\n\n"
                    "Если перенести программу в другое место, путь "
                    "исправится сам при следующем запуске.")
        else:
            _disable_autostart()

    # ---- системный трей -----------------------------------------------------
    def _start_tray(self):
        # Фон значка в трее НЕ из темы: панель задач Windows тёмная
        # независимо от того, светлая тема в самой программе или нет.
        image = Image.new("RGB", (64, 64), "#111111")
        d = ImageDraw.Draw(image)
        d.ellipse((6, 6, 58, 58), fill=self.colors["profit"])
        d.text((20, 24), "AI", fill="white")

        menu = pystray.Menu(
            pystray.MenuItem("Показать окно", self._show_window, default=True),
            pystray.MenuItem("Старт", lambda: self.root.after(0, self.start_bot)),
            pystray.MenuItem("Стоп", lambda: self.root.after(0, self.stop_bot)),
            pystray.MenuItem("Полный выход", lambda: self.root.after(0, self._hard_quit)),
        )
        self.tray_icon = pystray.Icon(APP_TITLE, image, APP_TITLE, menu)
        threading.Thread(target=self.tray_icon.run, daemon=True).start()

    def _show_window(self, *_):
        self.root.after(0, self.root.deiconify)

    def _on_close(self):
        if TRAY_AVAILABLE:
            self.root.withdraw()
        else:
            if messagebox.askyesno(APP_TITLE, "Закрыть программу? Бот (если запущен) будет остановлен."):
                self._hard_quit()

    def full_exit(self):
        if messagebox.askyesno(APP_TITLE, "Полностью закрыть программу? Бот (если запущен) будет остановлен, "
                                            "процесс исчезнет из Диспетчера задач."):
            self._hard_quit()

    def _hard_quit(self, *_):
        try:
            if self.stop_event:
                self.stop_event.set()
        except Exception:
            pass
        try:
            # Процессы счетов запущены отдельно от торгового цикла — их надо
            # остановить явно, иначе они переживут закрытие окна
            if getattr(self, "accounts_tab", None):
                self.accounts_tab.shutdown()
        except Exception:
            pass
        try:
            # Мост слушает порт в своём потоке — без явной остановки порт
            # остаётся занятым, и следующий запуск программы не может его
            # открыть. Владелец просил, чтобы «Выход» закрывал ВСЁ, что
            # запускалось.
            bridge_host.stop()
        except Exception:
            pass
        try:
            # tgr, а не telegram_reader: модуль импортирован под коротким
            # именем (см. верх файла). Здесь стояло полное имя — получался
            # NameError, который молча съедал except ниже, и чтение Telegram
            # на выходе не останавливалось вообще.
            tgr.stop()
        except Exception:
            pass
        try:
            # Окно терминала возвращаем на экран ОБЯЗАТЕЛЬНО. Спрятанное окно
            # исчезает и с панели задач: если оставить его так, человек не
            # сможет открыть MetaTrader вообще — ни мышью, ни из меню — и
            # решит, что программа сломала ему терминал.
            mt5c.show_terminal()
        except Exception:
            pass
        try:
            if self.tray_icon:
                self.tray_icon.stop()
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass
        os._exit(0)  # гарантированное завершение процесса — ничего не остаётся в Диспетчере задач

    def run(self):
        self.root.mainloop()


# =====================================================================
# Автозапуск с Windows (реестр HKCU\...\Run — не требует прав администратора)
# =====================================================================
_RUN_KEY = r"Software\Microsoft\Windows\CurrentVersion\Run"
_APP_REG_NAME = "AIScalperPro"


def _exe_path() -> str:
    return sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)


def _is_autostart_enabled() -> bool:
    if sys.platform != "win32":
        return False
    import winreg
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _RUN_KEY) as key:
            winreg.QueryValueEx(key, _APP_REG_NAME)
            return True
    except FileNotFoundError:
        return False
    except Exception:
        return False


def _autostart_command() -> str:
    """Что именно записано в автозапуск. Пусто — записи нет."""
    if sys.platform != "win32":
        return ""
    import winreg
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _RUN_KEY) as key:
            value, _ = winreg.QueryValueEx(key, _APP_REG_NAME)
            return str(value or "")
    except Exception:      # noqa: BLE001
        return ""


def autostart_needs_repair(recorded: str, current_exe: str) -> str:
    """Указывает ли автозапуск НЕ ТУДА. Пусто — всё в порядке.

    ЗАЧЕМ. Владелец: «не работает автозапуск программы с Windows». Галочка
    стоит, запись в реестре есть — а программа не стартует. Причина почти
    всегда одна: запись сделана, когда программа лежала в другом месте
    (запускали из «Загрузок», потом поставили установщиком, потом перенесли).
    Windows честно пытается запустить файл по старому пути, его там нет, и
    ничего не происходит — молча, без единого сообщения.

    Само по себе это не чинится: Windows не сообщает о неудаче, а человек
    видит только «галочка стоит, а не работает»."""
    have = str(recorded or "").strip().strip('"')
    want = str(current_exe or "").strip().strip('"')
    if not have:
        return ""                       # записи нет — это другой случай
    if not want:
        return ""
    if os.path.normcase(os.path.abspath(have)) == os.path.normcase(os.path.abspath(want)):
        return ""
    return (f"автозапуск указывал на «{have}», а программа сейчас лежит в "
            f"«{want}» — путь исправлен")


def _enable_autostart() -> str:
    """Включить автозапуск. Возвращает текст ошибки; пусто — получилось."""
    if sys.platform != "win32":
        return "Автозапуск с Windows доступен только на Windows."
    import winreg
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _RUN_KEY, 0, winreg.KEY_SET_VALUE) as key:
            winreg.SetValueEx(key, _APP_REG_NAME, 0, winreg.REG_SZ, f'"{_exe_path()}"')
    except Exception as e:
        log.warning("Не удалось включить автозапуск: %s", e)
        return f"Не удалось включить автозапуск: {e}"
    return ""


def repair_autostart() -> str:
    """Проверить и починить путь в автозапуске. Вызывается ПРИ КАЖДОМ ЗАПУСКЕ.

    Дёшево (одно чтение реестра) и снимает целый класс жалоб «галочка стоит, а
    не запускается»."""
    if sys.platform != "win32":
        return ""
    recorded = _autostart_command()
    if not recorded:
        return ""                       # автозапуск выключен — не навязываемся
    note = autostart_needs_repair(recorded, _exe_path())
    if not note:
        return ""
    if _enable_autostart():
        return ""
    log.info("Автозапуск: %s", note)
    return note


def _disable_autostart():
    if sys.platform != "win32":
        return
    import winreg
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _RUN_KEY, 0, winreg.KEY_SET_VALUE) as key:
            winreg.DeleteValue(key, _APP_REG_NAME)
    except FileNotFoundError:
        pass
    except Exception as e:
        log.warning("Не удалось выключить автозапуск: %s", e)


def _show_login() -> bool:
    """Экран входа ПЕРЕД открытием программы — тот же логин/пароль, что и у
    веб-дашборда (cfg.DASHBOARD_LOGIN / DASHBOARD_PASSWORD, задаются в config.py
    или на вкладке дашборда). Пока не введён верный логин/пароль — окно
    программы не открывается."""
    # ЦВЕТА БЕРЁМ ИЗ ui_theme, А НЕ ИЗ self. Это обычная функция, а не метод
    # класса App — никакого self здесь нет и быть не может. Раньше тут стояло
    # self.colors[...], и включение REQUIRE_LOGIN роняло программу на старте
    # с NameError, ещё до появления окна. Ловушка была тихой ровно потому, что
    # экран входа выключен по умолчанию, — а программа при этом сама
    # предлагает его включить (см. текст про заблокированные пароли счетов).
    colors = ui_theme.from_config(cfg)

    login_root = tk.Tk()
    login_root.title(APP_TITLE)
    login_root.geometry("340x260")
    login_root.resizable(False, False)
    login_root.configure(bg=colors["bg"])
    try:
        style = ttk.Style(login_root)
        style.theme_use("clam")
        style.configure(".", background=colors["bg"], foreground=colors["fg"], fieldbackground=colors["card"])
    except tk.TclError:
        pass

    ok_holder = {"ok": False}

    ttk.Label(login_root, text=APP_TITLE, font=("Segoe UI", 14, "bold")).pack(pady=(18, 10))

    form = ttk.Frame(login_root)
    form.pack(padx=30, fill="x")
    ttk.Label(form, text="Логин:").pack(anchor="w")
    login_var = tk.StringVar()
    login_entry = ttk.Entry(form, textvariable=login_var, width=28)
    login_entry.pack(fill="x")

    ttk.Label(form, text="Пароль:").pack(anchor="w", pady=(8, 0))
    pass_var = tk.StringVar()
    pass_entry = ttk.Entry(form, textvariable=pass_var, width=28, show="*")
    pass_entry.pack(fill="x")

    remembered = _load_remembered_password()
    if remembered:
        pass_var.set(remembered)
    remember_var = tk.BooleanVar(value=bool(remembered))
    ttk.Checkbutton(login_root, text="Запомнить пароль", variable=remember_var).pack(pady=(8, 0))

    status_var = tk.StringVar(value="")
    ttk.Label(login_root, textvariable=status_var, foreground=colors["loss"]).pack(pady=(8, 0))

    def try_login(*_):
        entered_login = login_var.get()
        entered_password = pass_var.get()
        stored_hash = getattr(cfg, "DASHBOARD_PASSWORD_HASH", "")
        if stored_hash:
            salt = getattr(cfg, "SECURITY_SALT", "")
            ok_login = entered_login == str(getattr(cfg, "DASHBOARD_LOGIN", ""))
            ok_pass = secure_store.verify_password(entered_password, salt, stored_hash)
        else:
            # Легаси-формат (миграция ещё не прошла) — сравнение как раньше.
            ok_login = entered_login == str(getattr(cfg, "DASHBOARD_LOGIN", ""))
            ok_pass = entered_password == str(getattr(cfg, "DASHBOARD_PASSWORD", ""))

        if not (ok_login and ok_pass):
            status_var.set("Неверный логин или пароль.")
            pass_var.set("")
            return

        try:
            secure_store.unlock_config(cfg, entered_password)
        except ValueError as e:
            status_var.set(str(e))
            pass_var.set("")
            return

        control.set_session_password(entered_password)
        if remember_var.get():
            _save_remembered_password(entered_password)
        else:
            _clear_remembered_password()
        ok_holder["ok"] = True
        login_root.destroy()

    ttk.Button(login_root, text="Войти", command=try_login).pack(pady=14)
    login_entry.bind("<Return>", lambda e: pass_entry.focus_set())
    pass_entry.bind("<Return>", try_login)
    login_root.protocol("WM_DELETE_WINDOW", login_root.destroy)
    login_entry.focus_set()
    login_root.mainloop()
    return ok_holder["ok"]


def _skip_login_unlock() -> bool:
    """Вход без пароля: пробуем открыть секреты запомненным паролем.

    Секреты (ключи AI, пароль MT5) зашифрованы паролем входа. Если экран
    входа пропущен, расшифровать их можно только сохранённым паролем
    («Запомнить пароль» на экране входа, хранится через Windows DPAPI).
    Если его нет — программа откроется, но зашифрованные ключи останутся
    недоступны, о чём честно пишем в журнал.
    """
    remembered = _load_remembered_password()
    if not remembered:
        log.warning("Вход без пароля: сохранённого пароля нет, зашифрованные "
                    "ключи (AI, пароль MT5) останутся недоступны. Введите их "
                    "заново в настройках или включите REQUIRE_LOGIN.")
        return False
    try:
        secure_store.unlock_config(cfg, remembered)
        control.set_session_password(remembered)
        log.info("Вход без пароля: секреты открыты сохранённым паролем.")
        return True
    except Exception as e:  # noqa: BLE001
        log.warning("Вход без пароля: сохранённый пароль не подошёл (%s)", e)
        return False


def selftest() -> int:
    """Проверка «программа вообще способна запуститься». Возвращает код выхода.

    ЗАЧЕМ ЭТО ЕСТЬ. Владелец трижды получал собранную программу, которая не
    открывалась вовсе: «Can't find a usable init.tcl», «No such file:
    base_library.zip», «Failed to remove temporary directory». Все три — про
    распаковку собранного файла, и все три видны СРАЗУ при первом запуске.

    Значит их можно поймать на сборочном сервере, а не на компьютере
    владельца. Эта проверка запускается в конце сборки: она поднимает окно
    (самая хрупкая часть — рисование), тут же его закрывает и выходит. Не
    вышло — сборка считается неудачной и до человека не доезжает.

    Ничего, кроме окна, здесь не трогается: ни терминал, ни счета, ни сеть."""
    try:
        import tkinter
        root = tkinter.Tk()
        root.withdraw()          # окно не показываем, нам важно, что оно создалось
        root.update_idletasks()
        root.destroy()
    except Exception as e:       # noqa: BLE001
        print(f"SELFTEST FAILED: {type(e).__name__}: {e}")
        return 1

    # ПРОВЕРКА ЗАПУЩЕННОЙ КОПИИ. Она выполняется РАНЬШЕ окна, и сломаться в
    # ней — значит не открыться вовсе. Именно это и случилось у владельца:
    # os.kill(pid, 0) на Windows дал «[WinError 6] The handle is invalid», и
    # программа упала системным окном с трассировкой. Окно тогда собиралось
    # прекрасно, поэтому проверка «поднимается ли окно» ничего не заметила.
    # Спрашиваем про заведомо чужой номер процесса: ответ неважен, важно, что
    # ответ вообще есть.
    try:
        single_instance.process_alive(0x7FFFFFFF)
    except Exception as e:       # noqa: BLE001
        print(f"SELFTEST FAILED: проверка копии: {type(e).__name__}: {e}")
        return 1

    print("SELFTEST OK")
    return 0


def main():
    # Проверка сборки. Стоит ДО freeze_support и до всего остального: она
    # должна отвечать даже у программы, у которой всё прочее сломано.
    if "--selftest" in sys.argv:
        sys.exit(selftest())

    # ОДНА КОМАНДА ДЛЯ ВЫГРУЗКИ ИСТОРИИ. То же самое делает кнопка на вкладке
    # «Система»; ключ нужен тем, кому удобнее из командной строки.
    if "--export-history" in sys.argv:
        import history_export
        mt5c.connect()
        print(history_export.describe(history_export.export_all(
            progress=lambda t: print(t, flush=True))))
        sys.exit(0)

    # КРИТИЧНО для собранного .exe: процессы счетов используют multiprocessing,
    # а на Windows дочерний процесс запускается повторным вызовом этого же
    # exe-файла. Без freeze_support() программа вместо запуска счёта начала бы
    # бесконечно открывать саму себя, пока не кончится память.
    # Вызывать нужно ПЕРВЫМ делом, до любой другой работы.
    multiprocessing.freeze_support()

    # ВТОРАЯ КОПИЯ НЕ НУЖНА. Владелец: «при запуске программы включается две».
    # Две копии подключаются к одному терминалу, ведут одни и те же позиции и
    # обе двигают стоп-лосс, каждая считая, что она одна. Стоит ПОСЛЕ
    # freeze_support: дочерние процессы счетов до этого места не доходят, и
    # замок их не касается.
    if not single_instance.acquire():
        try:
            import tkinter.messagebox as mb
            mb.showinfo(APP_TITLE,
                        "Программа уже запущена.\n\n"
                        "Вторая копия не нужна: обе подключались бы к одному "
                        "терминалу и вели одни и те же сделки. "
                        "Найдите окно программы или значок в трее.")
        except Exception:  # noqa: BLE001
            pass
        return
    atexit.register(single_instance.release)

    # САМОЕ ПЕРВОЕ ДЕЛО: если рядом лежит скачанная новая версия — ставим её.
    # Работающий .exe заменить нельзя, Windows его держит, поэтому подмена
    # возможна только здесь, пока программа ещё не «развернулась».
    # Раньше эта функция была написана, но её никто не вызывал: скачанный файл
    # так и лежал рядом с программой, а она продолжала запускаться старой.
    try:
        swapped = updater.apply_pending_swap()
        if swapped and "обновлена" in swapped:
            # Подмена удалась — но в памяти сейчас СТАРЫЙ код: он был прочитан
            # из файла до переименования. Стартуем заново, уже из нового файла.
            # Зацикливания не будет: файл .new после подмены исчез, и следующий
            # запуск подменять уже нечего.
            log.info("%s Перезапускаюсь в новой версии.", swapped)
            updater.restart_program()
        elif swapped:
            log.warning("%s", swapped)
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось применить скачанное обновление: %s", e)

    # Автозапуск мог указывать на прежнее место программы: её переносили,
    # ставили заново, обновляли. Windows про такую неудачу не сообщает вовсе —
    # человек видит только «галочка стоит, а не запускается».
    try:
        fixed = repair_autostart()
        if fixed:
            log.info("Автозапуск починен: %s", fixed)
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось проверить автозапуск: %s", e)

    # ДО всякой работы с настройками: если рядом с программой config.py нет
    # (запустили свежескачанный .exe из «Загрузок», перенесли на другой
    # компьютер, переустановили) — возвращаем его из постоянной папки
    # пользователя. Иначе программа создала бы заводской файл, и человек
    # увидел бы сброшенные настройки: «сбиваются последние настройки».
    try:
        restored = settings_backup.restore_if_missing()
        if restored:
            log.info("Настройки восстановлены из постоянной копии: %s", restored)
            _reload_cfg()
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось восстановить настройки: %s", e)

    # Дописываем в config.py настройки, появившиеся в новой версии. Без этого
    # новые поля на вкладке «Настройка» были бы пустыми, а «Сохранить» ругался
    # бы «Некорректные значения в полях: ...». Существующие значения не
    # трогаются — только добавляются недостающие (см. config_migrate.py).
    try:
        added = config_migrate.sync()
        changed = config_migrate.apply_one_time()
        # Дневной порог убытка лежит не в config.py, а в accounts.json —
        # у каждого счёта свой, и глобальную галочку он не читает.
        accounts_note = config_migrate.clear_account_daily_loss()
        if accounts_note:
            changed = list(changed) + [accounts_note]
        if added or changed:
            _reload_cfg()
        if added:
            log.info("Добавлены новые настройки: %s", ", ".join(added))
        for note in changed:
            log.info("Настройка изменена при обновлении: %s", note)
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось дописать новые настройки в config.py: %s", e)

    # Источник новостей налаживается сам: если включён новостной режим, а
    # сервис календаря в терминале не поставлен или не собран — программа
    # ставит и собирает его при запуске, не спрашивая. Всё, что нельзя
    # сделать снаружи (первый запуск сервиса в Навигаторе), честно
    # называется в интерфейсе на вкладке «Новости».
    try:
        news_state = news_autostart.ensure_ready(force=True)
        if news_state.get("news_mode") and not news_state.get("ready"):
            log.warning("Новости: %s", news_autostart.describe(news_state))
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось проверить источник новостей: %s", e)

    # Свежую копию настроек кладём в постоянную папку — она переживёт и
    # обновление, и перенос программы в другое место.
    try:
        settings_backup.save()
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось сохранить копию настроек: %s", e)

    _migrate_legacy_secrets()
    _harden_files()
    # REQUIRE_LOGIN=False — программа открывается без экрана входа.
    # Удобно, но защита паролем при этом не работает: любой, кто получил
    # доступ к компьютеру, откроет программу и увидит счета.
    if getattr(cfg, "REQUIRE_LOGIN", True):
        if not _show_login():
            return
    else:
        _skip_login_unlock()
    app = App()
    app.run()


if __name__ == "__main__":
    main()
